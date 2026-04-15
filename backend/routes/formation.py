"""
Module Formation & Sensibilisation - ARS 1000
GreenLink Agritech

Programme annuel, sessions, PV, listes de presence, attestations.
Conforme aux clauses 7.3, 7.4, 12.2-12.10, 13.1-13.5.
"""

from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from typing import Optional, List
from pydantic import BaseModel
from datetime import datetime, timezone
import logging
import io
import uuid

from database import db
from routes.auth import get_current_user
from routes.cooperative import verify_cooperative

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/formation", tags=["Formation & Sensibilisation"])

# ============= THEMES OBLIGATOIRES ARS 1000 =============

THEMES_OBLIGATOIRES = [
    {"code": "POL_DURABILITE", "titre": "Politique de durabilite du cacao", "clause": "7.3", "description": "Sensibilisation a la politique de management de l'entite reconnue et objectifs de durabilite.", "public_cible": "Direction, Responsable SMCD, Producteurs, Travailleurs"},
    {"code": "DROITS_HOMME", "titre": "Droits de l'homme & droits de l'enfant", "clause": "12.2-3, 12.5-6", "description": "Politique en matiere de droits de l'homme. Prevention du travail des enfants et pires formes de travail.", "public_cible": "Direction, Travailleurs, Producteurs"},
    {"code": "DISCRIMINATION", "titre": "Prevention discrimination, harcelement, abus", "clause": "12.2-3", "description": "Politique de non-discrimination, prevention du harcelement et des abus au travail.", "public_cible": "Direction, Travailleurs, Producteurs"},
    {"code": "TRAVAIL_ENFANTS", "titre": "Travail des enfants & pires formes", "clause": "12.5-6", "description": "Identification, suivi et remediation du travail des enfants. Systeme SSRTE.", "public_cible": "Producteurs, Travailleurs, Communautes"},
    {"code": "EGALITE_GENRE", "titre": "Egalite hommes-femmes & autonomisation des jeunes", "clause": "12.2d, 12.4, 12.7", "description": "Promotion de l'egalite des genres, autonomisation des femmes et des jeunes dans la cacaoculture.", "public_cible": "Direction, Producteurs, Femmes, Jeunes"},
    {"code": "SST", "titre": "Sante & Securite au Travail (SST)", "clause": "12.8-12.9", "description": "Equipements de protection, prevention des risques, regimes de securite sociale.", "public_cible": "Travailleurs, Producteurs, Applicateurs"},
    {"code": "DECHETS_ECOSYSTEMES", "titre": "Gestion des dechets & protection des ecosystemes", "clause": "13.4, 13.5", "description": "Bonnes pratiques de gestion des dechets, protection de la biodiversite et des ecosystemes.", "public_cible": "Producteurs, Travailleurs"},
    {"code": "TRACABILITE_BPA", "titre": "Tracabilite & bonnes pratiques agricoles", "clause": "7.3, 8.1", "description": "Formation aux bonnes pratiques agricoles, tracabilite du cacao de la parcelle a l'export.", "public_cible": "Producteurs, Responsable Tracabilite"},
    {"code": "RESILIENCE_CLIMAT", "titre": "Diversification & resilience climatique", "clause": "11.3", "description": "Strategies de diversification, adaptation au changement climatique, agroforesterie.", "public_cible": "Producteurs, Coach formateur"},
    {"code": "LIBERTE_ASSOCIATION", "titre": "Liberte d'association & negociations collectives", "clause": "12.10", "description": "Droit a la liberte d'association et politique de negociations collectives.", "public_cible": "Direction, Travailleurs, Producteurs"},
    {"code": "PROTECTION_EAU", "titre": "Protection des plans d'eau", "clause": "13.2", "description": "Mesures de protection des plans d'eau, zones tampons, prevention de la contamination.", "public_cible": "Producteurs, Coach formateur, Applicateurs"},
    {"code": "AGROCHIMIQUES", "titre": "Gestion des produits agrochimiques", "clause": "13.3", "description": "Administration et entreposage en securite des produits agrochimiques, EPI.", "public_cible": "Producteurs, Applicateurs, Responsable SMCD"},
]


# ============= PYDANTIC MODELS =============

class ProgrammeCreate(BaseModel):
    titre: str = "Programme annuel de formation"
    campagne: str = "2025-2026"
    objectifs: str = ""

class SessionCreate(BaseModel):
    programme_id: str = ""
    theme_code: str = ""
    theme_titre: str = ""
    date_session: str = ""
    lieu: str = ""
    formateur: str = ""
    public_cible: str = ""
    contenu: str = ""
    duree_heures: float = 0
    clause_ref: str = ""
    notes: str = ""

class SessionUpdate(BaseModel):
    statut: Optional[str] = None
    date_session: Optional[str] = None
    lieu: Optional[str] = None
    formateur: Optional[str] = None
    contenu: Optional[str] = None
    duree_heures: Optional[float] = None
    notes: Optional[str] = None

class ParticipantAdd(BaseModel):
    member_id: str = ""
    nom: str = ""
    prenom: str = ""
    role: str = ""
    telephone: str = ""
    signature: bool = False

class RevueCreate(BaseModel):
    audit_session_id: str = ""
    date_revue: str = ""
    participants: str = ""
    points_examines: str = ""
    decisions: str = ""
    actions: str = ""


def get_coop_id(user):
    return str(user.get("coop_id") or user.get("_id") or user.get("id", ""))


# ============= PROGRAMMES =============

@router.post("/programmes")
async def create_programme(data: ProgrammeCreate, current_user: dict = Depends(get_current_user)):
    """Creer un programme annuel de formation avec themes obligatoires pre-remplis"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    prog_id = str(uuid.uuid4())
    themes = []
    for t in THEMES_OBLIGATOIRES:
        themes.append({
            **t,
            "statut": "planifie",
            "sessions_count": 0,
            "participants_count": 0,
        })

    doc = {
        "programme_id": prog_id,
        "coop_id": coop_id,
        "titre": data.titre,
        "campagne": data.campagne,
        "objectifs": data.objectifs,
        "themes": themes,
        "statut": "actif",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    await db.formation_programmes.insert_one(doc)
    doc.pop("_id", None)
    return {"status": "success", "programme": doc}


@router.get("/programmes")
async def list_programmes(current_user: dict = Depends(get_current_user)):
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)
    progs = await db.formation_programmes.find({"coop_id": coop_id}, {"_id": 0}).sort("created_at", -1).to_list(50)
    return {"programmes": progs}


@router.get("/programmes/{programme_id}")
async def get_programme(programme_id: str, current_user: dict = Depends(get_current_user)):
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)
    prog = await db.formation_programmes.find_one({"programme_id": programme_id, "coop_id": coop_id}, {"_id": 0})
    if not prog:
        raise HTTPException(status_code=404, detail="Programme non trouve")
    return {"programme": prog}


# ============= SESSIONS =============

@router.post("/sessions")
async def create_session(data: SessionCreate, current_user: dict = Depends(get_current_user)):
    """Creer une session de formation/sensibilisation"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    session_id = str(uuid.uuid4())
    theme_info = next((t for t in THEMES_OBLIGATOIRES if t["code"] == data.theme_code), None)

    doc = {
        "session_id": session_id,
        "coop_id": coop_id,
        "programme_id": data.programme_id,
        "theme_code": data.theme_code,
        "theme_titre": data.theme_titre or (theme_info["titre"] if theme_info else ""),
        "clause_ref": data.clause_ref or (theme_info["clause"] if theme_info else ""),
        "date_session": data.date_session,
        "lieu": data.lieu,
        "formateur": data.formateur,
        "public_cible": data.public_cible or (theme_info["public_cible"] if theme_info else ""),
        "contenu": data.contenu or (theme_info["description"] if theme_info else ""),
        "duree_heures": data.duree_heures,
        "notes": data.notes,
        "statut": "planifiee",
        "participants": [],
        "pv_genere": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    await db.formation_sessions.insert_one(doc)
    doc.pop("_id", None)

    # Update programme theme stats
    if data.programme_id:
        await db.formation_programmes.update_one(
            {"programme_id": data.programme_id, "coop_id": coop_id, "themes.code": data.theme_code},
            {"$inc": {"themes.$.sessions_count": 1}, "$set": {"themes.$.statut": "en_cours"}}
        )

    return {"status": "success", "session": doc}


@router.get("/sessions")
async def list_sessions(
    current_user: dict = Depends(get_current_user),
    programme_id: Optional[str] = None,
    statut: Optional[str] = None,
    theme_code: Optional[str] = None,
    search: Optional[str] = None,
):
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)
    query = {"coop_id": coop_id}
    if programme_id:
        query["programme_id"] = programme_id
    if statut:
        query["statut"] = statut
    if theme_code:
        query["theme_code"] = theme_code
    if search:
        query["$or"] = [
            {"theme_titre": {"$regex": search, "$options": "i"}},
            {"formateur": {"$regex": search, "$options": "i"}},
            {"lieu": {"$regex": search, "$options": "i"}},
        ]

    sessions = await db.formation_sessions.find(query, {"_id": 0}).sort("date_session", -1).to_list(200)

    total = len(sessions)
    planifiees = sum(1 for s in sessions if s.get("statut") == "planifiee")
    completees = sum(1 for s in sessions if s.get("statut") == "completee")
    en_retard = sum(1 for s in sessions if s.get("statut") == "en_retard")
    total_participants = sum(len(s.get("participants", [])) for s in sessions)

    return {
        "sessions": sessions,
        "stats": {"total": total, "planifiees": planifiees, "completees": completees, "en_retard": en_retard, "total_participants": total_participants},
    }


@router.get("/sessions/{session_id}")
async def get_session(session_id: str, current_user: dict = Depends(get_current_user)):
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)
    s = await db.formation_sessions.find_one({"session_id": session_id, "coop_id": coop_id}, {"_id": 0})
    if not s:
        raise HTTPException(status_code=404, detail="Session non trouvee")
    return {"session": s}


@router.put("/sessions/{session_id}")
async def update_session(session_id: str, update: SessionUpdate, current_user: dict = Depends(get_current_user)):
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.formation_sessions.update_one(
        {"session_id": session_id, "coop_id": coop_id}, {"$set": update_data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Session non trouvee")

    # If marked completee, update programme
    if update_data.get("statut") == "completee":
        session = await db.formation_sessions.find_one({"session_id": session_id}, {"_id": 0})
        if session and session.get("programme_id"):
            n_part = len(session.get("participants", []))
            await db.formation_programmes.update_one(
                {"programme_id": session["programme_id"], "coop_id": coop_id, "themes.code": session.get("theme_code")},
                {"$inc": {"themes.$.participants_count": n_part}, "$set": {"themes.$.statut": "complete"}}
            )

    s = await db.formation_sessions.find_one({"session_id": session_id}, {"_id": 0})
    return {"status": "success", "session": s}


# ============= PARTICIPANTS =============

@router.post("/sessions/{session_id}/participants")
async def add_participants(session_id: str, participants: List[ParticipantAdd], current_user: dict = Depends(get_current_user)):
    """Ajouter des participants a une session (liste de presence)"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    docs = []
    for p in participants:
        docs.append({
            "participant_id": str(uuid.uuid4()),
            "member_id": p.member_id,
            "nom": p.nom,
            "prenom": p.prenom,
            "role": p.role,
            "telephone": p.telephone,
            "signature": p.signature,
            "present": True,
            "added_at": datetime.now(timezone.utc).isoformat(),
        })

    result = await db.formation_sessions.update_one(
        {"session_id": session_id, "coop_id": coop_id},
        {"$push": {"participants": {"$each": docs}}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Session non trouvee")

    return {"status": "success", "added": len(docs)}


@router.get("/sessions/{session_id}/participants")
async def get_participants(session_id: str, current_user: dict = Depends(get_current_user)):
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)
    s = await db.formation_sessions.find_one({"session_id": session_id, "coop_id": coop_id}, {"_id": 0, "participants": 1, "theme_titre": 1, "date_session": 1})
    if not s:
        raise HTTPException(status_code=404, detail="Session non trouvee")
    return {"participants": s.get("participants", []), "theme_titre": s.get("theme_titre", ""), "date_session": s.get("date_session", "")}


# ============= PV & ATTESTATIONS (PDF) =============

@router.get("/sessions/{session_id}/pv/pdf")
async def generate_pv_pdf(session_id: str, current_user: dict = Depends(get_current_user)):
    """Generer le PV de la session en PDF (conforme audit)"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    s = await db.formation_sessions.find_one({"session_id": session_id, "coop_id": coop_id}, {"_id": 0})
    if not s:
        raise HTTPException(status_code=404, detail="Session non trouvee")

    # Mark PV as generated
    await db.formation_sessions.update_one({"session_id": session_id}, {"$set": {"pv_genere": True}})

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    el = []

    title_s = ParagraphStyle("T", parent=styles["Title"], fontSize=16, textColor=colors.HexColor("#1A3622"))
    h2_s = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=12, textColor=colors.HexColor("#1A3622"))

    el.append(Paragraph("PROCES-VERBAL DE FORMATION / SENSIBILISATION", title_s))
    el.append(Spacer(1, 0.5*cm))

    info = [
        ["Theme", s.get("theme_titre", "")],
        ["Clause ARS 1000", s.get("clause_ref", "")],
        ["Date", s.get("date_session", "")],
        ["Lieu", s.get("lieu", "")],
        ["Formateur", s.get("formateur", "")],
        ["Public cible", s.get("public_cible", "")],
        ["Duree", f"{s.get('duree_heures', 0)} heures"],
    ]
    t = Table(info, colWidths=[5*cm, 11*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E8F0EA")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E5E0")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
    ]))
    el.append(t)
    el.append(Spacer(1, 0.5*cm))

    if s.get("contenu"):
        el.append(Paragraph("Contenu de la formation", h2_s))
        el.append(Paragraph(s["contenu"], styles["Normal"]))
        el.append(Spacer(1, 0.5*cm))

    # Liste de presence
    el.append(Paragraph("Liste de Presence", h2_s))
    el.append(Spacer(1, 0.3*cm))
    parts = s.get("participants", [])
    if parts:
        headers = ["N", "Nom", "Prenom", "Role", "Telephone", "Signature"]
        rows = [headers]
        for i, p in enumerate(parts, 1):
            rows.append([str(i), p.get("nom", ""), p.get("prenom", ""), p.get("role", ""), p.get("telephone", ""), "Oui" if p.get("signature") else ""])

        pt = Table(rows, colWidths=[1*cm, 3.5*cm, 3.5*cm, 3*cm, 3*cm, 2*cm])
        pt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A3622")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E5E0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
        ]))
        el.append(pt)
    else:
        el.append(Paragraph("Aucun participant enregistre.", styles["Normal"]))

    el.append(Spacer(1, 1*cm))
    el.append(Paragraph(f"Total participants: {len(parts)}", styles["Normal"]))
    el.append(Spacer(1, 1*cm))
    el.append(Paragraph("Signature du Formateur: ________________     Cachet Cooperative: ________________", styles["Normal"]))

    doc.build(el)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf", headers={
        "Content-Disposition": f"attachment; filename=pv_formation_{session_id[:8]}.pdf"
    })


@router.get("/attestation/{session_id}/{participant_id}/pdf")
async def generate_attestation_pdf(session_id: str, participant_id: str, current_user: dict = Depends(get_current_user)):
    """Generer une attestation individuelle de formation"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    s = await db.formation_sessions.find_one({"session_id": session_id, "coop_id": coop_id}, {"_id": 0})
    if not s:
        raise HTTPException(status_code=404, detail="Session non trouvee")

    participant = next((p for p in s.get("participants", []) if p.get("participant_id") == participant_id), None)
    if not participant:
        raise HTTPException(status_code=404, detail="Participant non trouve")

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=3*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    el = []

    center_s = ParagraphStyle("Center", parent=styles["Normal"], alignment=1, fontSize=11)
    title_s = ParagraphStyle("Title", parent=styles["Title"], fontSize=20, textColor=colors.HexColor("#1A3622"), alignment=1)
    big_s = ParagraphStyle("Big", parent=styles["Normal"], alignment=1, fontSize=14, leading=20)

    el.append(Paragraph("ATTESTATION DE FORMATION", title_s))
    el.append(Spacer(1, 1*cm))
    el.append(Paragraph("La Cooperative certifie que :", center_s))
    el.append(Spacer(1, 0.5*cm))

    nom_complet = f"{participant.get('prenom', '')} {participant.get('nom', '')}".strip()
    el.append(Paragraph(f"<b>{nom_complet}</b>", big_s))
    el.append(Spacer(1, 0.5*cm))

    el.append(Paragraph(f"a participe a la session de formation/sensibilisation portant sur :", center_s))
    el.append(Spacer(1, 0.3*cm))
    el.append(Paragraph(f"<b>{s.get('theme_titre', '')}</b>", big_s))
    el.append(Spacer(1, 0.3*cm))
    el.append(Paragraph(f"Clause ARS 1000 : {s.get('clause_ref', '')}", center_s))
    el.append(Spacer(1, 0.5*cm))

    el.append(Paragraph(f"Date : {s.get('date_session', '')}", center_s))
    el.append(Paragraph(f"Lieu : {s.get('lieu', '')}", center_s))
    el.append(Paragraph(f"Formateur : {s.get('formateur', '')}", center_s))
    el.append(Paragraph(f"Duree : {s.get('duree_heures', 0)} heures", center_s))
    el.append(Spacer(1, 2*cm))

    el.append(Paragraph(f"Fait le {datetime.now(timezone.utc).strftime('%d/%m/%Y')}", center_s))
    el.append(Spacer(1, 1.5*cm))
    el.append(Paragraph("Signature & Cachet de la Cooperative", center_s))
    el.append(Spacer(1, 0.5*cm))
    el.append(Paragraph("________________________________", center_s))

    doc.build(el)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf", headers={
        "Content-Disposition": f"attachment; filename=attestation_{nom_complet.replace(' ', '_')}.pdf"
    })


# ============= DASHBOARD =============

@router.get("/dashboard")
async def get_formation_dashboard(current_user: dict = Depends(get_current_user)):
    """Tableau de bord formation avec indicateurs"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    sessions = await db.formation_sessions.find({"coop_id": coop_id}, {"_id": 0}).to_list(500)
    progs = await db.formation_programmes.find({"coop_id": coop_id}, {"_id": 0}).to_list(10)

    total_sessions = len(sessions)
    completees = sum(1 for s in sessions if s.get("statut") == "completee")
    planifiees = sum(1 for s in sessions if s.get("statut") == "planifiee")
    en_retard = sum(1 for s in sessions if s.get("statut") == "en_retard")
    total_participants = sum(len(s.get("participants", [])) for s in sessions)
    total_pv = sum(1 for s in sessions if s.get("pv_genere"))

    # Couverture par theme
    theme_coverage = []
    for theme in THEMES_OBLIGATOIRES:
        theme_sessions = [s for s in sessions if s.get("theme_code") == theme["code"]]
        n_sessions = len(theme_sessions)
        n_participants = sum(len(s.get("participants", [])) for s in theme_sessions)
        completed = any(s.get("statut") == "completee" for s in theme_sessions)
        theme_coverage.append({
            "code": theme["code"],
            "titre": theme["titre"],
            "clause": theme["clause"],
            "sessions": n_sessions,
            "participants": n_participants,
            "statut": "complete" if completed else "planifie" if n_sessions > 0 else "non_planifie",
        })

    themes_complets = sum(1 for t in theme_coverage if t["statut"] == "complete")
    themes_total = len(THEMES_OBLIGATOIRES)
    taux_couverture = round(themes_complets / themes_total * 100) if themes_total > 0 else 0

    # Alertes
    alertes = []
    for t in theme_coverage:
        if t["statut"] == "non_planifie":
            alertes.append({
                "type": "manquant",
                "severity": "error",
                "message": f"Formation obligatoire non planifiee: {t['titre']} (clause {t['clause']})",
            })

    # Check overdue sessions
    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    for s in sessions:
        if s.get("statut") == "planifiee" and s.get("date_session") and s["date_session"] < now_str:
            await db.formation_sessions.update_one(
                {"session_id": s["session_id"]}, {"$set": {"statut": "en_retard"}}
            )
            alertes.append({
                "type": "en_retard",
                "severity": "warning",
                "message": f"Session en retard: {s.get('theme_titre', '')} prevue le {s.get('date_session', '')}",
            })

    return {
        "has_programme": len(progs) > 0,
        "kpis": {
            "total_sessions": total_sessions,
            "completees": completees,
            "planifiees": planifiees,
            "en_retard": en_retard,
            "total_participants": total_participants,
            "total_pv": total_pv,
            "taux_couverture": taux_couverture,
            "themes_complets": themes_complets,
            "themes_total": themes_total,
        },
        "theme_coverage": theme_coverage,
        "alertes": alertes,
        "programme": progs[0] if progs else None,
    }


# ============= SUIVI MEMBRE =============

@router.get("/member/{member_name}/history")
async def get_member_history(member_name: str, current_user: dict = Depends(get_current_user)):
    """Historique des formations suivies par un producteur/travailleur"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    pipeline = [
        {"$match": {"coop_id": coop_id}},
        {"$unwind": "$participants"},
        {"$match": {"$or": [
            {"participants.nom": {"$regex": member_name, "$options": "i"}},
            {"participants.prenom": {"$regex": member_name, "$options": "i"}},
        ]}},
        {"$project": {
            "_id": 0,
            "session_id": 1,
            "theme_titre": 1,
            "clause_ref": 1,
            "date_session": 1,
            "lieu": 1,
            "formateur": 1,
            "statut": 1,
            "participant": "$participants",
        }},
        {"$sort": {"date_session": -1}},
    ]
    results = await db.formation_sessions.aggregate(pipeline).to_list(100)
    return {"formations": results, "member": member_name}


# ============= EXPORT EXCEL =============

@router.get("/export/excel")
async def export_excel(current_user: dict = Depends(get_current_user)):
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    sessions = await db.formation_sessions.find({"coop_id": coop_id}, {"_id": 0}).sort("date_session", -1).to_list(500)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    header_fill = PatternFill(start_color="1A3622", end_color="1A3622", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=9)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Sessions sheet
    ws = wb.active
    ws.title = "Sessions de formation"
    headers = ["Theme", "Clause ARS", "Date", "Lieu", "Formateur", "Public cible", "Duree (h)", "Participants", "Statut", "PV"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for row_idx, s in enumerate(sessions, 2):
        vals = [s.get("theme_titre", ""), s.get("clause_ref", ""), s.get("date_session", ""), s.get("lieu", ""), s.get("formateur", ""), s.get("public_cible", ""), s.get("duree_heures", 0), len(s.get("participants", [])), s.get("statut", ""), "Oui" if s.get("pv_genere") else "Non"]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.border = thin_border

    for i in range(1, 11):
        ws.column_dimensions[chr(64 + i)].width = 18

    # Participants sheet
    ws2 = wb.create_sheet("Listes de presence")
    p_headers = ["Session", "Date", "Nom", "Prenom", "Role", "Telephone", "Signature"]
    for col, h in enumerate(p_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    row = 2
    for s in sessions:
        for p in s.get("participants", []):
            ws2.cell(row=row, column=1, value=s.get("theme_titre", "")).border = thin_border
            ws2.cell(row=row, column=2, value=s.get("date_session", "")).border = thin_border
            ws2.cell(row=row, column=3, value=p.get("nom", "")).border = thin_border
            ws2.cell(row=row, column=4, value=p.get("prenom", "")).border = thin_border
            ws2.cell(row=row, column=5, value=p.get("role", "")).border = thin_border
            ws2.cell(row=row, column=6, value=p.get("telephone", "")).border = thin_border
            ws2.cell(row=row, column=7, value="Oui" if p.get("signature") else "Non").border = thin_border
            row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={
        "Content-Disposition": f"attachment; filename=formations_ars1000.xlsx"
    })


# ============= THEMES LIST (for frontend) =============

@router.get("/themes")
async def get_themes():
    """Liste des themes obligatoires ARS 1000"""
    return {"themes": THEMES_OBLIGATOIRES}
