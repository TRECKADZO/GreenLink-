"""
Module Gestion des Membres & Enregistrement ARS 1000
Clauses 4.2.2, 4.2.3, 4.3

Procedure d'adhesion digitale, bulletin/contrat, perimetre SM.
S'appuie sur la collection coop_members existante.
"""

from fastapi import APIRouter, HTTPException, Depends, Query, Request
from fastapi.responses import StreamingResponse
from typing import Optional, List
from pydantic import BaseModel
from datetime import datetime, timezone
from bson import ObjectId
import logging
import io
import uuid

from database import db
from routes.auth import get_current_user
from routes.cooperative import verify_cooperative

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/membres", tags=["Membres & Enregistrement ARS 1000"])

# ============= CHAMPS REGISTRE ARS 1000 =============
# Structure conforme au registre officiel ARS 1000 YAKRO

SECTIONS_REGISTRE = {
    "identification": {
        "titre": "IDENTIFICATION DU PRODUCTEUR",
        "champs": [
            {"field": "section", "label": "Section", "required": False},
            {"field": "nom", "label": "Nom", "required": True},
            {"field": "prenom", "label": "Prenom", "required": True},
            {"field": "numero_enregistrement", "label": "N d'enregistrement", "required": False},
            {"field": "cni_number", "label": "N CNI", "required": False},
            {"field": "date_naissance", "label": "Date de naissance", "required": False},
            {"field": "sexe", "label": "Sexe", "required": True},
            {"field": "contact", "label": "Contact", "required": True},
        ],
    },
    "cacaoyere": {
        "titre": "INFORMATIONS SUR LA CACAOYERE",
        "champs": [
            {"field": "nombre_champs", "label": "Nombre de champs", "required": False},
            {"field": "code_cacaoyere", "label": "Code de la cacaoyere", "required": False},
            {"field": "date_creation_cacaoyere", "label": "Date de creation", "required": False},
            {"field": "date_enregistrement", "label": "Date d'enregistrement", "required": False},
            {"field": "superficie_ha", "label": "Superficie (ha)", "required": False},
            {"field": "culture", "label": "Culture", "required": False},
            {"field": "densite_pieds", "label": "Densite (pieds)", "required": False},
            {"field": "polygone_disponible", "label": "Polygone disponible (oui/non)", "required": False},
            {"field": "gps_latitude", "label": "GPS Latitude", "required": False},
            {"field": "gps_longitude", "label": "GPS Longitude", "required": False},
            {"field": "autres_cultures", "label": "Autres cultures presentes", "required": False},
            {"field": "date_audit_interne", "label": "Date d'audit interne", "required": False},
        ],
    },
    "production": {
        "titre": "INFORMATIONS DE PRODUCTION",
        "champs": [
            {"field": "recolte_precedente_kg", "label": "Recolte annee precedente (Kg)", "required": False},
            {"field": "volume_vendu_precedent_kg", "label": "Volume vendu campagne precedente (Kg)", "required": False},
            {"field": "estimation_rendement_kg_ha", "label": "Estimation rendement (Kg/ha)", "required": False},
            {"field": "volume_certifier_kg", "label": "Volume a certifier (Kg)", "required": False},
        ],
    },
    "travailleurs": {
        "titre": "TRAVAILLEURS AGRICOLES PERMANENTS",
        "champs": [
            {"field": "nb_travailleurs", "label": "Nombre", "required": False},
            {"field": "travailleurs_liste", "label": "Liste des travailleurs", "required": False},
        ],
    },
    "menage": {
        "titre": "COMPOSITION DU MENAGE",
        "champs": [
            {"field": "membres_menage", "label": "Membres du menage", "required": False},
        ],
    },
}


ETAPES_ADHESION = [
    {"etape": 1, "titre": "Identification du Producteur", "description": "Informations personnelles du producteur conformes a la norme ARS 1000."},
    {"etape": 2, "titre": "Informations Cacaoyere & Production", "description": "Donnees sur les parcelles, la production et les objectifs de certification."},
    {"etape": 3, "titre": "Travailleurs & Menage", "description": "Travailleurs agricoles permanents et composition du menage."},
    {"etape": 4, "titre": "Bulletin d'adhesion & Validation", "description": "Signature du bulletin et validation par le Responsable SMCD."},
]


def get_coop_id(user):
    return str(user.get("coop_id") or user.get("_id") or user.get("id", ""))


# ============= MODELS =============

class TravailleurItem(BaseModel):
    nom: str = ""
    prenom: str = ""
    sexe: str = ""
    date_naissance: str = ""

class MembreMenageItem(BaseModel):
    nom: str = ""
    prenom: str = ""
    sexe: str = ""
    date_naissance: str = ""
    qualite_filiation: str = ""  # Conjointe, enfant, neveu, niece, frere, soeur
    frequentation_ecole: str = ""  # oui/non
    raison_non_scolarisation: str = ""
    nom_ecole: str = ""
    classe: str = ""

class AdhesionCreate(BaseModel):
    # === IDENTIFICATION DU PRODUCTEUR ===
    section: str = ""
    nom: str = ""
    prenom: str = ""
    numero_enregistrement: str = ""
    cni_number: str = ""
    code_national_ccc: str = ""  # Code officiel du Conseil Cafe-Cacao (CI)
    date_naissance: str = ""
    sexe: str = ""
    contact: str = ""
    localite: str = ""
    campement: str = ""
    loc_region: str = ""
    loc_departement: str = ""
    loc_sous_prefecture: str = ""
    # === INFORMATIONS SUR LA CACAOYERE ===
    nombre_champs: int = 0
    code_cacaoyere: str = ""
    date_creation_cacaoyere: str = ""
    date_enregistrement: str = ""
    superficie_ha: float = 0
    culture: str = "Cacao"
    densite_pieds: int = 0
    polygone_disponible: str = "non"
    gps_latitude: str = ""
    gps_longitude: str = ""
    autres_cultures: str = ""
    date_audit_interne: str = ""
    # === INFORMATIONS DE PRODUCTION ===
    recolte_precedente_kg: float = 0
    volume_vendu_precedent_kg: float = 0
    estimation_rendement_kg_ha: float = 0
    volume_certifier_kg: float = 0
    # === TRAVAILLEURS AGRICOLES PERMANENTS ===
    nb_travailleurs: int = 0
    travailleurs_liste: List[TravailleurItem] = []
    # === COMPOSITION DU MENAGE ===
    membres_menage: List[MembreMenageItem] = []
    # === BULLETIN D'ADHESION ===
    signature_producteur: bool = False
    temoin_1_nom: str = ""
    temoin_1_signature: bool = False
    temoin_2_nom: str = ""
    temoin_2_signature: bool = False
    notes: str = ""
    # Legacy compat
    full_name: str = ""
    phone_number: str = ""
    village: str = ""
    department: str = ""
    zone: str = ""
    nombre_parcelles: int = 0
    hectares_approx: float = 0
    gps_parcelle: str = ""
    nombre_travailleurs: int = 0
    statut_producteur: str = "actif"
    sensibilisation_faite: bool = False
    sensibilisation_date: str = ""
    sensibilisation_accuse: bool = False

class MemberUpdate(BaseModel):
    section: Optional[str] = None
    nom: Optional[str] = None
    prenom: Optional[str] = None
    numero_enregistrement: Optional[str] = None
    cni_number: Optional[str] = None
    code_national_ccc: Optional[str] = None
    date_naissance: Optional[str] = None
    sexe: Optional[str] = None
    contact: Optional[str] = None
    localite: Optional[str] = None
    campement: Optional[str] = None
    nombre_champs: Optional[int] = None
    code_cacaoyere: Optional[str] = None
    date_creation_cacaoyere: Optional[str] = None
    date_enregistrement: Optional[str] = None
    superficie_ha: Optional[float] = None
    culture: Optional[str] = None
    densite_pieds: Optional[int] = None
    polygone_disponible: Optional[str] = None
    gps_latitude: Optional[str] = None
    gps_longitude: Optional[str] = None
    autres_cultures: Optional[str] = None
    date_audit_interne: Optional[str] = None
    recolte_precedente_kg: Optional[float] = None
    volume_vendu_precedent_kg: Optional[float] = None
    estimation_rendement_kg_ha: Optional[float] = None
    volume_certifier_kg: Optional[float] = None
    nb_travailleurs: Optional[int] = None
    travailleurs_liste: Optional[List[TravailleurItem]] = None
    membres_menage: Optional[List[MembreMenageItem]] = None
    statut_producteur: Optional[str] = None
    notes: Optional[str] = None
    full_name: Optional[str] = None
    phone_number: Optional[str] = None
    village: Optional[str] = None
    department: Optional[str] = None
    zone: Optional[str] = None
    nombre_parcelles: Optional[int] = None
    hectares_approx: Optional[float] = None

class PerimetreUpdate(BaseModel):
    description: str = ""
    producteurs_inclus: int = 0
    parcelles_incluses: int = 0
    exclusions: str = ""
    date_validation: str = ""
    valide_par: str = ""


# ============= PROCEDURE D'ADHESION =============

@router.post("/adhesion")
async def create_adhesion(data: AdhesionCreate, current_user: dict = Depends(get_current_user)):
    """Creer une demande d'adhesion complete - Registre ARS 1000"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    adhesion_id = str(uuid.uuid4())
    code_membre = f"MBR-{uuid.uuid4().hex[:6].upper()}"

    # Build full_name from nom+prenom or legacy
    full_name = f"{data.nom} {data.prenom}".strip() if (data.nom or data.prenom) else data.full_name
    contact = data.contact or data.phone_number
    village = data.localite or data.village
    superficie = data.superficie_ha or data.hectares_approx

    # Determine etape
    etape = 1
    if full_name and contact:
        etape = 2
    if data.nombre_champs > 0 or data.superficie_ha > 0:
        etape = 3
    if data.signature_producteur:
        etape = 4

    doc = {
        "adhesion_id": adhesion_id,
        "code_membre": code_membre,
        "coop_id": coop_id,
        # IDENTIFICATION DU PRODUCTEUR
        "section": data.section,
        "nom": data.nom,
        "prenom": data.prenom,
        "full_name": full_name,
        "numero_enregistrement": data.numero_enregistrement,
        "cni_number": data.cni_number,
        "code_national_ccc": data.code_national_ccc,
        "date_naissance": data.date_naissance,
        "sexe": data.sexe,
        "contact": contact,
        "phone_number": contact,
        "localite": village,
        "village": village,
        "campement": data.campement,
        "loc_region": data.loc_region,
        "loc_departement": data.loc_departement,
        "loc_sous_prefecture": data.loc_sous_prefecture,
        "department": data.department or data.loc_departement,
        "zone": data.zone or data.section,
        # INFORMATIONS SUR LA CACAOYERE
        "nombre_champs": data.nombre_champs,
        "nombre_parcelles": data.nombre_parcelles or data.nombre_champs,
        "code_cacaoyere": data.code_cacaoyere,
        "date_creation_cacaoyere": data.date_creation_cacaoyere,
        "date_enregistrement": data.date_enregistrement or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "superficie_ha": superficie,
        "hectares_approx": superficie,
        "culture": data.culture,
        "densite_pieds": data.densite_pieds,
        "polygone_disponible": data.polygone_disponible,
        "gps_latitude": data.gps_latitude,
        "gps_longitude": data.gps_longitude,
        "gps_parcelle": f"{data.gps_latitude},{data.gps_longitude}" if data.gps_latitude else data.gps_parcelle,
        "autres_cultures": data.autres_cultures,
        "date_audit_interne": data.date_audit_interne,
        # INFORMATIONS DE PRODUCTION
        "recolte_precedente_kg": data.recolte_precedente_kg,
        "volume_vendu_precedent_kg": data.volume_vendu_precedent_kg,
        "estimation_rendement_kg_ha": data.estimation_rendement_kg_ha,
        "volume_certifier_kg": data.volume_certifier_kg,
        # TRAVAILLEURS AGRICOLES PERMANENTS
        "nb_travailleurs": data.nb_travailleurs or data.nombre_travailleurs,
        "nombre_travailleurs": data.nb_travailleurs or data.nombre_travailleurs,
        "travailleurs_liste": [t.model_dump() for t in data.travailleurs_liste],
        # COMPOSITION DU MENAGE
        "membres_menage": [m.model_dump() for m in data.membres_menage],
        # BULLETIN
        "sensibilisation_faite": data.sensibilisation_faite,
        "sensibilisation_date": data.sensibilisation_date,
        "sensibilisation_accuse": data.sensibilisation_accuse,
        "signature_producteur": data.signature_producteur,
        "temoin_1_nom": data.temoin_1_nom,
        "temoin_1_signature": data.temoin_1_signature,
        "temoin_2_nom": data.temoin_2_nom,
        "temoin_2_signature": data.temoin_2_signature,
        "statut_producteur": data.statut_producteur,
        "date_adhesion": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        # META
        "etape_courante": etape,
        "statut": "en_cours" if etape < 4 else "en_attente_validation",
        "validation": {"validee": False, "validee_par": "", "date_validation": ""},
        "notes": data.notes,
        "historique": [],
        "created_by": current_user.get("full_name", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    await db.membres_adhesions.insert_one(doc)
    doc.pop("_id", None)

    return {"status": "success", "adhesion": doc, "code_membre": code_membre}


@router.get("/adhesions")
async def list_adhesions(
    current_user: dict = Depends(get_current_user),
    statut: Optional[str] = None,
    search: Optional[str] = None,
):
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    query = {"coop_id": coop_id}
    if statut:
        query["statut"] = statut
    if search:
        query["$or"] = [
            {"full_name": {"$regex": search, "$options": "i"}},
            {"code_membre": {"$regex": search, "$options": "i"}},
            {"village": {"$regex": search, "$options": "i"}},
        ]

    adhesions = await db.membres_adhesions.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)

    total = len(adhesions)
    en_cours = sum(1 for a in adhesions if a.get("statut") == "en_cours")
    en_attente = sum(1 for a in adhesions if a.get("statut") == "en_attente_validation")
    valides = sum(1 for a in adhesions if a.get("statut") == "valide")
    retrait = sum(1 for a in adhesions if a.get("statut") == "retrait")

    return {
        "adhesions": adhesions,
        "stats": {"total": total, "en_cours": en_cours, "en_attente_validation": en_attente, "valides": valides, "retrait": retrait},
    }


@router.put("/adhesion/{adhesion_id}/valider")
async def valider_adhesion(adhesion_id: str, current_user: dict = Depends(get_current_user)):
    """Valider une adhesion (etape 4 - par Resp SMCD ou CA/CG)"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    result = await db.membres_adhesions.update_one(
        {"adhesion_id": adhesion_id, "coop_id": coop_id},
        {"$set": {
            "statut": "valide",
            "etape_courante": 4,
            "validation.validee": True,
            "validation.validee_par": current_user.get("full_name", ""),
            "validation.date_validation": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Adhesion non trouvee")

    return {"status": "success", "message": "Adhesion validee"}


@router.put("/adhesion/{adhesion_id}/retrait")
async def retrait_membre(adhesion_id: str, current_user: dict = Depends(get_current_user)):
    """Marquer un retrait de membre"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    result = await db.membres_adhesions.update_one(
        {"adhesion_id": adhesion_id, "coop_id": coop_id},
        {"$set": {"statut": "retrait", "date_retrait": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Adhesion non trouvee")

    return {"status": "success"}


# ============= BULLETIN PDF =============

@router.get("/adhesion/{adhesion_id}/bulletin/pdf")
async def generate_bulletin_pdf(adhesion_id: str, current_user: dict = Depends(get_current_user)):
    """Generer le bulletin/contrat d'adhesion en PDF"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    a = await db.membres_adhesions.find_one({"adhesion_id": adhesion_id, "coop_id": coop_id}, {"_id": 0})
    if not a:
        raise HTTPException(status_code=404, detail="Adhesion non trouvee")

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    el = []

    title_s = ParagraphStyle("T", parent=styles["Title"], fontSize=16, textColor=colors.HexColor("#1A3622"), alignment=1)
    center_s = ParagraphStyle("C", parent=styles["Normal"], alignment=1, fontSize=10)
    normal_s = ParagraphStyle("N", parent=styles["Normal"], fontSize=9, leading=12)

    el.append(Paragraph("BULLETIN D'ADHESION", title_s))
    el.append(Paragraph("Contrat de membre - Norme ARS 1000", center_s))
    el.append(Spacer(1, 0.8*cm))

    info = [
        ["Code membre", a.get("code_membre", "")],
        ["Nom complet (a)", a.get("full_name", "")],
        ["Date de naissance (b)", a.get("date_naissance", "")],
        ["Sexe (c)", a.get("sexe", "")],
        ["N CNI (d)", a.get("cni_number", "")],
        ["Telephone (e)", a.get("phone_number", "")],
        ["Village/Section (f)", a.get("localite", "") or a.get("village", "")],
        ["Campement", a.get("campement", "")],
        ["Departement (g)", a.get("department", "")],
        ["Zone (h)", a.get("zone", "")],
        ["Parcelles (i)", str(a.get("nombre_parcelles", 0))],
        ["Superficie ha (j)", str(a.get("hectares_approx", 0))],
        ["GPS (k)", a.get("gps_parcelle", "")],
        ["Travailleurs (l)", str(a.get("nombre_travailleurs", 0))],
        ["Date adhesion (m)", a.get("date_adhesion", "")],
    ]
    t = Table(info, colWidths=[5*cm, 11*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E8F0EA")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E5E0")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
    ]))
    el.append(t)
    el.append(Spacer(1, 0.8*cm))

    el.append(Paragraph("En signant ce bulletin, je m'engage a respecter les statuts, le reglement interieur et la charte de la cooperative, ainsi que les exigences de la norme ARS 1000 pour la durabilite du cacao.", normal_s))
    el.append(Spacer(1, 1*cm))

    sigs = [
        ["Signature du producteur", "Oui" if a.get("signature_producteur") else "______"],
        ["Temoin 1: " + a.get("temoin_1_nom", "______"), "Oui" if a.get("temoin_1_signature") else "______"],
        ["Temoin 2: " + a.get("temoin_2_nom", "______"), "Oui" if a.get("temoin_2_signature") else "______"],
    ]
    st = Table(sigs, colWidths=[8*cm, 8*cm])
    st.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E5E0")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (-1, -1), 12),
    ]))
    el.append(st)

    el.append(Spacer(1, 0.5*cm))
    v = a.get("validation", {})
    if v.get("validee"):
        el.append(Paragraph(f"Valide par: {v.get('validee_par', '')} le {v.get('date_validation', '')}", normal_s))

    el.append(Spacer(1, 0.5*cm))
    el.append(Paragraph("Cachet de la Cooperative: ________________", normal_s))

    doc.build(el)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf", headers={
        "Content-Disposition": f"attachment; filename=bulletin_{a.get('code_membre', '')}.pdf"
    })


# ============= BASE DE DONNEES MEMBRES =============

@router.get("/registre")
async def get_registre_membres(
    current_user: dict = Depends(get_current_user),
    statut: Optional[str] = None,
    village: Optional[str] = None,
    sexe: Optional[str] = None,
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
):
    """Registre complet des membres (champs 4.2.3.2 a-n)"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    query = {"coop_id": coop_id}
    if statut:
        query["statut"] = statut
    if village:
        query["village"] = {"$regex": village, "$options": "i"}
    if sexe:
        query["sexe"] = sexe
    if search:
        query["$or"] = [
            {"full_name": {"$regex": search, "$options": "i"}},
            {"code_membre": {"$regex": search, "$options": "i"}},
            {"phone_number": {"$regex": search, "$options": "i"}},
            {"cni_number": {"$regex": search, "$options": "i"}},
            {"code_national_ccc": {"$regex": search, "$options": "i"}},
        ]

    total = await db.membres_adhesions.count_documents(query)
    membres = await db.membres_adhesions.find(query, {"_id": 0}).sort("full_name", 1).skip(skip).limit(limit).to_list(limit)

    return {"membres": membres, "total": total, "skip": skip, "limit": limit}


@router.put("/registre/{adhesion_id}")
async def update_membre(adhesion_id: str, update: MemberUpdate, current_user: dict = Depends(get_current_user)):
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()

    history = {"date": datetime.now(timezone.utc).isoformat(), "user": current_user.get("full_name", ""), "fields": list(update_data.keys())}

    result = await db.membres_adhesions.update_one(
        {"adhesion_id": adhesion_id, "coop_id": coop_id},
        {"$set": update_data, "$push": {"historique": history}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Membre non trouve")

    m = await db.membres_adhesions.find_one({"adhesion_id": adhesion_id}, {"_id": 0})
    return {"status": "success", "membre": m}


# ============= PERIMETRE SM (clause 4.3) =============

@router.post("/perimetre")
async def save_perimetre(data: PerimetreUpdate, current_user: dict = Depends(get_current_user)):
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    # Count members
    total_membres = await db.membres_adhesions.count_documents({"coop_id": coop_id, "statut": "valide"})
    total_hectares_pipeline = [
        {"$match": {"coop_id": coop_id, "statut": "valide"}},
        {"$group": {"_id": None, "total": {"$sum": "$hectares_approx"}}},
    ]
    ha_result = await db.membres_adhesions.aggregate(total_hectares_pipeline).to_list(1)
    total_ha = ha_result[0]["total"] if ha_result else 0

    doc = {
        "perimetre_id": str(uuid.uuid4()),
        "coop_id": coop_id,
        "description": data.description,
        "producteurs_inclus": data.producteurs_inclus or total_membres,
        "parcelles_incluses": data.parcelles_incluses,
        "superficie_totale_ha": round(total_ha, 2),
        "exclusions": data.exclusions,
        "date_validation": data.date_validation,
        "valide_par": data.valide_par,
        "auto_stats": {"total_membres_valides": total_membres, "total_hectares": round(total_ha, 2)},
        "created_by": current_user.get("full_name", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }

    await db.membres_perimetres.insert_one(doc)
    doc.pop("_id", None)
    return {"status": "success", "perimetre": doc}


@router.get("/perimetre")
async def get_perimetres(current_user: dict = Depends(get_current_user)):
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)
    perimetres = await db.membres_perimetres.find({"coop_id": coop_id}, {"_id": 0}).sort("created_at", -1).to_list(50)
    return {"perimetres": perimetres}


# ============= DASHBOARD =============

@router.get("/dashboard")
async def get_membres_dashboard(current_user: dict = Depends(get_current_user)):
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    total = await db.membres_adhesions.count_documents({"coop_id": coop_id})
    valides = await db.membres_adhesions.count_documents({"coop_id": coop_id, "statut": "valide"})
    en_cours = await db.membres_adhesions.count_documents({"coop_id": coop_id, "statut": "en_cours"})
    en_attente = await db.membres_adhesions.count_documents({"coop_id": coop_id, "statut": "en_attente_validation"})
    retrait = await db.membres_adhesions.count_documents({"coop_id": coop_id, "statut": "retrait"})

    # By sex
    sexe_pipeline = [
        {"$match": {"coop_id": coop_id, "statut": {"$ne": "retrait"}}},
        {"$group": {"_id": "$sexe", "count": {"$sum": 1}}},
    ]
    sexe_result = await db.membres_adhesions.aggregate(sexe_pipeline).to_list(10)
    par_sexe = {r["_id"]: r["count"] for r in sexe_result if r["_id"]}

    # By village
    village_pipeline = [
        {"$match": {"coop_id": coop_id, "statut": {"$ne": "retrait"}}},
        {"$group": {"_id": "$village", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 10},
    ]
    village_result = await db.membres_adhesions.aggregate(village_pipeline).to_list(10)
    par_village = [{"village": r["_id"] or "Non defini", "count": r["count"]} for r in village_result]

    # Total hectares
    ha_pipeline = [
        {"$match": {"coop_id": coop_id, "statut": "valide"}},
        {"$group": {"_id": None, "total": {"$sum": "$hectares_approx"}}},
    ]
    ha_result = await db.membres_adhesions.aggregate(ha_pipeline).to_list(1)
    total_ha = round(ha_result[0]["total"], 1) if ha_result else 0

    # Perimetre
    perimetre = await db.membres_perimetres.find_one({"coop_id": coop_id}, {"_id": 0}, sort=[("created_at", -1)])

    return {
        "kpis": {
            "total": total,
            "valides": valides,
            "en_cours": en_cours,
            "en_attente_validation": en_attente,
            "retrait": retrait,
            "hommes": par_sexe.get("M", 0),
            "femmes": par_sexe.get("F", 0),
            "total_hectares": total_ha,
        },
        "par_village": par_village,
        "perimetre": perimetre,
    }


# ============= EXPORT EXCEL =============

@router.get("/export/excel")
async def export_registre_excel(current_user: dict = Depends(get_current_user)):
    """
    Export du Registre ARS 1000 au format Excel conforme au template officiel
    (Registre des membres ARS 1000 - YAKRO).
    Sections : Identification Producteur / Société Coopérative / Production 2024-2025 /
               Composition du Ménage / Informations complémentaires.
    """
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    membres = await db.membres_adhesions.find({"coop_id": coop_id}, {"_id": 0}).sort("full_name", 1).to_list(5000)
    coop_info = await db.coop_info_ars.find_one({"coop_id": coop_id}, {"_id": 0}) or {}
    coop_name = current_user.get("coop_name") or coop_info.get("nom_cooperative") or current_user.get("full_name", "")
    campagne = coop_info.get("campagne", "2024/2025")
    sigle = coop_info.get("sigle", "")
    siege = coop_info.get("siege", current_user.get("headquarters_region", ""))
    niveau_cert = coop_info.get("niveau_certification", "Bronze")
    nb_magasins = coop_info.get("nb_magasins_stockage", 0)
    nb_cacaoyeres = coop_info.get("nb_cacaoyeres", 0)

    nb_producteurs = len(membres)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.cell.cell import MergedCell

    wb = Workbook()
    ws = wb.active
    ws.title = f"Registre {campagne.replace('/', '-')}"
    hf = PatternFill(start_color="1A3622", end_color="1A3622", fill_type="solid")
    sf = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True, size=10)
    sfont = Font(color="000000", bold=True, size=9)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # ----- Layout columns (match YAKRO template exactly) -----
    # Section A: Identification du producteur (7 col)
    # Section B: Informations sur la societe cooperative (5 col)
    # Section C: Informations de production (4 col)
    # Section D: Composition du menage (6 col)
    # Section E: Informations complementaires (11 col)
    # Row 1 = section headers (merged), Row 2 = column titles

    col_groups = [
        ("IDENTIFICATION DU PRODUCTEUR", [
            "N° CNI", "Code National CCC", "Nom Prénom", "Date de naissance", "Sexe",
            "N° Section", "Nombre de sections (site de production)", "Nombre de champs",
        ]),
        ("INFORMATIONS SUR LA SOCIETE COOPERATIVE", [
            "Nombre de producteurs membre", "Nom de la société coopérative",
            "Code de la cacaoyère", "Localité", "Date de création",
        ]),
        ("INFORMATIONS DE PRODUCTION 2024/2025", [
            "Autres cultures présentes", "Culture", "Superficie (ha)",
            "Volume à certifier pour la campagne en cours (Kg)",
        ]),
        ("COMPOSITION DU MENAGE", [
            "Noms Prénoms (membres du menage)",
            "Travailleurs agricoles permanents",
            "Pour enfants en âge de scolarisation",
            "Fréquentation d'une école? (oui/non)",
            "Nom de l'école", "Classe",
        ]),
        ("INFORMATIONS COMPLEMENTAIRES", [
            "Nombre de cacaoyère", "Niveau de certification demandé",
            "Nombre de magasin de stockage de fèves", "Sigle",
            "N° d'enrégistrement", "Date d'enrégistrement", "Date d'audit interne",
            "GPS Latitude", "GPS Longitude", "Siège", "Campagne",
        ]),
    ]

    # Title row 0
    total_cols = sum(len(cols) for _, cols in col_groups)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"REGISTRE DES MEMBRES ARS 1000 — {coop_name} — Campagne {campagne}"
    title_cell.fill = hf
    title_cell.font = Font(color="FFFFFF", bold=True, size=12)
    title_cell.alignment = center
    ws.row_dimensions[1].height = 24

    # Section headers row 2
    col_idx = 1
    for section_title, cols in col_groups:
        start = col_idx
        end = col_idx + len(cols) - 1
        ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
        cell = ws.cell(row=2, column=start)
        cell.value = section_title
        cell.fill = hf
        cell.font = hfont
        cell.alignment = center
        cell.border = tb
        col_idx = end + 1

    # Column titles row 3
    flat_cols = []
    for _, cols in col_groups:
        flat_cols.extend(cols)
    for i, header in enumerate(flat_cols, 1):
        cell = ws.cell(row=3, column=i, value=header)
        cell.fill = sf
        cell.font = sfont
        cell.border = tb
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[3].height = 40

    def _compact(items, key):
        return ", ".join(str((i or {}).get(key, "")).strip() for i in (items or []) if isinstance(i, dict) and (i or {}).get(key))

    # Data rows (start at row 4)
    for row_idx, m in enumerate(membres, 4):
        menage = m.get("membres_menage", []) or []
        travailleurs = m.get("travailleurs_liste", []) or []
        # Culture principale + autres
        autres_cultures_raw = m.get("autres_cultures") or []
        if isinstance(autres_cultures_raw, list):
            autres_str = ", ".join(str(ac) if not isinstance(ac, dict) else ac.get("nom", ac.get("libelle", "")) for ac in autres_cultures_raw)
        else:
            autres_str = str(autres_cultures_raw)

        noms_menage = _compact(menage, "full_name") or _compact(menage, "nom")
        ecole_list = _compact(menage, "frequentation_ecole")
        nom_ecole_list = _compact(menage, "nom_ecole")
        classe_list = _compact(menage, "classe")
        nb_travailleurs = m.get("nb_travailleurs") or m.get("nombre_travailleurs") or len(travailleurs) or ""

        vals = [
            # IDENTIFICATION DU PRODUCTEUR
            m.get("cni_number", ""),
            m.get("code_national_ccc", ""),
            m.get("full_name", "") or f"{m.get('nom','')} {m.get('prenom','')}".strip(),
            m.get("date_naissance", ""),
            m.get("sexe", ""),
            m.get("section", "") or m.get("zone", ""),
            coop_info.get("nb_sections", "") or "",
            m.get("nombre_champs", "") or m.get("nombre_parcelles", ""),
            # INFOS SOCIETE COOPERATIVE
            nb_producteurs,
            coop_name,
            m.get("code_cacaoyere", ""),
            m.get("localite", "") or m.get("village", ""),
            m.get("date_creation_cacaoyere", ""),
            # INFOS DE PRODUCTION
            autres_str,
            m.get("culture", ""),
            m.get("superficie_ha", "") or m.get("hectares_approx", ""),
            m.get("volume_certifier_kg", "") or m.get("recolte_precedente_kg", ""),
            # COMPOSITION DU MENAGE
            noms_menage,
            nb_travailleurs,
            sum(1 for p in menage if str(p.get("frequentation_ecole", "")).strip() or p.get("age_scolarisation")),
            ecole_list,
            nom_ecole_list,
            classe_list,
            # INFOS COMPLEMENTAIRES
            nb_cacaoyeres,
            niveau_cert,
            nb_magasins,
            sigle,
            m.get("numero_enregistrement", ""),
            m.get("date_enregistrement", ""),
            m.get("date_audit_interne", ""),
            m.get("gps_latitude", ""),
            m.get("gps_longitude", ""),
            siege,
            campagne,
        ]

        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = tb
            cell.alignment = Alignment(wrap_text=True, vertical='center')

    # Auto column widths (excluding merged cells)
    for col in ws.columns:
        cells = [c for c in col if not isinstance(c, MergedCell)]
        if cells:
            lens = [len(str(c.value or "")) for c in cells]
            max_len = max(lens) if lens else 10
            ws.column_dimensions[cells[0].column_letter].width = min(max(max_len + 2, 12), 30)

    ws.freeze_panes = "A4"

    # Footer sheet with summary
    ws2 = wb.create_sheet("Recapitulatif")
    ws2.column_dimensions['A'].width = 45
    ws2.column_dimensions['B'].width = 30
    summary_rows = [
        ("Nom de la cooperative", coop_name),
        ("Sigle", sigle),
        ("Siege", siege),
        ("Campagne", campagne),
        ("Niveau de certification demande", niveau_cert),
        ("Nombre de producteurs membres", nb_producteurs),
        ("Nombre de sections", coop_info.get("nb_sections", "")),
        ("Nombre de cacaoyeres", nb_cacaoyeres),
        ("Nombre de magasins de stockage de feves", nb_magasins),
        ("Date de generation", datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")),
        ("Genere par", current_user.get("full_name", current_user.get("email", ""))),
    ]
    ws2["A1"] = "RECAPITULATIF DU REGISTRE"
    ws2["A1"].font = Font(bold=True, size=14, color="1A3622")
    for i, (label, val) in enumerate(summary_rows, 3):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"registre_ars1000_{coop_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ============= ETAPES & CHAMPS REFERENCE =============

@router.get("/reference/etapes")
async def get_etapes():
    return {"etapes": ETAPES_ADHESION}

@router.get("/reference/champs")
async def get_champs():
    return {"sections": SECTIONS_REGISTRE}

@router.get("/reference/sections")
async def get_sections_registre():
    return {"sections": SECTIONS_REGISTRE}



# ============= INFOS COOPERATIVE ARS 1000 =============

class CoopInfoARS(BaseModel):
    campagne: str = ""
    sigle: str = ""
    siege: str = ""
    nb_sections: int = 0
    nb_magasins_stockage: int = 0
    nb_cacaoyeres: int = 0
    niveau_certification: str = "Bronze"  # Bronze, Argent, Or

@router.get("/coop-info")
async def get_coop_info(current_user: dict = Depends(get_current_user)):
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)
    info = await db.coop_info_ars.find_one({"coop_id": coop_id}, {"_id": 0})
    if not info:
        # Auto-populate from user data
        nb_prod = await db.membres_adhesions.count_documents({"coop_id": coop_id})
        info = {
            "coop_id": coop_id,
            "nom_cooperative": current_user.get("coop_name", "") or current_user.get("full_name", ""),
            "campagne": "2024/2025",
            "sigle": "",
            "siege": current_user.get("headquarters_region", ""),
            "nb_sections": 0,
            "nb_magasins_stockage": 0,
            "nb_producteurs": nb_prod,
            "nb_cacaoyeres": 0,
            "niveau_certification": "Bronze",
        }
    return {"info": info}

@router.put("/coop-info")
async def update_coop_info(data: CoopInfoARS, current_user: dict = Depends(get_current_user)):
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)
    nb_prod = await db.membres_adhesions.count_documents({"coop_id": coop_id})

    doc = {
        "coop_id": coop_id,
        "nom_cooperative": current_user.get("coop_name", "") or current_user.get("full_name", ""),
        "campagne": data.campagne,
        "sigle": data.sigle,
        "siege": data.siege,
        "nb_sections": data.nb_sections,
        "nb_magasins_stockage": data.nb_magasins_stockage,
        "nb_producteurs": nb_prod,
        "nb_cacaoyeres": data.nb_cacaoyeres,
        "niveau_certification": data.niveau_certification,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    await db.coop_info_ars.update_one(
        {"coop_id": coop_id},
        {"$set": doc},
        upsert=True,
    )
    return {"status": "success", "info": doc}
