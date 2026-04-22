"""
Module Gestion des Risques & Durabilite - ARS 1000-1 Clause 6.1
Cartographie des Risques du SMCD - Processus en 5 etapes

Etape 1: Mesure des criteres (G: Gravite 1-5, F: Frequence 1-5)
Etape 2: Determination du niveau de risque NR = G x F et decision
Etape 3: Evaluation des mesures de prevention existantes (M: A-E)
Etape 4: Evaluation finale du risque (EFR = croisement NR/M)
Etape 5: Decision face au risque et plan de traitement
"""

from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from typing import Optional, List
from pydantic import BaseModel
from datetime import datetime, timezone
import logging
import io
import uuid

from database import db
from routes.auth import get_current_user
from routes.cooperative import verify_cooperative

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/risques", tags=["Risques & Durabilite"])


# ============= REFERENTIEL ARS 1000-1 CH 6.1 =============

CONTEXTES = [
    {"code": "ENJEUX", "label": "Questions pertinentes pour l'existence (Enjeux)"},
    {"code": "FAIBLESSES", "label": "Faiblesses issues de l'analyse 4.1"},
    {"code": "MENACES", "label": "Menaces issues de l'analyse 4.1"},
    {"code": "PLANTATION", "label": "Plantation"},
    {"code": "TRANSPORT", "label": "Transport"},
    {"code": "STOCKAGE", "label": "Stockage"},
    {"code": "TRANSFORMATION", "label": "Transformation"},
    {"code": "ADMINISTRATION", "label": "Administration"},
]

CATEGORIES_RISQUES = [
    {"code": "ENVIRONNEMENT", "label": "Environnement", "exemples": "Deforestation, degradation sols, pollution eau, perte biodiversite"},
    {"code": "SOCIAL", "label": "Social", "exemples": "Travail enfants, travail force, discrimination, SST"},
    {"code": "ECONOMIQUE", "label": "Economique", "exemples": "Volatilite prix, acces marche, endettement producteurs"},
    {"code": "CLIMATIQUE", "label": "Climatique", "exemples": "Secheresse, inondations, maladies cacaoyer, temperature"},
    {"code": "GOUVERNANCE", "label": "Gouvernance", "exemples": "Corruption, transparence, conformite reglementaire"},
    {"code": "TRACABILITE", "label": "Tracabilite", "exemples": "Melange lots, perte tracabilite, fraude origine"},
    {"code": "SST", "label": "Sante & Securite au Travail", "exemples": "Blessures, intoxication, EPI, produits chimiques"},
]

# Etape 1: Echelles G et F (1 a 5)
ECHELLE_GRAVITE = [
    {"valeur": 1, "label": "Negligeable", "description": "Consequence minime, pas d'impact notable"},
    {"valeur": 2, "label": "Mineure", "description": "Impact faible, gerable sans difficulte"},
    {"valeur": 3, "label": "Moderee", "description": "Impact significatif, necessite attention"},
    {"valeur": 4, "label": "Majeure", "description": "Impact grave, pertes importantes"},
    {"valeur": 5, "label": "Critique", "description": "Catastrophique, menace l'existence"},
]

ECHELLE_FREQUENCE = [
    {"valeur": 1, "label": "Rare", "description": "Moins d'une fois par an"},
    {"valeur": 2, "label": "Peu probable", "description": "Une fois par an"},
    {"valeur": 3, "label": "Possible", "description": "Plusieurs fois par an"},
    {"valeur": 4, "label": "Probable", "description": "Mensuel"},
    {"valeur": 5, "label": "Quasi certain", "description": "Hebdomadaire ou plus"},
]

# Etape 2: Echelle de decision NR (G x F)
ECHELLE_NR = [
    {"code": "I", "min": 1, "max": 2, "label": "Tres faible", "couleur": "#10B981"},
    {"code": "II", "min": 3, "max": 4, "label": "Faible", "couleur": "#6EE7B7"},
    {"code": "III", "min": 5, "max": 6, "label": "Peu faible", "couleur": "#FBBF24"},
    {"code": "IV", "min": 7, "max": 8, "label": "Fort", "couleur": "#F59E0B"},
    {"code": "V", "min": 9, "max": 25, "label": "Tres fort", "couleur": "#EF4444"},
]

# Etape 3: Echelle M (mesure de prevention existante)
ECHELLE_M = [
    {"code": "A", "min": 0.1, "max": 0.2, "label": "Tres pertinente", "couleur": "#10B981"},
    {"code": "B", "min": 0.3, "max": 0.4, "label": "Pertinente", "couleur": "#6EE7B7"},
    {"code": "C", "min": 0.5, "max": 0.6, "label": "Peu pertinente", "couleur": "#FBBF24"},
    {"code": "D", "min": 0.6, "max": 0.7, "label": "Tres peu pertinente", "couleur": "#F59E0B"},
    {"code": "E", "min": 0.7, "max": 1.0, "label": "Pas pertinente", "couleur": "#EF4444"},
]

# Etape 5: Decisions
DECISIONS = {
    "vert": "Mesure de maitrise tres pertinente. Aucune action supplementaire n'est a envisager.",
    "jaune": "Mesure de maitrise pertinente. D'autres mesures peuvent etre envisagees si simples et peu couteuses.",
    "rouge": "Mesure insuffisante. Action a mener immediatement, sans delai, quelles que soient les ressources necessaires.",
}

PRIORITES = [
    {"valeur": 1, "label": "Faible"},
    {"valeur": 2, "label": "Moyen"},
    {"valeur": 3, "label": "Fort"},
]


def get_coop_id(user):
    return str(user.get("coop_id") or user.get("_id") or user.get("id", ""))


def _calc_nr(g: int, f: int) -> int:
    return g * f


def _code_nr(nr: int) -> str:
    for e in ECHELLE_NR:
        if e["min"] <= nr <= e["max"]:
            return e["code"]
    return "V" if nr > 8 else "I"


def _label_nr(code: str) -> str:
    for e in ECHELLE_NR:
        if e["code"] == code:
            return e["label"]
    return ""


def _decision_efr(nr_code: str, m_code: str) -> str:
    """Determine la couleur de decision EFR selon la matrice NR x M"""
    nr_idx = ["I", "II", "III", "IV", "V"].index(nr_code) if nr_code in ["I", "II", "III", "IV", "V"] else 0
    m_idx = ["A", "B", "C", "D", "E"].index(m_code) if m_code in ["A", "B", "C", "D", "E"] else 2
    # Matrice simplifiee
    if nr_idx <= 1 and m_idx <= 1:
        return "vert"
    if nr_idx >= 3 or m_idx >= 3:
        return "rouge"
    return "jaune"


# ============= MODELS =============

class ActionTraitement(BaseModel):
    libelle: str = ""
    responsable: str = ""
    contributeur: str = ""
    echeance: str = ""

class Surveillance(BaseModel):
    methode: str = ""
    responsable: str = ""
    echeance: str = ""

class EvaluationEfficacite(BaseModel):
    critere: str = ""
    methode: str = ""
    responsable: str = ""
    date: str = ""
    enregistrements: str = ""
    efficace: str = ""  # Oui / Non / En cours
    decision: str = ""


class RisqueCreate(BaseModel):
    # Contexte
    contexte: str = ""
    libelle_enjeu: str = ""
    # Identification
    titre: str
    categorie: str = ""
    danger: str = ""
    description: str = ""
    # Etape 1: G et F
    gravite: int = 3
    frequence: int = 3
    # Etape 3: Mesure prevention existante
    mesure_prevention: str = ""
    mesure_m_code: str = "C"
    # Priorite
    priorite: int = 2
    # Causes
    causes: str = ""
    # Actions de traitement
    actions: List[ActionTraitement] = []
    # Surveillance
    surveillance: Surveillance = Surveillance()
    # Objectifs
    objectif_g: int = 0
    objectif_f: int = 0
    # Evaluation efficacite
    evaluation: EvaluationEfficacite = EvaluationEfficacite()
    # Legacy compat
    zone: str = ""
    parties_prenantes: str = ""
    cause_racine: str = ""
    probabilite: str = "Possible"
    impact: str = "Modere"


class RisqueUpdate(BaseModel):
    titre: Optional[str] = None
    description: Optional[str] = None
    danger: Optional[str] = None
    gravite: Optional[int] = None
    frequence: Optional[int] = None
    mesure_prevention: Optional[str] = None
    mesure_m_code: Optional[str] = None
    priorite: Optional[int] = None
    causes: Optional[str] = None
    actions: Optional[List[ActionTraitement]] = None
    surveillance: Optional[Surveillance] = None
    objectif_g: Optional[int] = None
    objectif_f: Optional[int] = None
    evaluation: Optional[EvaluationEfficacite] = None
    statut: Optional[str] = None
    zone: Optional[str] = None
    probabilite: Optional[str] = None
    impact: Optional[str] = None


class MitigationCreate(BaseModel):
    risque_id: str
    action: str
    responsable: str = ""
    echeance: str = ""
    ressources: str = ""


class IndicateurCreate(BaseModel):
    nom: str
    categorie: str = "ENVIRONNEMENT"
    valeur: float = 0
    unite: str = ""
    cible: float = 0
    periode: str = ""


def _build_risque_doc(data: RisqueCreate, coop_id: str, created_by: str) -> dict:
    """Build complete risk document following ARS 1000-1 Ch 6.1"""
    g = max(1, min(5, data.gravite))
    f = max(1, min(5, data.frequence))
    nr = _calc_nr(g, f)
    nr_code = _code_nr(nr)
    m_code = data.mesure_m_code if data.mesure_m_code in ["A", "B", "C", "D", "E"] else "C"
    efr_decision = _decision_efr(nr_code, m_code)

    # Legacy score for backwards compat
    legacy_niveaux = {"vert": "Faible", "jaune": "Moyen", "rouge": "Critique"}

    return {
        "risque_id": str(uuid.uuid4()),
        "coop_id": coop_id,
        # Contexte
        "contexte": data.contexte,
        "libelle_enjeu": data.libelle_enjeu,
        # Identification
        "titre": data.titre,
        "categorie": data.categorie,
        "danger": data.danger,
        "description": data.description,
        # Etape 1: Notation (G x F)
        "gravite": g,
        "frequence": f,
        "nr": nr,
        "nr_code": nr_code,
        "nr_label": _label_nr(nr_code),
        # Etape 3: Mesure de prevention
        "mesure_prevention": data.mesure_prevention,
        "mesure_m_code": m_code,
        # Etape 4: EFR
        "efr_nr": nr,
        "efr_m": m_code,
        "efr_decision": efr_decision,
        "efr_decision_texte": DECISIONS.get(efr_decision, ""),
        # Priorite de traitement
        "priorite": max(1, min(3, data.priorite)),
        # Causes
        "causes": data.causes or data.cause_racine,
        # Actions de traitement
        "actions": [a.model_dump() for a in data.actions],
        # Surveillance
        "surveillance": data.surveillance.model_dump(),
        # Objectifs
        "objectif_g": data.objectif_g,
        "objectif_f": data.objectif_f,
        # Evaluation efficacite
        "evaluation": data.evaluation.model_dump(),
        # Legacy compat
        "probabilite": data.probabilite,
        "impact": data.impact,
        "score": nr,
        "niveau": legacy_niveaux.get(efr_decision, "Moyen"),
        "zone": data.zone,
        "parties_prenantes": data.parties_prenantes,
        "cause_racine": data.causes or data.cause_racine,
        "statut": "ouvert",
        "mitigations": [],
        # Meta
        "created_by": created_by,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }


# ============= RISQUES CRUD =============

@router.post("/registre")
async def create_risque(data: RisqueCreate, current_user: dict = Depends(get_current_user)):
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)
    doc = _build_risque_doc(data, coop_id, current_user.get("full_name", ""))
    await db.risques_registre.insert_one(doc)
    doc.pop("_id", None)
    return {"status": "success", "risque": doc}


@router.get("/registre")
async def list_risques(
    current_user: dict = Depends(get_current_user),
    categorie: Optional[str] = None,
    statut: Optional[str] = None,
    niveau: Optional[str] = None,
):
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)
    query = {"coop_id": coop_id}
    if categorie:
        query["categorie"] = categorie
    if statut:
        query["statut"] = statut
    if niveau:
        query["niveau"] = niveau

    risques = await db.risques_registre.find(query, {"_id": 0}).sort("nr", -1).to_list(500)

    total = len(risques)
    critiques = sum(1 for r in risques if r.get("efr_decision") == "rouge" or r.get("niveau") == "Critique")
    eleves = sum(1 for r in risques if r.get("nr_code") in ["IV", "V"] or r.get("niveau") == "Eleve")
    ouverts = sum(1 for r in risques if r.get("statut") == "ouvert")
    mitigees = sum(1 for r in risques if r.get("statut") in ["mitige", "en_mitigation"])

    return {
        "risques": risques,
        "stats": {"total": total, "critiques": critiques, "eleves": eleves, "ouverts": ouverts, "mitigees": mitigees},
    }


@router.put("/registre/{risque_id}")
async def update_risque(risque_id: str, update: RisqueUpdate, current_user: dict = Depends(get_current_user)):
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    update_data = {k: v for k, v in update.model_dump().items() if v is not None}

    # Recalc NR if G or F changed
    if "gravite" in update_data or "frequence" in update_data:
        r = await db.risques_registre.find_one({"risque_id": risque_id, "coop_id": coop_id})
        if r:
            g = update_data.get("gravite", r.get("gravite", 3))
            f = update_data.get("frequence", r.get("frequence", 3))
            m_code = update_data.get("mesure_m_code", r.get("mesure_m_code", "C"))
            nr = _calc_nr(g, f)
            nr_code = _code_nr(nr)
            efr_decision = _decision_efr(nr_code, m_code)
            update_data.update({
                "nr": nr, "nr_code": nr_code, "nr_label": _label_nr(nr_code),
                "efr_nr": nr, "efr_m": m_code, "efr_decision": efr_decision,
                "efr_decision_texte": DECISIONS.get(efr_decision, ""),
                "score": nr,
            })

    if "mesure_m_code" in update_data:
        r = await db.risques_registre.find_one({"risque_id": risque_id, "coop_id": coop_id})
        if r:
            nr_code = r.get("nr_code", "III")
            m_code = update_data["mesure_m_code"]
            efr_decision = _decision_efr(nr_code, m_code)
            update_data.update({
                "efr_m": m_code, "efr_decision": efr_decision,
                "efr_decision_texte": DECISIONS.get(efr_decision, ""),
            })

    # Serialize nested models
    if "actions" in update_data and isinstance(update_data["actions"], list):
        update_data["actions"] = [a.model_dump() if hasattr(a, "model_dump") else a for a in update_data["actions"]]
    if "surveillance" in update_data and hasattr(update_data["surveillance"], "model_dump"):
        update_data["surveillance"] = update_data["surveillance"].model_dump()
    if "evaluation" in update_data and hasattr(update_data["evaluation"], "model_dump"):
        update_data["evaluation"] = update_data["evaluation"].model_dump()

    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.risques_registre.update_one(
        {"risque_id": risque_id, "coop_id": coop_id}, {"$set": update_data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Risque non trouve")

    r = await db.risques_registre.find_one({"risque_id": risque_id}, {"_id": 0})
    return {"status": "success", "risque": r}


# ============= MITIGATIONS =============

@router.post("/registre/{risque_id}/mitigation")
async def add_mitigation(risque_id: str, data: MitigationCreate, current_user: dict = Depends(get_current_user)):
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    mit = {
        "mitigation_id": str(uuid.uuid4()),
        "action": data.action,
        "responsable": data.responsable,
        "echeance": data.echeance,
        "ressources": data.ressources,
        "statut": "planifie",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }

    result = await db.risques_registre.update_one(
        {"risque_id": risque_id, "coop_id": coop_id},
        {"$push": {"mitigations": mit}, "$set": {"statut": "en_mitigation", "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Risque non trouve")

    return {"status": "success", "mitigation": mit}


# ============= INDICATEURS =============

@router.post("/indicateurs")
async def add_indicateur(data: IndicateurCreate, current_user: dict = Depends(get_current_user)):
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    doc = {
        "indicateur_id": str(uuid.uuid4()),
        "coop_id": coop_id,
        "nom": data.nom,
        "categorie": data.categorie,
        "valeur": data.valeur,
        "unite": data.unite,
        "cible": data.cible,
        "periode": data.periode or datetime.now(timezone.utc).strftime("%Y-%m"),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.risques_indicateurs.insert_one(doc)
    doc.pop("_id", None)
    return {"status": "success", "indicateur": doc}


@router.get("/indicateurs")
async def list_indicateurs(current_user: dict = Depends(get_current_user)):
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)
    inds = await db.risques_indicateurs.find({"coop_id": coop_id}, {"_id": 0}).sort("created_at", -1).to_list(200)
    return {"indicateurs": inds}


# ============= DASHBOARD =============

@router.get("/dashboard")
async def get_risques_dashboard(current_user: dict = Depends(get_current_user)):
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    risques = await db.risques_registre.find({"coop_id": coop_id}, {"_id": 0}).to_list(500)
    indicateurs = await db.risques_indicateurs.find({"coop_id": coop_id}, {"_id": 0}).to_list(200)

    total = len(risques)
    critiques = sum(1 for r in risques if r.get("efr_decision") == "rouge" or r.get("niveau") == "Critique")
    eleves = sum(1 for r in risques if r.get("nr_code") in ["IV", "V"] or r.get("niveau") == "Eleve")
    ouverts = sum(1 for r in risques if r.get("statut") == "ouvert")
    mitigees = sum(1 for r in risques if r.get("statut") in ["mitige", "en_mitigation"])

    # Matrice 5x5 (G x F)
    matrice = [[0]*5 for _ in range(5)]
    for r in risques:
        g = r.get("gravite", 0)
        f = r.get("frequence", 0)
        if 1 <= g <= 5 and 1 <= f <= 5:
            matrice[g-1][f-1] += 1

    # Repartition par EFR
    repartition_efr = {"vert": 0, "jaune": 0, "rouge": 0}
    for r in risques:
        d = r.get("efr_decision", "jaune")
        if d in repartition_efr:
            repartition_efr[d] += 1

    # Par categorie
    par_categorie = {}
    for r in risques:
        cat = r.get("categorie", "Autre")
        if cat not in par_categorie:
            par_categorie[cat] = {"count": 0, "score_moyen": 0, "critiques": 0}
        par_categorie[cat]["count"] += 1
        par_categorie[cat]["score_moyen"] += r.get("nr", r.get("score", 0))
        if r.get("efr_decision") == "rouge":
            par_categorie[cat]["critiques"] += 1
    for cat in par_categorie:
        if par_categorie[cat]["count"] > 0:
            par_categorie[cat]["score_moyen"] = round(par_categorie[cat]["score_moyen"] / par_categorie[cat]["count"], 1)

    categories_list = [{"categorie": k, **v} for k, v in par_categorie.items()]

    return {
        "kpis": {
            "total": total, "critiques": critiques, "eleves": eleves,
            "ouverts": ouverts, "mitigees": mitigees,
            "indicateurs": len(indicateurs),
        },
        "matrice_5x5": matrice,
        "repartition_efr": repartition_efr,
        "par_categorie": categories_list,
        "categories_reference": CATEGORIES_RISQUES,
        "top_risques": sorted(risques, key=lambda x: x.get("nr", x.get("score", 0)), reverse=True)[:5],
    }


# ============= REFERENCE =============

@router.get("/reference/categories")
async def get_categories():
    return {"categories": CATEGORIES_RISQUES}


@router.get("/reference/contextes")
async def get_contextes():
    return {"contextes": CONTEXTES}


@router.get("/reference/echelles")
async def get_echelles():
    return {
        "gravite": ECHELLE_GRAVITE,
        "frequence": ECHELLE_FREQUENCE,
        "nr": ECHELLE_NR,
        "m": ECHELLE_M,
        "priorites": PRIORITES,
        "decisions": DECISIONS,
    }


@router.get("/reference/niveaux")
async def get_niveaux():
    """Legacy endpoint compat"""
    return {
        "probabilite": [e["label"] for e in ECHELLE_FREQUENCE],
        "impact": [e["label"] for e in ECHELLE_GRAVITE],
    }


# ============= EXPORT EXCEL =============

@router.get("/export/excel")
async def export_cartographie_excel(current_user: dict = Depends(get_current_user)):
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    risques = await db.risques_registre.find({"coop_id": coop_id}, {"_id": 0}).sort("nr", -1).to_list(500)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Cartographie des Risques"
    hf = PatternFill(start_color="1A3622", end_color="1A3622", fill_type="solid")
    sf = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True, size=8)
    sfont = Font(color="000000", bold=True, size=8)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Title
    ws.merge_cells("A1:AC1")
    c = ws["A1"]
    c.value = "CARTOGRAPHIE DES RISQUES DU SMCD - ARS 1000-1 Clause 6.1"
    c.fill = hf
    c.font = Font(color="FFFFFF", bold=True, size=11)
    c.alignment = Alignment(horizontal='center')

    # Section headers row 2
    sections = [
        ("A", "D", "IDENTIFICATION"),
        ("E", "G", "NOTATION NR (G x F)"),
        ("H", "I", "MESURE PREVENTION"),
        ("J", "L", "EFR"),
        ("M", "M", "PRIORITE"),
        ("N", "N", "CAUSES"),
        ("O", "R", "ACTIONS DE TRAITEMENT"),
        ("S", "U", "SURVEILLANCE"),
        ("V", "W", "OBJECTIFS"),
        ("X", "AC", "EVALUATION EFFICACITE"),
    ]
    for start, end, title in sections:
        cell = ws[f"{start}2"]
        cell.value = title
        cell.fill = sf
        cell.font = sfont
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        if start != end:
            ws.merge_cells(f"{start}2:{end}2")

    # Column headers row 3
    headers = [
        "Contexte", "Enjeu / Activite", "Danger", "Risque",
        "G", "F", "NR (GxF)",
        "Mesure prevention", "M",
        "NR", "M", "Decision",
        "Priorite",
        "Causes",
        "Action", "Responsable", "Contributeur", "Echeance",
        "Methode", "Responsable", "Echeance",
        "G cible", "F cible",
        "Critere efficacite", "Methode", "Responsable", "Date", "Enregistrements", "Efficace?"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = PatternFill(start_color="E8F0EA", end_color="E8F0EA", fill_type="solid")
        cell.font = Font(bold=True, size=7)
        cell.border = tb
        cell.alignment = Alignment(wrap_text=True)

    for row, r in enumerate(risques, 4):
        surv = r.get("surveillance", {})
        ev = r.get("evaluation", {})
        acts = r.get("actions", [])
        act = acts[0] if acts else {}

        vals = [
            r.get("contexte", ""), r.get("libelle_enjeu", ""), r.get("danger", ""), r.get("titre", ""),
            r.get("gravite", ""), r.get("frequence", ""), r.get("nr", ""),
            r.get("mesure_prevention", ""), r.get("mesure_m_code", ""),
            r.get("efr_nr", ""), r.get("efr_m", ""), r.get("efr_decision", ""),
            r.get("priorite", ""),
            r.get("causes", ""),
            act.get("libelle", ""), act.get("responsable", ""), act.get("contributeur", ""), act.get("echeance", ""),
            surv.get("methode", ""), surv.get("responsable", ""), surv.get("echeance", ""),
            r.get("objectif_g", ""), r.get("objectif_f", ""),
            ev.get("critere", ""), ev.get("methode", ""), ev.get("responsable", ""), ev.get("date", ""), ev.get("enregistrements", ""), ev.get("efficace", ""),
        ]
        for col, val in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=val).border = tb

    # Auto-width
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        cells = [cell for cell in col if not isinstance(cell, MergedCell)]
        if cells:
            max_len = max(len(str(cell.value or "")) for cell in cells)
            ws.column_dimensions[cells[0].column_letter].width = min(max(max_len + 2, 8), 22)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={
        "Content-Disposition": "attachment; filename=cartographie_risques_ars1000.xlsx"
    })



# ============= AUTO-DIAGNOSTIC RISQUES (Inspire Rainforest Alliance S3) =============
# Questionnaire Oui/Non adapte au contexte ARS 1000 Cacao CI
# Chaque reponse "a risque" genere une fiche risque pre-remplie

AUTO_DIAGNOSTIC_QUESTIONS = [
    # THEME: TRACABILITE
    {"id": "T1", "theme": "TRACABILITE", "question": "Utilisez-vous des intermediaires ou sous-traitants dans votre chaine d'approvisionnement?", "reponse_risque": "oui",
     "risque_titre": "Perte de tracabilite via intermediaires", "categorie": "TRACABILITE", "gravite_suggeree": 4, "frequence_suggeree": 3,
     "actions_suggerees": ["Mettre en place un flux de tracabilite documente incluant tous les acteurs", "Former tous les acteurs sur la procedure de tracabilite", "Verifier la calibration des balances et la tenue des registres chez les intermediaires", "Surveiller les intermediaires pendant la periode de recolte"]},
    {"id": "T2", "theme": "TRACABILITE", "question": "Les producteurs ont-ils des difficultes a conserver les recus/documents de tracabilite?", "reponse_risque": "oui",
     "risque_titre": "Defaillance de la tenue des registres producteurs", "categorie": "TRACABILITE", "gravite_suggeree": 3, "frequence_suggeree": 4,
     "actions_suggerees": ["Inclure la gestion des recus dans le plan de formation", "Fournir des pochettes plastiques pour la conservation", "Afficher des rappels pour encourager la conservation", "Verifier la tenue des recus lors des inspections internes"]},
    {"id": "T3", "theme": "TRACABILITE", "question": "Gerez-vous a la fois du produit certifie et non certifie?", "reponse_risque": "oui",
     "risque_titre": "Risque de melange produit certifie/non certifie", "categorie": "TRACABILITE", "gravite_suggeree": 5, "frequence_suggeree": 3,
     "actions_suggerees": ["Implementer un systeme d'identification physique des produits certifies (etiquettes, tags sur sacs)", "Documenter les regles de separation dans les registres", "Former tous les acteurs sur la procedure de separation"]},
    {"id": "T4", "theme": "TRACABILITE", "question": "Les producteurs vendent-ils a plusieurs acheteurs differents?", "reponse_risque": "oui",
     "risque_titre": "Fiabilite des volumes declares compromise", "categorie": "TRACABILITE", "gravite_suggeree": 3, "frequence_suggeree": 4,
     "actions_suggerees": ["Collecter les informations sur les volumes recoltes mensuellement", "Verifier les volumes lors des inspections internes", "Croiser les donnees de livraison avec les declarations"]},
    # THEME: AGROCHIMIQUES
    {"id": "A1", "theme": "AGROCHIMIQUES", "question": "Est-il courant dans la region d'utiliser des produits agrochimiques de la liste des substances interdites?", "reponse_risque": "oui",
     "risque_titre": "Utilisation de pesticides interdits", "categorie": "SST", "gravite_suggeree": 5, "frequence_suggeree": 3,
     "actions_suggerees": ["Former les producteurs sur la liste des agrochimiques interdits ARS 1000", "Verifier l'utilisation lors des inspections internes", "Surveiller pendant la periode d'application", "Mettre en place un systeme de collecte des stocks de produits interdits"]},
    {"id": "A2", "theme": "AGROCHIMIQUES", "question": "Les producteurs appliquent-ils d'abord des methodes biologiques/physiques avant les agrochimiques (IPM)?", "reponse_risque": "non",
     "risque_titre": "Surutil. agrochimiques sans gestion integree (IPM)", "categorie": "ENVIRONNEMENT", "gravite_suggeree": 4, "frequence_suggeree": 3,
     "actions_suggerees": ["Former les membres sur la Gestion Integree des Ravageurs (IPM)", "Identifier des sources de produits de lutte biologique", "Surveiller l'utilisation des agrochimiques et la tenue des registres", "Contacter un service de vulgarisation pour la procedure IPM"]},
    {"id": "A3", "theme": "AGROCHIMIQUES", "question": "Les producteurs et travailleurs portent-ils systematiquement les EPI lors de l'application des agrochimiques?", "reponse_risque": "non",
     "risque_titre": "Non-port des EPI lors des traitements phytosanitaires", "categorie": "SST", "gravite_suggeree": 5, "frequence_suggeree": 4,
     "actions_suggerees": ["Assurer la disponibilite des EPI pour tous les applicateurs", "Former sur l'utilisation correcte des EPI et des agrochimiques", "Explorer la creation d'equipes de pulverisation specialisees", "Surveiller le port des EPI pendant les periodes d'application"]},
    # THEME: ENVIRONNEMENT
    {"id": "E1", "theme": "ENVIRONNEMENT", "question": "Y a-t-il des zones avec des pentes superieures a 1m de denivele sur 3m?", "reponse_risque": "oui",
     "risque_titre": "Erosion des sols sur pentes fortes", "categorie": "ENVIRONNEMENT", "gravite_suggeree": 4, "frequence_suggeree": 3,
     "actions_suggerees": ["Planter une couverture vegetale native", "Implementer des plantations en courbes de niveau", "Installer des barrieres vivantes et systemes de drainage"]},
    {"id": "E2", "theme": "ENVIRONNEMENT", "question": "Y a-t-il des zones avec de l'eau stagnante prolongee apres la pluie?", "reponse_risque": "oui",
     "risque_titre": "Engorgement des sols et asphyxie racinaire", "categorie": "CLIMATIQUE", "gravite_suggeree": 3, "frequence_suggeree": 3,
     "actions_suggerees": ["Ameliorer le drainage par des tranchees", "Ameliorer la structure du sol", "Evaluer l'adequation des cultures pour ces zones"]},
    {"id": "E3", "theme": "ENVIRONNEMENT", "question": "La secheresse est-elle un facteur limitant pour la production?", "reponse_risque": "oui",
     "risque_titre": "Impact de la secheresse sur la productivite", "categorie": "CLIMATIQUE", "gravite_suggeree": 4, "frequence_suggeree": 4,
     "actions_suggerees": ["Maintenir le sol couvert pour reduire l'evapotranspiration", "Utiliser des cultures a enracinement profond", "Favoriser les cultures mixtes avec arbres d'ombrage", "Minimiser les pertes d'eau en cas d'irrigation"]},
    {"id": "E4", "theme": "ENVIRONNEMENT", "question": "Les sites de production des membres s'etendent-ils vers de nouvelles zones (deforestation)?", "reponse_risque": "oui",
     "risque_titre": "Risque de deforestation par expansion des parcelles", "categorie": "ENVIRONNEMENT", "gravite_suggeree": 5, "frequence_suggeree": 3,
     "actions_suggerees": ["Sensibiliser sur le maintien des ecosystemes naturels", "Marquer clairement les limites des ecosystemes naturels et zones tampons", "Surveiller regulierement que les activites n'empietent pas sur ces zones"]},
    # THEME: SOCIAL
    {"id": "S1", "theme": "SOCIAL", "question": "Y a-t-il des travailleurs migrants, minorites ethniques ou personnes ne parlant pas la langue dominante?", "reponse_risque": "oui",
     "risque_titre": "Risque de discrimination des populations vulnerables", "categorie": "SOCIAL", "gravite_suggeree": 4, "frequence_suggeree": 3,
     "actions_suggerees": ["Evaluer la presence et les besoins de ces populations", "Enregistrer les specificites: type de population, nombre, langue", "Publier les postes dans les langues appropriees", "Former l'encadrement sur les biais inconscients et pratiques anti-discriminatoires"]},
    {"id": "S2", "theme": "SOCIAL", "question": "L'encadrement a-t-il pris des mesures pour prevenir la violence et le harcelement au travail?", "reponse_risque": "non",
     "risque_titre": "Absence de prevention violence/harcelement au travail", "categorie": "SOCIAL", "gravite_suggeree": 4, "frequence_suggeree": 3,
     "actions_suggerees": ["Former les responsables sur le comportement respectueux", "Former les travailleurs sur les concepts de harcelement", "Identifier les zones a risque sur le lieu de travail", "Mettre en place un mecanisme de plainte accessible"]},
    {"id": "S3", "theme": "SOCIAL", "question": "Y a-t-il un risque que l'age des travailleurs ne soit pas verifie a l'embauche?", "reponse_risque": "oui",
     "risque_titre": "Risque de travail des enfants par defaut de verification", "categorie": "SOCIAL", "gravite_suggeree": 5, "frequence_suggeree": 3,
     "actions_suggerees": ["Communiquer a tous les producteurs la methode de verification de l'age", "Verifier l'age sur base de documents d'identite, registres scolaires ou medicaux", "Controler lors des inspections internes la liste des travailleurs"]},
    {"id": "S4", "theme": "SOCIAL", "question": "Le mecanisme de plainte est-il visible et accessible a tous les producteurs et travailleurs?", "reponse_risque": "non",
     "risque_titre": "Mecanisme de plainte inaccessible", "categorie": "GOUVERNANCE", "gravite_suggeree": 3, "frequence_suggeree": 4,
     "actions_suggerees": ["Afficher les informations du mecanisme de plainte dans les langues locales", "S'assurer que l'information est visible et accessible a tous", "Verifier et mettre a jour regulierement les affichages"]},
    # THEME: ECONOMIQUE
    {"id": "EC1", "theme": "ECONOMIQUE", "question": "Le rendement moyen des producteurs est-il inferieur a la moyenne nationale?", "reponse_risque": "oui",
     "risque_titre": "Rendements sous-optimaux des producteurs", "categorie": "ECONOMIQUE", "gravite_suggeree": 3, "frequence_suggeree": 4,
     "actions_suggerees": ["Former le personnel a identifier les contraintes de productivite au champ", "Etablir des parcelles de demonstration (modeles)", "Mettre en place des essais de rajeunissement, fertilisation et lutte phytosanitaire"]},
    {"id": "EC2", "theme": "ECONOMIQUE", "question": "Tous les producteurs ont-ils acces aux intrants agricoles et aux connaissances pour optimiser la productivite?", "reponse_risque": "non",
     "risque_titre": "Deficit d'acces aux intrants et connaissances", "categorie": "ECONOMIQUE", "gravite_suggeree": 3, "frequence_suggeree": 4,
     "actions_suggerees": ["Identifier les besoins principaux des membres en intrants et formation", "Former sur la gestion financiere et les couts de production", "Faciliter l'acces aux services financiers (prets investissements)"]},
    {"id": "EC3", "theme": "ECONOMIQUE", "question": "Estimez-vous que tous les producteurs gagnent un revenu decent de la production du cacao?", "reponse_risque": "non",
     "risque_titre": "Revenu des producteurs insuffisant (revenu vital)", "categorie": "ECONOMIQUE", "gravite_suggeree": 4, "frequence_suggeree": 4,
     "actions_suggerees": ["Evaluer le revenu net total d'un echantillon representatif de menages", "Former sur la gestion financiere et la comprehension des couts", "Accompagner les decisions de diversification des revenus"]},
]


@router.get("/auto-diagnostic/questions")
async def get_auto_diagnostic_questions():
    """Retourne le questionnaire d'auto-diagnostic adapte ARS 1000"""
    # Group by theme
    themes = {}
    for q in AUTO_DIAGNOSTIC_QUESTIONS:
        t = q["theme"]
        if t not in themes:
            themes[t] = []
        themes[t].append(q)
    return {"questions": AUTO_DIAGNOSTIC_QUESTIONS, "par_theme": themes, "total": len(AUTO_DIAGNOSTIC_QUESTIONS)}


@router.post("/auto-diagnostic/evaluer")
async def evaluer_auto_diagnostic(reponses: dict, current_user: dict = Depends(get_current_user)):
    """Evalue les reponses du questionnaire et genere les risques identifies"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    risques_identifies = []
    score_par_theme = {}

    for q in AUTO_DIAGNOSTIC_QUESTIONS:
        theme = q["theme"]
        if theme not in score_par_theme:
            score_par_theme[theme] = {"total": 0, "a_risque": 0}
        score_par_theme[theme]["total"] += 1

        reponse = reponses.get(q["id"], "")
        est_a_risque = (reponse.lower() == q["reponse_risque"].lower())

        if est_a_risque:
            score_par_theme[theme]["a_risque"] += 1
            risques_identifies.append({
                "question_id": q["id"],
                "theme": theme,
                "question": q["question"],
                "titre": q["risque_titre"],
                "categorie": q["categorie"],
                "gravite_suggeree": q["gravite_suggeree"],
                "frequence_suggeree": q["frequence_suggeree"],
                "actions_suggerees": q["actions_suggerees"],
            })

    # Score global
    total_q = len(AUTO_DIAGNOSTIC_QUESTIONS)
    total_risques = len(risques_identifies)
    score_global = round(((total_q - total_risques) / max(total_q, 1)) * 100)

    # Score par theme
    for theme in score_par_theme:
        t = score_par_theme[theme]
        t["score"] = round(((t["total"] - t["a_risque"]) / max(t["total"], 1)) * 100)

    # Save diagnostic result
    doc = {
        "diagnostic_id": str(uuid.uuid4()),
        "coop_id": coop_id,
        "reponses": reponses,
        "risques_identifies": risques_identifies,
        "score_global": score_global,
        "score_par_theme": score_par_theme,
        "total_questions": total_q,
        "total_risques": total_risques,
        "created_by": current_user.get("full_name", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.risques_auto_diagnostic.insert_one(doc)
    doc.pop("_id", None)

    return {"status": "success", "diagnostic": doc}


@router.post("/auto-diagnostic/generer-risques")
async def generer_risques_from_diagnostic(diagnostic_id: str, current_user: dict = Depends(get_current_user)):
    """Genere les fiches risque ARS 1000 a partir du diagnostic"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    diag = await db.risques_auto_diagnostic.find_one({"diagnostic_id": diagnostic_id, "coop_id": coop_id})
    if not diag:
        raise HTTPException(status_code=404, detail="Diagnostic non trouve")

    created = []
    for r in diag.get("risques_identifies", []):
        from routes.risques_durabilite import RisqueCreate, ActionTraitement, _build_risque_doc
        actions = [ActionTraitement(libelle=a) for a in r.get("actions_suggerees", [])]
        data = RisqueCreate(
            titre=r["titre"],
            categorie=r["categorie"],
            contexte="AUTO-DIAGNOSTIC",
            libelle_enjeu=r["question"],
            gravite=r["gravite_suggeree"],
            frequence=r["frequence_suggeree"],
            mesure_m_code="C",
            priorite=2 if r["gravite_suggeree"] < 4 else 3,
            causes=f"Identifie par auto-diagnostic: {r['question']}",
            actions=actions,
        )
        doc = _build_risque_doc(data, coop_id, current_user.get("full_name", ""))
        doc["source"] = "auto-diagnostic"
        doc["diagnostic_id"] = diagnostic_id
        await db.risques_registre.insert_one(doc)
        doc.pop("_id", None)
        created.append(doc)

    return {"status": "success", "risques_crees": len(created), "risques": created}


@router.get("/auto-diagnostic/historique")
async def get_diagnostic_historique(current_user: dict = Depends(get_current_user)):
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)
    docs = await db.risques_auto_diagnostic.find({"coop_id": coop_id}, {"_id": 0}).sort("created_at", -1).to_list(20)
    return {"diagnostics": docs}



# ============= EXPORT PDF / EXCEL (Auto-Diagnostic) =============

@router.get("/auto-diagnostic/{diagnostic_id}/pdf")
async def export_diagnostic_pdf(diagnostic_id: str, current_user: dict = Depends(get_current_user)):
    """PDF officiel d'un auto-diagnostic (ARS 1000 Ch. 6.1)."""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)
    diag = await db.risques_auto_diagnostic.find_one({"diagnostic_id": diagnostic_id, "coop_id": coop_id})
    if not diag:
        raise HTTPException(status_code=404, detail="Diagnostic non trouve")

    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_CENTER
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.platypus.flowables import HRFlowable

    FOREST = colors.HexColor("#1A3622")
    GOLD = colors.HexColor("#D4AF37")
    MUTED = colors.HexColor("#6B7280")
    BORDER = colors.HexColor("#E5E5E0")
    LIGHT = colors.HexColor("#FAF9F6")

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle("H1", parent=styles["Heading1"], fontSize=16, textColor=FOREST, alignment=TA_CENTER, fontName="Helvetica-Bold"))
    styles.add(ParagraphStyle("Sub", parent=styles["Normal"], fontSize=9, textColor=MUTED, alignment=TA_CENTER, spaceAfter=10))
    styles.add(ParagraphStyle("Sec", parent=styles["Heading2"], fontSize=11, textColor=FOREST, spaceBefore=10, spaceAfter=6, fontName="Helvetica-Bold"))
    styles.add(ParagraphStyle("Body", parent=styles["Normal"], fontSize=9, leading=11))
    styles.add(ParagraphStyle("Cell", parent=styles["Normal"], fontSize=8, leading=10))
    styles.add(ParagraphStyle("Small", parent=styles["Normal"], fontSize=7, textColor=MUTED, leading=9))

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.2 * cm, rightMargin=1.2 * cm,
        topMargin=1.2 * cm, bottomMargin=1.2 * cm,
        title="Auto-Diagnostic Risques & Durabilite",
    )
    story = []

    coop_name = current_user.get("coop_name") or current_user.get("full_name", "")
    created_at = diag.get("created_at", datetime.now(timezone.utc).isoformat())[:19].replace("T", " ")
    score = diag.get("score_global", 0)
    total_q = diag.get("total_questions", 0)
    total_r = diag.get("total_risques", 0)

    story.append(Paragraph("Auto-Diagnostic Risques & Durabilite", styles["H1"]))
    story.append(Paragraph(f"{coop_name} - Genere le {created_at} - ARS 1000 Chapitre 6.1", styles["Sub"]))
    story.append(HRFlowable(width="100%", thickness=1, color=GOLD, spaceBefore=2, spaceAfter=10))

    # Headline box
    score_color = "#10B981" if score >= 70 else ("#F59E0B" if score >= 40 else "#EF4444")
    headline = Table([[
        Paragraph(f'<font size="24" color="{score_color}"><b>{score}%</b></font>', styles["Body"]),
        Paragraph(f'<b>Score Global</b><br/>{total_r} risque(s) identifie(s) sur {total_q} questions', styles["Body"]),
    ]], colWidths=[3.5 * cm, 14 * cm])
    headline.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 1, BORDER),
        ("BACKGROUND", (0, 0), (0, 0), LIGHT),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
        ("LEFTPADDING", (0, 0), (-1, -1), 14),
    ]))
    story.append(headline)

    # Score par theme
    story.append(Paragraph("Score par theme", styles["Sec"]))
    theme_rows = [[
        Paragraph("<b>Theme</b>", styles["Cell"]),
        Paragraph("<b>Questions</b>", styles["Cell"]),
        Paragraph("<b>Risques</b>", styles["Cell"]),
        Paragraph("<b>Score</b>", styles["Cell"]),
    ]]
    for theme, s in (diag.get("score_par_theme") or {}).items():
        theme_rows.append([
            Paragraph(theme, styles["Cell"]),
            Paragraph(str(s.get("total", 0)), styles["Cell"]),
            Paragraph(str(s.get("a_risque", 0)), styles["Cell"]),
            Paragraph(f"{s.get('score', 0)}%", styles["Cell"]),
        ])
    t = Table(theme_rows, colWidths=[9 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), FOREST),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, BORDER),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
    ]))
    story.append(t)

    # Risques identifies
    risques = diag.get("risques_identifies") or []
    if risques:
        story.append(PageBreak())
        story.append(Paragraph(f"Risques identifies ({len(risques)})", styles["Sec"]))
        rows = [[
            Paragraph("<b>#</b>", styles["Cell"]),
            Paragraph("<b>Titre / Categorie</b>", styles["Cell"]),
            Paragraph("<b>G</b>", styles["Cell"]),
            Paragraph("<b>F</b>", styles["Cell"]),
            Paragraph("<b>Actions suggerees</b>", styles["Cell"]),
        ]]
        for i, r in enumerate(risques, 1):
            actions_str = "<br/>".join(f"- {a}" for a in (r.get("actions_suggerees") or []))
            rows.append([
                Paragraph(str(i), styles["Cell"]),
                Paragraph(f"<b>{r.get('titre', '')}</b><br/><font size=7 color='#6B7280'>{r.get('categorie', '')}</font>", styles["Cell"]),
                Paragraph(str(r.get("gravite_suggeree", "")), styles["Cell"]),
                Paragraph(str(r.get("frequence_suggeree", "")), styles["Cell"]),
                Paragraph(actions_str or "-", styles["Cell"]),
            ])
        rt = Table(rows, colWidths=[0.8 * cm, 6 * cm, 1 * cm, 1 * cm, 8.7 * cm])
        rt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), FOREST),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, BORDER),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
        ]))
        story.append(rt)

    # Signature footer
    story.append(Spacer(1, 1 * cm))
    story.append(HRFlowable(width="100%", thickness=0.3, color=BORDER))
    story.append(Paragraph(
        f"<b>Cooperative :</b> {coop_name} &nbsp;&nbsp; <b>Responsable :</b> {diag.get('created_by', '-')} &nbsp;&nbsp; <b>Date :</b> {created_at}",
        styles["Small"],
    ))
    story.append(Paragraph(
        "Signature du responsable HSE / Coordinateur ARS 1000 : __________________________    Cachet : __________________________",
        styles["Small"],
    ))
    story.append(Paragraph(
        "Diagnostic genere selon le referentiel ARS 1000 Cacao Durable - Chapitre 6.1 Processus d'evaluation des risques.",
        styles["Small"],
    ))

    doc.build(story)
    buf.seek(0)
    filename = f"auto_diagnostic_{diagnostic_id[:8]}.pdf"
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/auto-diagnostic/{diagnostic_id}/excel")
async def export_diagnostic_excel(diagnostic_id: str, current_user: dict = Depends(get_current_user)):
    """Excel d'un auto-diagnostic (reponses + risques identifies)."""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)
    diag = await db.risques_auto_diagnostic.find_one({"diagnostic_id": diagnostic_id, "coop_id": coop_id})
    if not diag:
        raise HTTPException(status_code=404, detail="Diagnostic non trouve")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    hf = PatternFill(start_color="1A3622", end_color="1A3622", fill_type="solid")
    sf = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True, size=10)
    sfont = Font(color="000000", bold=True, size=9)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    coop_name = current_user.get("coop_name") or current_user.get("full_name", "")

    # Sheet 1: Resume
    ws = wb.active
    ws.title = "Resume"
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 30
    ws["A1"] = "AUTO-DIAGNOSTIC RISQUES & DURABILITE"
    ws["A1"].font = Font(bold=True, size=14, color="1A3622")
    ws.merge_cells("A1:B1")
    info = [
        ("Cooperative", coop_name),
        ("Date", diag.get("created_at", "")[:19].replace("T", " ")),
        ("Responsable", diag.get("created_by", "")),
        ("Score global", f"{diag.get('score_global', 0)}%"),
        ("Total questions", diag.get("total_questions", 0)),
        ("Total risques identifies", diag.get("total_risques", 0)),
    ]
    for i, (k, v) in enumerate(info, 3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)

    # Sheet 2: Scores par theme
    ws2 = wb.create_sheet("Scores par theme")
    headers_t = ["Theme", "Total questions", "A risque", "Score %"]
    for i, h in enumerate(headers_t, 1):
        cell = ws2.cell(row=1, column=i, value=h)
        cell.fill = hf
        cell.font = hfont
        cell.border = tb
        cell.alignment = center
    for row, (theme, s) in enumerate((diag.get("score_par_theme") or {}).items(), 2):
        ws2.cell(row=row, column=1, value=theme).border = tb
        ws2.cell(row=row, column=2, value=s.get("total", 0)).border = tb
        ws2.cell(row=row, column=3, value=s.get("a_risque", 0)).border = tb
        ws2.cell(row=row, column=4, value=s.get("score", 0)).border = tb
    for col_letter in ["A", "B", "C", "D"]:
        ws2.column_dimensions[col_letter].width = 25 if col_letter == "A" else 15

    # Sheet 3: Risques identifies
    ws3 = wb.create_sheet("Risques identifies")
    headers_r = ["#", "Titre", "Categorie", "Question source", "Gravite", "Frequence", "Actions suggerees"]
    for i, h in enumerate(headers_r, 1):
        cell = ws3.cell(row=1, column=i, value=h)
        cell.fill = hf
        cell.font = hfont
        cell.border = tb
        cell.alignment = center
    for row, r in enumerate(diag.get("risques_identifies") or [], 2):
        actions_str = " | ".join(r.get("actions_suggerees") or [])
        values = [
            row - 1,
            r.get("titre", ""),
            r.get("categorie", ""),
            r.get("question", ""),
            r.get("gravite_suggeree", ""),
            r.get("frequence_suggeree", ""),
            actions_str,
        ]
        for i, v in enumerate(values, 1):
            cell = ws3.cell(row=row, column=i, value=v)
            cell.border = tb
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = {"A": 6, "B": 35, "C": 20, "D": 40, "E": 10, "F": 12, "G": 50}
    for col_letter, w in widths.items():
        ws3.column_dimensions[col_letter].width = w

    # Sheet 4: Reponses detaillees
    ws4 = wb.create_sheet("Reponses")
    hdrs = ["Question ID", "Reponse"]
    for i, h in enumerate(hdrs, 1):
        cell = ws4.cell(row=1, column=i, value=h)
        cell.fill = sf
        cell.font = sfont
        cell.border = tb
    reponses = diag.get("reponses") or {}
    for row, (qid, rep) in enumerate(reponses.items(), 2):
        ws4.cell(row=row, column=1, value=qid).border = tb
        ws4.cell(row=row, column=2, value=rep).border = tb
    ws4.column_dimensions["A"].width = 20
    ws4.column_dimensions["B"].width = 10

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"auto_diagnostic_{diagnostic_id[:8]}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )



# ============= COMPARAISON 2 DIAGNOSTICS =============

@router.get("/auto-diagnostic/{diagnostic_id}/comparer")
async def comparer_diagnostic(diagnostic_id: str, current_user: dict = Depends(get_current_user)):
    """
    Compare le diagnostic courant avec le precedent de la meme cooperative.
    Retourne : score delta, risques nouveaux / resolus / persistants,
    score par theme avant/apres, progression globale.
    """
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    current = await db.risques_auto_diagnostic.find_one(
        {"diagnostic_id": diagnostic_id, "coop_id": coop_id},
        {"_id": 0},
    )
    if not current:
        raise HTTPException(status_code=404, detail="Diagnostic courant non trouve")

    # Previous = any diagnostic before this one (by created_at desc)
    previous = await db.risques_auto_diagnostic.find_one(
        {"coop_id": coop_id, "created_at": {"$lt": current.get("created_at", "")}},
        {"_id": 0},
        sort=[("created_at", -1)],
    )
    if not previous:
        return {
            "has_previous": False,
            "current": current,
            "message": "Aucun diagnostic precedent pour comparaison",
        }

    # Keys = question_id identifying a risk
    curr_risks_by_q = {r.get("question"): r for r in (current.get("risques_identifies") or [])}
    prev_risks_by_q = {r.get("question"): r for r in (previous.get("risques_identifies") or [])}

    nouveaux = [curr_risks_by_q[q] for q in curr_risks_by_q if q not in prev_risks_by_q]
    resolus = [prev_risks_by_q[q] for q in prev_risks_by_q if q not in curr_risks_by_q]
    persistants = [curr_risks_by_q[q] for q in curr_risks_by_q if q in prev_risks_by_q]

    # Theme progression
    themes = {}
    curr_themes = current.get("score_par_theme") or {}
    prev_themes = previous.get("score_par_theme") or {}
    all_themes = set(curr_themes.keys()) | set(prev_themes.keys())
    for t in all_themes:
        c = curr_themes.get(t, {"score": 0, "total": 0, "a_risque": 0})
        p = prev_themes.get(t, {"score": 0, "total": 0, "a_risque": 0})
        themes[t] = {
            "score_current": c.get("score", 0),
            "score_previous": p.get("score", 0),
            "delta": c.get("score", 0) - p.get("score", 0),
            "risques_current": c.get("a_risque", 0),
            "risques_previous": p.get("a_risque", 0),
        }

    score_delta = current.get("score_global", 0) - previous.get("score_global", 0)
    return {
        "has_previous": True,
        "current": {
            "diagnostic_id": current.get("diagnostic_id"),
            "created_at": current.get("created_at"),
            "score_global": current.get("score_global", 0),
            "total_risques": current.get("total_risques", 0),
            "total_questions": current.get("total_questions", 0),
            "created_by": current.get("created_by", ""),
        },
        "previous": {
            "diagnostic_id": previous.get("diagnostic_id"),
            "created_at": previous.get("created_at"),
            "score_global": previous.get("score_global", 0),
            "total_risques": previous.get("total_risques", 0),
            "total_questions": previous.get("total_questions", 0),
            "created_by": previous.get("created_by", ""),
        },
        "score_delta": score_delta,
        "progression": "amelioration" if score_delta > 0 else ("degradation" if score_delta < 0 else "stable"),
        "nouveaux": nouveaux,
        "resolus": resolus,
        "persistants": persistants,
        "themes": themes,
    }
