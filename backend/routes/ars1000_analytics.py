"""
ARS 1000 Analytics & Export endpoints for Super Admin dashboard
"""
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone
from typing import Optional
import io
import logging

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/admin/analytics/ars1000", tags=["ARS 1000 Analytics"])


def get_db():
    from server import db
    return db


def get_current_user():
    from routes.auth import get_current_user as auth_get_current_user
    return auth_get_current_user


@router.get("/stats")
async def get_ars1000_stats(current_user=Depends(get_current_user())):
    """Aggregate ARS 1000 stats for the super admin dashboard"""
    if current_user.get("user_type") not in ["admin", "cooperative"]:
        raise HTTPException(status_code=403, detail="Acces refuse")

    db = get_db()

    # PDC Stats
    total_pdc = await db.pdc.count_documents({})
    pdc_by_status = {}
    for status in ["brouillon", "en_cours", "visite_terrain", "complete_agent", "valide"]:
        pdc_by_status[status] = await db.pdc.count_documents({"statut": status})

    # Average conformity
    pipeline_conf = [
        {"$match": {"pourcentage_conformite": {"$exists": True, "$gt": 0}}},
        {"$group": {"_id": None, "avg": {"$avg": "$pourcentage_conformite"}, "min": {"$min": "$pourcentage_conformite"}, "max": {"$max": "$pourcentage_conformite"}}}
    ]
    conf_result = await db.pdc.aggregate(pipeline_conf).to_list(1)
    avg_conformite = round(conf_result[0]["avg"], 1) if conf_result else 0
    min_conformite = round(conf_result[0]["min"], 1) if conf_result else 0
    max_conformite = round(conf_result[0]["max"], 1) if conf_result else 0

    # Conformity distribution
    conf_distribution = {"excellent": 0, "bon": 0, "moyen": 0, "faible": 0}
    async for pdc in db.pdc.find({"pourcentage_conformite": {"$exists": True}}, {"pourcentage_conformite": 1}):
        score = pdc.get("pourcentage_conformite", 0)
        if score >= 80:
            conf_distribution["excellent"] += 1
        elif score >= 60:
            conf_distribution["bon"] += 1
        elif score >= 40:
            conf_distribution["moyen"] += 1
        else:
            conf_distribution["faible"] += 1

    # PDC completion rate (fiches filled)
    fiches_completion = {"identification": 0, "epargne": 0, "menage_detail": 0, "exploitation": 0, "cultures": 0, "inventaire_arbres": 0, "arbres_ombrage_resume": 0, "materiel_detail": 0, "matrice_strategique_detail": 0, "programme_annuel": 0}
    total_pdcs_checked = 0
    async for pdc in db.pdc.find({}, {"identification": 1, "epargne": 1, "menage_detail": 1, "exploitation": 1, "cultures": 1, "inventaire_arbres": 1, "arbres_ombrage_resume": 1, "materiel_detail": 1, "matrice_strategique_detail": 1, "programme_annuel": 1}):
        total_pdcs_checked += 1
        ident = pdc.get("identification", {})
        if ident and (ident.get("nom") or ident.get("prenoms")):
            fiches_completion["identification"] += 1
        if pdc.get("epargne"):
            fiches_completion["epargne"] += 1
        if isinstance(pdc.get("menage_detail"), list) and len(pdc.get("menage_detail", [])) > 0:
            fiches_completion["menage_detail"] += 1
        if pdc.get("exploitation") and pdc["exploitation"].get("superficie_totale_ha"):
            fiches_completion["exploitation"] += 1
        if isinstance(pdc.get("cultures"), list) and len(pdc.get("cultures", [])) > 0:
            fiches_completion["cultures"] += 1
        if isinstance(pdc.get("inventaire_arbres"), list) and len(pdc.get("inventaire_arbres", [])) > 0:
            fiches_completion["inventaire_arbres"] += 1
        if pdc.get("arbres_ombrage_resume"):
            fiches_completion["arbres_ombrage_resume"] += 1
        if isinstance(pdc.get("materiel_detail"), list) and len(pdc.get("materiel_detail", [])) > 0:
            fiches_completion["materiel_detail"] += 1
        if isinstance(pdc.get("matrice_strategique_detail"), list) and len(pdc.get("matrice_strategique_detail", [])) > 0:
            fiches_completion["matrice_strategique_detail"] += 1
        if isinstance(pdc.get("programme_annuel"), list) and len(pdc.get("programme_annuel", [])) > 0:
            fiches_completion["programme_annuel"] += 1

    fiches_pct = {}
    for k, v in fiches_completion.items():
        fiches_pct[k] = round((v / total_pdcs_checked) * 100, 1) if total_pdcs_checked > 0 else 0

    # Harvest declarations
    total_declarations = await db.ars1000_declarations_recoltes.count_documents({})
    decl_by_status = {}
    for s in ["soumise", "validee", "rejetee"]:
        decl_by_status[s] = await db.ars1000_declarations_recoltes.count_documents({"statut": s})

    pipeline_harvest = [
        {"$match": {"statut": "validee"}},
        {"$group": {"_id": None, "total_kg": {"$sum": "$quantite_kg"}, "count": {"$sum": 1}}}
    ]
    harvest_result = await db.ars1000_declarations_recoltes.aggregate(pipeline_harvest).to_list(1)
    total_kg_validated = harvest_result[0]["total_kg"] if harvest_result else 0

    # Grade distribution
    pipeline_grade = [
        {"$group": {"_id": "$grade", "count": {"$sum": 1}}}
    ]
    grade_dist = {}
    async for g in db.ars1000_declarations_recoltes.aggregate(pipeline_grade):
        grade_dist[g["_id"] or "non_grade"] = g["count"]

    # Certifications
    total_certifications = await db.certification.count_documents({})
    cert_by_level = {}
    for lvl in ["bronze", "argent", "or"]:
        cert_by_level[lvl] = await db.certification.count_documents({"niveau": lvl})

    # Reclamations
    total_reclamations = await db.reclamations.count_documents({})
    recl_by_status = {}
    for s in ["ouverte", "en_cours", "resolue", "fermee"]:
        recl_by_status[s] = await db.reclamations.count_documents({"statut": s})

    # Agroforesterie summary
    total_arbres_inventories = 0
    total_ombrage = 0
    async for pdc in db.pdc.find({}, {"inventaire_arbres": 1, "arbres_ombrage": 1, "arbres_ombrage_resume": 1}):
        inv = pdc.get("inventaire_arbres", [])
        if isinstance(inv, list):
            total_arbres_inventories += len(inv)
        ao = pdc.get("arbres_ombrage", {})
        aor = pdc.get("arbres_ombrage_resume", {})
        total_ombrage += int(ao.get("nombre_total", 0) or 0) + int(aor.get("total", 0) or 0)

    # Agent visits
    pipeline_visits = [
        {"$match": {"last_visit_at": {"$exists": True}}},
        {"$count": "total"}
    ]
    visit_result = await db.pdc.aggregate(pipeline_visits).to_list(1)
    total_visites_terrain = visit_result[0]["total"] if visit_result else 0

    # Top cooperatives by PDC count
    pipeline_top_coop = [
        {"$group": {"_id": "$coop_id", "count": {"$sum": 1}, "avg_conf": {"$avg": "$pourcentage_conformite"}}},
        {"$sort": {"count": -1}},
        {"$limit": 10}
    ]
    top_coops = []
    async for c in db.pdc.aggregate(pipeline_top_coop):
        if c["_id"]:
            coop = await db.users.find_one({"_id": c["_id"]}, {"_id": 0, "cooperative_name": 1, "full_name": 1})
            name = (coop or {}).get("cooperative_name") or (coop or {}).get("full_name") or str(c["_id"])
            top_coops.append({"name": name, "pdc_count": c["count"], "avg_conformite": round(c.get("avg_conf", 0) or 0, 1)})

    return {
        "pdc": {
            "total": total_pdc,
            "by_status": pdc_by_status,
            "conformite_moyenne": avg_conformite,
            "conformite_min": min_conformite,
            "conformite_max": max_conformite,
            "distribution": conf_distribution,
            "fiches_completion_pct": fiches_pct,
            "visites_terrain": total_visites_terrain,
        },
        "recoltes": {
            "total_declarations": total_declarations,
            "by_status": decl_by_status,
            "total_kg_validated": total_kg_validated,
            "grade_distribution": grade_dist,
        },
        "certifications": {
            "total": total_certifications,
            "by_level": cert_by_level,
        },
        "reclamations": {
            "total": total_reclamations,
            "by_status": recl_by_status,
        },
        "agroforesterie": {
            "total_arbres_inventories": total_arbres_inventories,
            "total_ombrage": total_ombrage,
        },
        "top_cooperatives": top_coops,
    }


@router.get("/export/excel/{pdc_id}")
async def export_pdc_excel(pdc_id: str, current_user=Depends(get_current_user())):
    """Export a single farmer PDC to Excel with all 7 fiches"""
    from bson import ObjectId
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if current_user.get("user_type") not in ["admin", "cooperative"]:
        raise HTTPException(status_code=403, detail="Acces refuse")

    db = get_db()
    try:
        pdc = await db.pdc.find_one({"_id": ObjectId(pdc_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="ID invalide")

    if not pdc:
        raise HTTPException(status_code=404, detail="PDC introuvable")

    # Fetch farmer info
    farmer = await db.users.find_one({"_id": pdc.get("farmer_id")}, {"_id": 0, "full_name": 1, "phone_number": 1})
    farmer_name = (farmer or {}).get("full_name", "Planteur")

    wb = Workbook()
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    sub_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

    def style_cell(ws, row, col, value=""):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True)
        return cell

    # ===== FICHE 1: Identification =====
    ws1 = wb.active
    ws1.title = "F1-Identification"
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 35

    ident = pdc.get("identification", {})
    ws1.merge_cells('A1:B1')
    ws1.cell(row=1, column=1, value="FICHE 1 : IDENTIFICATION DU PRODUCTEUR").font = Font(bold=True, size=13)
    ws1.cell(row=1, column=1).fill = header_fill
    ws1.cell(row=1, column=1).font = header_font

    fields = [
        ("Nom", ident.get("nom", "")), ("Prenoms", ident.get("prenoms", "")),
        ("Contact (Tel)", ident.get("contact_tel") or ident.get("telephone", "")),
        ("Code National", ident.get("code_national") or ident.get("numero_identification", "")),
        ("Code Groupe", ident.get("code_groupe", "")),
        ("Nom Entite", ident.get("nom_entite", "")), ("Code Entite", ident.get("code_entite", "")),
        ("Delegation Regionale", ident.get("delegation_regionale", "")),
        ("Region", ident.get("region", "")), ("Departement", ident.get("department", "")),
        ("Sous-Prefecture", ident.get("sous_prefecture", "")),
        ("Village", ident.get("village", "")), ("Campement", ident.get("campement", "")),
        ("Genre", ident.get("genre", "")), ("Date de naissance", ident.get("date_naissance", "")),
        ("Statut foncier", ident.get("statut_foncier", "")),
    ]
    for i, (label, val) in enumerate(fields, start=2):
        style_cell(ws1, i, 1, label).font = Font(bold=True)
        style_cell(ws1, i, 2, str(val or ""))

    # Epargne section
    epargne = pdc.get("epargne", {})
    if epargne:
        r = len(fields) + 3
        ws1.merge_cells(f'A{r}:E{r}')
        ws1.cell(row=r, column=1, value="SITUATION DE L'EPARGNE").font = Font(bold=True, size=11)
        r += 1
        for c, h in enumerate(["Type", "Compte", "Argent", "Financement", "Montant"], 1):
            style_cell(ws1, r, c, h).font = Font(bold=True)
            ws1.cell(row=r, column=c).fill = sub_fill
        r += 1
        for cat in ["mobile_money", "microfinance", "banque", "autres"]:
            ep = epargne.get(cat, {})
            style_cell(ws1, r, 1, cat.replace("_", " ").title())
            style_cell(ws1, r, 2, "Oui" if ep.get("compte") else "Non")
            style_cell(ws1, r, 3, "Oui" if ep.get("argent_compte") else "Non")
            style_cell(ws1, r, 4, "Oui" if ep.get("financement") else "Non")
            style_cell(ws1, r, 5, str(ep.get("montant", "") or ""))
            r += 1

    # ===== FICHE 2: Menage =====
    ws2 = wb.create_sheet("F2-Menage")
    ws2.merge_cells('A1:I1')
    ws2.cell(row=1, column=1, value="FICHE 2 : SITUATION DU MENAGE").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill

    menage_headers = ["Membre", "Nombre", "A l'ecole", "Aucun", "Primaire", "Secondaire", "Universitaire", "Plein temps", "Occasionnel"]
    for c, h in enumerate(menage_headers, 1):
        style_cell(ws2, 2, c, h).font = Font(bold=True)
        ws2.cell(row=2, column=c).fill = sub_fill
        ws2.column_dimensions[chr(64 + c)].width = 14

    menage_data = pdc.get("menage_detail", [])
    if not isinstance(menage_data, list):
        menage_data = []
    for i, row in enumerate(menage_data, start=3):
        style_cell(ws2, i, 1, row.get("type", ""))
        for j, f in enumerate(["nombre", "a_ecole", "aucun", "primaire", "secondaire", "universitaire", "plein_temps", "occasionnel"], 2):
            style_cell(ws2, i, j, str(row.get(f, "") or ""))

    # ===== FICHE 3: Exploitation =====
    ws3 = wb.create_sheet("F3-Exploitation")
    ws3.merge_cells('A1:F1')
    ws3.cell(row=1, column=1, value="FICHE 3 : DESCRIPTION DE L'EXPLOITATION").font = header_font
    ws3.cell(row=1, column=1).fill = header_fill

    exploitation = pdc.get("exploitation", {})
    exp_fields = [
        ("Superficie totale (ha)", exploitation.get("superficie_totale_ha", "")),
        ("Superficie cultivee (ha)", exploitation.get("superficie_cultivee_ha", "")),
        ("Superficie foret (ha)", exploitation.get("superficie_foret_ha", "")),
        ("Superficie jachere (ha)", exploitation.get("superficie_jachere_ha", "")),
        ("Source d'eau", exploitation.get("source_eau", "")),
        ("Type source", exploitation.get("type_source_eau", "")),
    ]
    for i, (label, val) in enumerate(exp_fields, start=2):
        style_cell(ws3, i, 1, label).font = Font(bold=True)
        style_cell(ws3, i, 2, str(val or ""))
    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 20

    # Cultures table
    cultures = pdc.get("cultures", [])
    if isinstance(cultures, list) and cultures:
        r = len(exp_fields) + 4
        ws3.merge_cells(f'A{r}:F{r}')
        ws3.cell(row=r, column=1, value="CULTURES").font = Font(bold=True, size=11)
        r += 1
        cult_h = ["Culture", "Superficie (ha)", "Annee creation", "Source materiel", "Production (kg)", "Revenu (FCFA)"]
        for c, h in enumerate(cult_h, 1):
            style_cell(ws3, r, c, h).font = Font(bold=True)
            ws3.cell(row=r, column=c).fill = sub_fill
        r += 1
        for cult in cultures:
            style_cell(ws3, r, 1, cult.get("nom", ""))
            style_cell(ws3, r, 2, str(cult.get("superficie", "") or ""))
            style_cell(ws3, r, 3, str(cult.get("annee_creation", "") or ""))
            style_cell(ws3, r, 4, str(cult.get("source_materiel", "") or ""))
            style_cell(ws3, r, 5, str(cult.get("production_kg", "") or ""))
            style_cell(ws3, r, 6, str(cult.get("revenu_fcfa", "") or ""))
            r += 1

    # ===== FICHE 4: Inventaire Arbres =====
    ws4 = wb.create_sheet("F4-Inventaire")
    ws4.merge_cells('A1:F1')
    ws4.cell(row=1, column=1, value="FICHE 4 : INVENTAIRE DES ARBRES").font = header_font
    ws4.cell(row=1, column=1).fill = header_fill

    inv_h = ["N", "Nom botanique", "Nom local", "Circonference (cm)", "Origine", "Decision"]
    for c, h in enumerate(inv_h, 1):
        style_cell(ws4, 2, c, h).font = Font(bold=True)
        ws4.cell(row=2, column=c).fill = sub_fill
        ws4.column_dimensions[chr(64 + c)].width = 18

    inv_arbres = pdc.get("inventaire_arbres", [])
    if isinstance(inv_arbres, list):
        for i, arb in enumerate(inv_arbres, start=3):
            style_cell(ws4, i, 1, i - 2)
            style_cell(ws4, i, 2, arb.get("nom_botanique", ""))
            style_cell(ws4, i, 3, arb.get("nom_local", ""))
            style_cell(ws4, i, 4, str(arb.get("circonference", "") or ""))
            style_cell(ws4, i, 5, arb.get("origine", ""))
            style_cell(ws4, i, 6, arb.get("decision", ""))

    # ===== FICHE 5: Arbres Ombrage =====
    ws5 = wb.create_sheet("F5-Ombrage")
    ws5.merge_cells('A1:B1')
    ws5.cell(row=1, column=1, value="FICHE 5 : ARBRES D'OMBRAGE").font = header_font
    ws5.cell(row=1, column=1).fill = header_fill
    ws5.column_dimensions['A'].width = 30
    ws5.column_dimensions['B'].width = 15

    ao_resume = pdc.get("arbres_ombrage_resume", {}) or {}
    ao_data = pdc.get("arbres_ombrage", {}) or {}
    for i, (label, val) in enumerate([
        ("Strate 1 (basse)", ao_resume.get("strate1", ao_data.get("strate_basse", ""))),
        ("Strate 2 (moyenne)", ao_resume.get("strate2", ao_data.get("strate_moyenne", ""))),
        ("Strate 3 (haute)", ao_resume.get("strate3", ao_data.get("strate_haute", ""))),
        ("TOTAL", ao_resume.get("total", ao_data.get("nombre_total", ""))),
    ], start=2):
        style_cell(ws5, i, 1, label).font = Font(bold=True)
        style_cell(ws5, i, 2, str(val or ""))

    # ===== FICHE 6: Materiel =====
    ws6 = wb.create_sheet("F6-Materiel")
    ws6.merge_cells('A1:H1')
    ws6.cell(row=1, column=1, value="FICHE 6 : MATERIEL AGRICOLE").font = header_font
    ws6.cell(row=1, column=1).fill = header_fill

    mat_h = ["Type", "Designation", "Quantite", "Annee", "Cout", "Bon", "Acceptable", "Mauvais"]
    for c, h in enumerate(mat_h, 1):
        style_cell(ws6, 2, c, h).font = Font(bold=True)
        ws6.cell(row=2, column=c).fill = sub_fill
        ws6.column_dimensions[chr(64 + c)].width = 16

    mat_detail = pdc.get("materiel_detail", [])
    if isinstance(mat_detail, list):
        for i, m in enumerate(mat_detail, start=3):
            style_cell(ws6, i, 1, m.get("type", ""))
            style_cell(ws6, i, 2, m.get("designation", ""))
            style_cell(ws6, i, 3, str(m.get("quantite", "") or ""))
            style_cell(ws6, i, 4, str(m.get("annee", "") or ""))
            style_cell(ws6, i, 5, str(m.get("cout", "") or ""))
            style_cell(ws6, i, 6, str(m.get("bon", "") or ""))
            style_cell(ws6, i, 7, str(m.get("acceptable", "") or ""))
            style_cell(ws6, i, 8, str(m.get("mauvais", "") or ""))

    # ===== FICHE 7: Planification =====
    ws7 = wb.create_sheet("F7-Planification")
    ws7.merge_cells('A1:K1')
    ws7.cell(row=1, column=1, value="FICHE 7a : MATRICE STRATEGIQUE").font = header_font
    ws7.cell(row=1, column=1).fill = header_fill

    strat_h = ["Axe", "Objectifs", "Activites", "Cout", "A1", "A2", "A3", "A4", "A5", "Responsable", "Partenaires"]
    for c, h in enumerate(strat_h, 1):
        style_cell(ws7, 2, c, h).font = Font(bold=True)
        ws7.cell(row=2, column=c).fill = sub_fill

    strat = pdc.get("matrice_strategique_detail", [])
    r = 3
    if isinstance(strat, list):
        for s in strat:
            style_cell(ws7, r, 1, s.get("axe", ""))
            style_cell(ws7, r, 2, s.get("objectifs", ""))
            style_cell(ws7, r, 3, s.get("activites", ""))
            style_cell(ws7, r, 4, str(s.get("cout", "") or ""))
            for j, a in enumerate(["a1", "a2", "a3", "a4", "a5"], 5):
                style_cell(ws7, r, j, "X" if s.get(a) else "")
            style_cell(ws7, r, 10, s.get("responsable", ""))
            style_cell(ws7, r, 11, s.get("partenaires", ""))
            r += 1

    # Programme annuel
    r += 2
    ws7.merge_cells(f'A{r}:K{r}')
    ws7.cell(row=r, column=1, value="FICHE 7b : PROGRAMME ANNUEL D'ACTION").font = header_font
    ws7.cell(row=r, column=1).fill = header_fill
    r += 1
    prog_h = ["Axe", "Activite", "Sous-activite", "Indicateur", "T1", "T2", "T3", "T4", "Execution", "Appui", "Cout"]
    for c, h in enumerate(prog_h, 1):
        style_cell(ws7, r, c, h).font = Font(bold=True)
        ws7.cell(row=r, column=c).fill = sub_fill
    r += 1

    prog = pdc.get("programme_annuel", [])
    if isinstance(prog, list):
        for p in prog:
            style_cell(ws7, r, 1, p.get("axe", ""))
            style_cell(ws7, r, 2, p.get("activite", ""))
            style_cell(ws7, r, 3, p.get("sous_activite", ""))
            style_cell(ws7, r, 4, p.get("indicateur", ""))
            for j, t in enumerate(["t1", "t2", "t3", "t4"], 5):
                style_cell(ws7, r, j, "X" if p.get(t) else "")
            style_cell(ws7, r, 9, p.get("execution", ""))
            style_cell(ws7, r, 10, p.get("appui", ""))
            style_cell(ws7, r, 11, str(p.get("cout", "") or ""))
            r += 1

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_name = farmer_name.replace(" ", "_").replace("/", "_")
    filename = f"PDC_ARS1000_{safe_name}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
