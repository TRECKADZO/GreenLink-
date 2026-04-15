"""
Module Audit Interne & Non-conformites - ARS 1000
GreenLink Agritech

Checklist d'audit digitale, gestion NC, rapports, revue de direction.
Conforme a la checklist Excel officielle (feuilles ARS 1000-1, ARS 1000-2, Resultats).
"""

from fastapi import APIRouter, HTTPException, Depends, Query, status
from fastapi.responses import StreamingResponse
from typing import Optional, List
from pydantic import BaseModel, Field
from datetime import datetime, timezone
from bson import ObjectId
import logging
import io
import json
import uuid
import os
from pathlib import Path

from database import db
from routes.auth import get_current_user
from routes.cooperative import verify_cooperative

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/audit", tags=["Audit Interne ARS 1000"])

SEED_FILE = Path(__file__).parent.parent / "data_seed" / "ars1000_checklist_data.json"

# ============= PYDANTIC MODELS =============

class AuditSessionCreate(BaseModel):
    titre: str = "Audit interne ARS 1000"
    campagne: str = "2025-2026"
    auditeur: str = ""
    date_debut: str = ""
    niveau_certification: str = "Bronze"
    notes: str = ""

class ChecklistItemUpdate(BaseModel):
    preuves_audit: Optional[str] = None
    constatation: Optional[str] = None
    conformite: Optional[str] = None  # C / NC / NA
    notes_auditeur: Optional[str] = None

class NonConformiteCreate(BaseModel):
    audit_session_id: str
    checklist_item_id: str = ""
    clause: str = ""
    norme: str = "ARS 1000-1"
    constatation: str = ""
    type_nc: str = "Mineure"  # Majeure / Mineure
    cause_profonde: str = ""
    corrections: str = ""
    actions_correctives: str = ""
    responsable: str = ""
    date_resolution_prevue: str = ""

class NonConformiteUpdate(BaseModel):
    constatation: Optional[str] = None
    type_nc: Optional[str] = None
    cause_profonde: Optional[str] = None
    corrections: Optional[str] = None
    actions_correctives: Optional[str] = None
    responsable: Optional[str] = None
    date_resolution_prevue: Optional[str] = None
    statut: Optional[str] = None  # ouvert / en_cours / resolu
    date_resolution: Optional[str] = None

class RevueDirectionCreate(BaseModel):
    audit_session_id: str = ""
    date_revue: str = ""
    participants: str = ""
    points_examines: str = ""
    decisions: str = ""
    actions: str = ""
    prochaine_revue: str = ""


def get_coop_id(user):
    return str(user.get("coop_id") or user.get("_id") or user.get("id", ""))


# ============= SEED CHECKLIST =============

async def seed_checklist_if_needed(coop_id: str, session_id: str, norme: str):
    """Charge les items de la checklist depuis le JSON si pas encore en base"""
    existing = await db.audit_checklist_items.count_documents({
        "session_id": session_id, "norme": norme
    })
    if existing > 0:
        return existing

    if not SEED_FILE.exists():
        return 0

    with open(SEED_FILE, 'r') as f:
        data = json.load(f)

    key = "ars1000_1" if norme == "ARS 1000-1" else "ars1000_2"
    items = data.get(key, [])

    docs = []
    for idx, item in enumerate(items):
        docs.append({
            "item_id": str(uuid.uuid4()),
            "session_id": session_id,
            "coop_id": coop_id,
            "norme": norme,
            "ordre": idx + 1,
            "clause": item.get("clause", ""),
            "section": item.get("section", ""),
            "titre": item.get("titre", ""),
            "contenu_detaille": item.get("contenu_detaille", ""),
            "resume": item.get("resume", ""),
            "moyens": item.get("moyens", ""),
            "matieres": item.get("matieres", ""),
            "precision": item.get("precision", ""),
            "niveau": item.get("niveau", "").strip(),
            "type_exigence": item.get("type_exigence", ""),
            "cible": item.get("cible", ""),
            "etape": item.get("etape", ""),
            "preuves_audit": "",
            "constatation": "",
            "conformite": "",
            "notes_auditeur": "",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        })

    if docs:
        await db.audit_checklist_items.insert_many(docs)

    return len(docs)


# ============= AUDIT SESSIONS =============

@router.post("/sessions")
async def create_audit_session(data: AuditSessionCreate, current_user: dict = Depends(get_current_user)):
    """Creer une nouvelle session d'audit interne"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    session_id = str(uuid.uuid4())
    doc = {
        "session_id": session_id,
        "coop_id": coop_id,
        "titre": data.titre,
        "campagne": data.campagne,
        "auditeur": data.auditeur,
        "date_debut": data.date_debut or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "niveau_certification": data.niveau_certification,
        "notes": data.notes,
        "statut": "en_cours",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    await db.audit_sessions.insert_one(doc)
    doc.pop("_id", None)

    count1 = await seed_checklist_if_needed(coop_id, session_id, "ARS 1000-1")
    count2 = await seed_checklist_if_needed(coop_id, session_id, "ARS 1000-2")

    return {
        "status": "success",
        "session": doc,
        "checklist_items_created": {"ars1000_1": count1, "ars1000_2": count2},
    }


@router.get("/sessions")
async def list_audit_sessions(current_user: dict = Depends(get_current_user)):
    """Lister les sessions d'audit de la cooperative"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    sessions = await db.audit_sessions.find(
        {"coop_id": coop_id}, {"_id": 0}
    ).sort("created_at", -1).to_list(100)

    return {"sessions": sessions}


@router.get("/sessions/{session_id}")
async def get_audit_session(session_id: str, current_user: dict = Depends(get_current_user)):
    """Detail d'une session d'audit"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    session = await db.audit_sessions.find_one(
        {"session_id": session_id, "coop_id": coop_id}, {"_id": 0}
    )
    if not session:
        raise HTTPException(status_code=404, detail="Session non trouvee")

    return {"session": session}


@router.put("/sessions/{session_id}/close")
async def close_audit_session(session_id: str, current_user: dict = Depends(get_current_user)):
    """Cloturer une session d'audit"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    result = await db.audit_sessions.update_one(
        {"session_id": session_id, "coop_id": coop_id},
        {"$set": {"statut": "cloture", "date_cloture": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Session non trouvee")

    return {"status": "success", "message": "Session cloturee"}


# ============= CHECKLIST =============

@router.get("/sessions/{session_id}/checklist")
async def get_checklist(
    session_id: str,
    current_user: dict = Depends(get_current_user),
    norme: str = "ARS 1000-1",
    niveau: Optional[str] = None,
    cible: Optional[str] = None,
    etape: Optional[str] = None,
    conformite: Optional[str] = None,
):
    """Recuperer la checklist d'audit avec filtres"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    query = {"session_id": session_id, "coop_id": coop_id, "norme": norme}
    if niveau:
        query["niveau"] = {"$regex": niveau, "$options": "i"}
    if cible:
        query["cible"] = {"$regex": cible, "$options": "i"}
    if etape:
        query["etape"] = {"$regex": etape, "$options": "i"}
    if conformite:
        query["conformite"] = conformite

    items = await db.audit_checklist_items.find(
        query, {"_id": 0}
    ).sort("ordre", 1).to_list(500)

    total = len(items)
    conformes = sum(1 for i in items if i.get("conformite") == "C")
    non_conformes = sum(1 for i in items if i.get("conformite") == "NC")
    na = sum(1 for i in items if i.get("conformite") == "NA")
    non_evalues = total - conformes - non_conformes - na
    applicable = total - na
    taux = round(conformes / applicable * 100, 1) if applicable > 0 else 0

    return {
        "items": items,
        "stats": {
            "total": total,
            "conformes": conformes,
            "non_conformes": non_conformes,
            "na": na,
            "non_evalues": non_evalues,
            "taux_conformite": taux,
        },
    }


@router.put("/checklist/{item_id}")
async def update_checklist_item(
    item_id: str,
    update: ChecklistItemUpdate,
    current_user: dict = Depends(get_current_user),
):
    """Mettre a jour un item de la checklist (conformite, preuves, constatation)"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Aucune donnee a mettre a jour")

    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["updated_by"] = str(current_user.get("_id", ""))

    # Add to history
    history_entry = {
        "date": datetime.now(timezone.utc).isoformat(),
        "user": current_user.get("full_name", ""),
        "changes": update_data,
    }

    result = await db.audit_checklist_items.update_one(
        {"item_id": item_id, "coop_id": coop_id},
        {"$set": update_data, "$push": {"historique": history_entry}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Item non trouve")

    item = await db.audit_checklist_items.find_one({"item_id": item_id}, {"_id": 0})
    return {"status": "success", "item": item}


# ============= NON-CONFORMITES =============

@router.post("/non-conformites")
async def create_nc(nc: NonConformiteCreate, current_user: dict = Depends(get_current_user)):
    """Creer une non-conformite"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    existing = await db.audit_non_conformites.count_documents({"coop_id": coop_id})
    nc_number = existing + 1

    doc = {
        "nc_id": str(uuid.uuid4()),
        "nc_number": nc_number,
        "coop_id": coop_id,
        "audit_session_id": nc.audit_session_id,
        "checklist_item_id": nc.checklist_item_id,
        "clause": nc.clause,
        "norme": nc.norme,
        "constatation": nc.constatation,
        "type_nc": nc.type_nc,
        "cause_profonde": nc.cause_profonde,
        "corrections": nc.corrections,
        "actions_correctives": nc.actions_correctives,
        "responsable": nc.responsable,
        "date_resolution_prevue": nc.date_resolution_prevue,
        "date_resolution": "",
        "statut": "ouvert",
        "created_by": current_user.get("full_name", ""),
        "historique": [],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    await db.audit_non_conformites.insert_one(doc)
    doc.pop("_id", None)

    return {"status": "success", "nc": doc}


@router.get("/non-conformites")
async def list_non_conformites(
    current_user: dict = Depends(get_current_user),
    session_id: Optional[str] = None,
    statut: Optional[str] = None,
    type_nc: Optional[str] = None,
):
    """Lister les non-conformites"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    query = {"coop_id": coop_id}
    if session_id:
        query["audit_session_id"] = session_id
    if statut:
        query["statut"] = statut
    if type_nc:
        query["type_nc"] = type_nc

    ncs = await db.audit_non_conformites.find(
        query, {"_id": 0}
    ).sort("nc_number", 1).to_list(500)

    total = len(ncs)
    ouvertes = sum(1 for n in ncs if n.get("statut") == "ouvert")
    en_cours = sum(1 for n in ncs if n.get("statut") == "en_cours")
    resolues = sum(1 for n in ncs if n.get("statut") == "resolu")
    majeures = sum(1 for n in ncs if n.get("type_nc") == "Majeure")
    mineures = sum(1 for n in ncs if n.get("type_nc") == "Mineure")

    return {
        "non_conformites": ncs,
        "stats": {
            "total": total,
            "ouvertes": ouvertes,
            "en_cours": en_cours,
            "resolues": resolues,
            "majeures": majeures,
            "mineures": mineures,
        },
    }


@router.put("/non-conformites/{nc_id}")
async def update_nc(nc_id: str, update: NonConformiteUpdate, current_user: dict = Depends(get_current_user)):
    """Mettre a jour une non-conformite"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()

    history_entry = {
        "date": datetime.now(timezone.utc).isoformat(),
        "user": current_user.get("full_name", ""),
        "changes": update_data,
    }

    result = await db.audit_non_conformites.update_one(
        {"nc_id": nc_id, "coop_id": coop_id},
        {"$set": update_data, "$push": {"historique": history_entry}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="NC non trouvee")

    nc = await db.audit_non_conformites.find_one({"nc_id": nc_id}, {"_id": 0})
    return {"status": "success", "nc": nc}


# ============= DASHBOARD CONFORMITE =============

@router.get("/dashboard")
async def get_audit_dashboard(
    current_user: dict = Depends(get_current_user),
    session_id: Optional[str] = None,
):
    """Tableau de bord de conformite"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    # Get latest session if not specified
    if not session_id:
        latest = await db.audit_sessions.find_one(
            {"coop_id": coop_id}, {"_id": 0}, sort=[("created_at", -1)]
        )
        if latest:
            session_id = latest["session_id"]

    if not session_id:
        return {
            "has_session": False,
            "message": "Aucune session d'audit. Creez-en une pour commencer.",
        }

    # Compute stats for both normes
    results = {}
    for norme in ["ARS 1000-1", "ARS 1000-2"]:
        items = await db.audit_checklist_items.find(
            {"session_id": session_id, "coop_id": coop_id, "norme": norme}, {"_id": 0}
        ).to_list(500)

        total = len(items)
        conformes = sum(1 for i in items if i.get("conformite") == "C")
        non_conformes = sum(1 for i in items if i.get("conformite") == "NC")
        na = sum(1 for i in items if i.get("conformite") == "NA")
        non_evalues = total - conformes - non_conformes - na
        applicable = total - na
        taux = round(conformes / applicable * 100, 1) if applicable > 0 else 0

        results[norme] = {
            "total": total,
            "conformes": conformes,
            "non_conformes": non_conformes,
            "na": na,
            "non_evalues": non_evalues,
            "taux_conformite": taux,
        }

    # Global stats
    total_all = sum(r["total"] for r in results.values())
    conformes_all = sum(r["conformes"] for r in results.values())
    nc_all = sum(r["non_conformes"] for r in results.values())
    na_all = sum(r["na"] for r in results.values())
    applicable_all = total_all - na_all
    taux_global = round(conformes_all / applicable_all * 100, 1) if applicable_all > 0 else 0

    # NC stats
    ncs = await db.audit_non_conformites.find(
        {"coop_id": coop_id, "audit_session_id": session_id}, {"_id": 0}
    ).to_list(500)
    nc_ouvertes = sum(1 for n in ncs if n.get("statut") == "ouvert")
    nc_en_cours = sum(1 for n in ncs if n.get("statut") == "en_cours")
    nc_resolues = sum(1 for n in ncs if n.get("statut") == "resolu")

    # By section
    sections_pipeline = [
        {"$match": {"session_id": session_id, "coop_id": coop_id}},
        {"$group": {
            "_id": {"norme": "$norme", "section": "$section"},
            "total": {"$sum": 1},
            "conformes": {"$sum": {"$cond": [{"$eq": ["$conformite", "C"]}, 1, 0]}},
            "non_conformes": {"$sum": {"$cond": [{"$eq": ["$conformite", "NC"]}, 1, 0]}},
        }},
        {"$sort": {"_id.norme": 1, "_id.section": 1}},
    ]
    sections = await db.audit_checklist_items.aggregate(sections_pipeline).to_list(100)
    by_section = []
    for s in sections:
        if s["_id"]["section"]:
            applicable = s["total"] - 0  # NA not tracked in aggregation for simplicity
            by_section.append({
                "norme": s["_id"]["norme"],
                "section": s["_id"]["section"],
                "total": s["total"],
                "conformes": s["conformes"],
                "non_conformes": s["non_conformes"],
                "taux": round(s["conformes"] / s["total"] * 100, 1) if s["total"] > 0 else 0,
            })

    session = await db.audit_sessions.find_one(
        {"session_id": session_id, "coop_id": coop_id}, {"_id": 0}
    )

    return {
        "has_session": True,
        "session": session,
        "resultats": results,
        "global": {
            "total": total_all,
            "conformes": conformes_all,
            "non_conformes": nc_all,
            "na": na_all,
            "non_evalues": total_all - conformes_all - nc_all - na_all,
            "taux_conformite": taux_global,
        },
        "non_conformites": {
            "total": len(ncs),
            "ouvertes": nc_ouvertes,
            "en_cours": nc_en_cours,
            "resolues": nc_resolues,
        },
        "par_section": by_section,
    }


# ============= RESULTATS (comme feuille Excel) =============

@router.get("/sessions/{session_id}/resultats")
async def get_audit_results(session_id: str, current_user: dict = Depends(get_current_user)):
    """Resultats d'audit au format de la feuille Excel 'Resultats d'audit'"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    results = {}
    for norme in ["ARS 1000-1", "ARS 1000-2"]:
        items = await db.audit_checklist_items.find(
            {"session_id": session_id, "coop_id": coop_id, "norme": norme}, {"_id": 0}
        ).to_list(500)
        total = len(items)
        c = sum(1 for i in items if i.get("conformite") == "C")
        nc = sum(1 for i in items if i.get("conformite") == "NC")
        na = sum(1 for i in items if i.get("conformite") == "NA")
        applicable = total - na
        results[norme] = {
            "total": total,
            "conformes": {"nombre": c, "taux": round(c / applicable * 100, 1) if applicable > 0 else 0},
            "non_conformes": {"nombre": nc, "taux": round(nc / applicable * 100, 1) if applicable > 0 else 0},
            "na": na,
            "pourcentage_conformite": round(c / applicable * 100, 1) if applicable > 0 else 0,
        }

    # NC table
    ncs = await db.audit_non_conformites.find(
        {"coop_id": coop_id, "audit_session_id": session_id}, {"_id": 0}
    ).sort("nc_number", 1).to_list(500)

    return {"resultats": results, "tableau_nc": ncs}


# ============= REVUE DE DIRECTION =============

@router.post("/revue-direction")
async def create_revue(data: RevueDirectionCreate, current_user: dict = Depends(get_current_user)):
    """Creer un PV de revue de direction (clause 9.3)"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    doc = {
        "revue_id": str(uuid.uuid4()),
        "coop_id": coop_id,
        "audit_session_id": data.audit_session_id,
        "date_revue": data.date_revue or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "participants": data.participants,
        "points_examines": data.points_examines,
        "decisions": data.decisions,
        "actions": data.actions,
        "prochaine_revue": data.prochaine_revue,
        "created_by": current_user.get("full_name", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }

    await db.audit_revues_direction.insert_one(doc)
    doc.pop("_id", None)
    return {"status": "success", "revue": doc}


@router.get("/revue-direction")
async def list_revues(current_user: dict = Depends(get_current_user)):
    """Lister les revues de direction"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    revues = await db.audit_revues_direction.find(
        {"coop_id": coop_id}, {"_id": 0}
    ).sort("created_at", -1).to_list(100)

    return {"revues": revues}


# ============= EXPORT PDF =============

@router.get("/sessions/{session_id}/export/pdf")
async def export_audit_pdf(session_id: str, current_user: dict = Depends(get_current_user)):
    """Exporter le rapport d'audit complet en PDF"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    session = await db.audit_sessions.find_one(
        {"session_id": session_id, "coop_id": coop_id}, {"_id": 0}
    )
    if not session:
        raise HTTPException(status_code=404, detail="Session non trouvee")

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle("Title", parent=styles["Title"], fontSize=18, textColor=colors.HexColor("#1A3622"))
    h2_style = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=14, textColor=colors.HexColor("#1A3622"))

    elements.append(Paragraph("RAPPORT D'AUDIT INTERNE ARS 1000", title_style))
    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph(f"Session: {session.get('titre', '')}", styles["Normal"]))
    elements.append(Paragraph(f"Campagne: {session.get('campagne', '')} | Auditeur: {session.get('auditeur', '')}", styles["Normal"]))
    elements.append(Paragraph(f"Date: {session.get('date_debut', '')} | Niveau: {session.get('niveau_certification', '')}", styles["Normal"]))
    elements.append(Spacer(1, 1*cm))

    # Results for each norme
    for norme in ["ARS 1000-1", "ARS 1000-2"]:
        elements.append(Paragraph(f"Resultats {norme}", h2_style))
        elements.append(Spacer(1, 0.3*cm))

        items = await db.audit_checklist_items.find(
            {"session_id": session_id, "coop_id": coop_id, "norme": norme}, {"_id": 0}
        ).sort("ordre", 1).to_list(500)

        total = len(items)
        c_count = sum(1 for i in items if i.get("conformite") == "C")
        nc_count = sum(1 for i in items if i.get("conformite") == "NC")
        na_count = sum(1 for i in items if i.get("conformite") == "NA")
        applicable = total - na_count
        taux = round(c_count / applicable * 100, 1) if applicable > 0 else 0

        kpi_data = [
            ["Total exigences", str(total)],
            ["Conformes", f"{c_count} ({taux}%)"],
            ["Non-conformes", str(nc_count)],
            ["Non applicable", str(na_count)],
        ]
        kpi_table = Table(kpi_data, colWidths=[8*cm, 6*cm])
        kpi_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E8F0EA")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E5E0")),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(kpi_table)
        elements.append(Spacer(1, 0.5*cm))

        # Checklist items with findings
        evaluated = [i for i in items if i.get("conformite")]
        if evaluated:
            headers = ["Clause", "Titre", "Conformite", "Constatation"]
            rows = [headers]
            for i in evaluated:
                rows.append([
                    str(i.get("clause", ""))[:10],
                    str(i.get("titre", ""))[:40],
                    i.get("conformite", ""),
                    str(i.get("constatation", ""))[:50],
                ])

            cl_table = Table(rows, colWidths=[2*cm, 7*cm, 2*cm, 5.5*cm])
            cl_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A3622")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E5E0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
            ]))
            elements.append(cl_table)

        elements.append(PageBreak())

    # Non-conformites table
    elements.append(Paragraph("Tableau des Non-conformites", h2_style))
    elements.append(Spacer(1, 0.3*cm))

    ncs = await db.audit_non_conformites.find(
        {"coop_id": coop_id, "audit_session_id": session_id}, {"_id": 0}
    ).sort("nc_number", 1).to_list(500)

    if ncs:
        nc_headers = ["N", "Clause", "Type", "Constatation", "Statut"]
        nc_rows = [nc_headers]
        for n in ncs:
            nc_rows.append([
                str(n.get("nc_number", "")),
                str(n.get("clause", "")),
                n.get("type_nc", ""),
                str(n.get("constatation", ""))[:50],
                n.get("statut", ""),
            ])
        nc_table = Table(nc_rows, colWidths=[1.2*cm, 2*cm, 2.5*cm, 8*cm, 2.5*cm])
        nc_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A3622")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E5E0")),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(nc_table)
    else:
        elements.append(Paragraph("Aucune non-conformite enregistree.", styles["Normal"]))

    doc.build(elements)
    buf.seek(0)

    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=rapport_audit_{session_id[:8]}.pdf"},
    )


# ============= EXPORT EXCEL =============

@router.get("/sessions/{session_id}/export/excel")
async def export_audit_excel(session_id: str, current_user: dict = Depends(get_current_user)):
    """Exporter la checklist complete en Excel (format identique a l'original)"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    header_fill = PatternFill(start_color="1A3622", end_color="1A3622", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=9)
    nc_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    c_fill = PatternFill(start_color="E8F0EA", end_color="E8F0EA", fill_type="solid")
    na_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for idx, norme in enumerate(["ARS 1000-1", "ARS 1000-2"]):
        if idx == 0:
            ws = wb.active
            ws.title = norme
        else:
            ws = wb.create_sheet(norme)

        headers = ["Clause", "Titre", "Contenu detaille", "Resume", "Moyens", "Matieres",
                    "Precision de conformite", "Niveau", "Preuves d'audit", "Constatation",
                    "Conformite (C/NC/NA)", "Type d'exigence", "Cible", "Etape"]

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

        items = await db.audit_checklist_items.find(
            {"session_id": session_id, "coop_id": coop_id, "norme": norme}, {"_id": 0}
        ).sort("ordre", 1).to_list(500)

        for row_idx, item in enumerate(items, 2):
            vals = [
                item.get("clause", ""), item.get("titre", ""),
                item.get("contenu_detaille", ""), item.get("resume", ""),
                item.get("moyens", ""), item.get("matieres", ""),
                item.get("precision", ""), item.get("niveau", ""),
                item.get("preuves_audit", ""), item.get("constatation", ""),
                item.get("conformite", ""), item.get("type_exigence", ""),
                item.get("cible", ""), item.get("etape", ""),
            ]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col, value=v)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Color row by conformite
            conf = item.get("conformite", "")
            if conf == "NC":
                for col in range(1, 15):
                    ws.cell(row=row_idx, column=col).fill = nc_fill
            elif conf == "C":
                for col in range(1, 15):
                    ws.cell(row=row_idx, column=col).fill = c_fill
            elif conf == "NA":
                for col in range(1, 15):
                    ws.cell(row=row_idx, column=col).fill = na_fill

        # Adjust widths
        widths = [8, 30, 40, 30, 12, 25, 30, 10, 25, 25, 12, 12, 20, 8]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i) if i < 27 else 'A'].width = w

    # Resultats sheet
    ws_res = wb.create_sheet("Resultats d'audit")
    ws_res["A1"] = "RESULTATS D'AUDIT ARS 1000"
    ws_res["A1"].font = Font(bold=True, size=14, color="1A3622")
    ws_res.merge_cells("A1:F1")

    ws_res["A3"] = ""
    ws_res["B3"] = "ARS 1000-1"
    ws_res["D3"] = "ARS 1000-2"
    ws_res["B4"] = "Nombre"
    ws_res["C4"] = "Taux (%)"
    ws_res["D4"] = "Nombre"
    ws_res["E4"] = "Taux (%)"

    for col in range(1, 6):
        ws_res.cell(row=3, column=col).font = Font(bold=True)
        ws_res.cell(row=4, column=col).font = Font(bold=True)

    for norme_idx, norme in enumerate(["ARS 1000-1", "ARS 1000-2"]):
        items = await db.audit_checklist_items.find(
            {"session_id": session_id, "coop_id": coop_id, "norme": norme}, {"_id": 0}
        ).to_list(500)
        total = len(items)
        c_n = sum(1 for i in items if i.get("conformite") == "C")
        nc_n = sum(1 for i in items if i.get("conformite") == "NC")
        na_n = sum(1 for i in items if i.get("conformite") == "NA")
        app = total - na_n

        col_offset = norme_idx * 2 + 2
        ws_res.cell(row=5, column=1, value="Conformes")
        ws_res.cell(row=5, column=col_offset, value=c_n)
        ws_res.cell(row=5, column=col_offset + 1, value=round(c_n / app * 100, 1) if app > 0 else 0)
        ws_res.cell(row=6, column=1, value="Non-conformes")
        ws_res.cell(row=6, column=col_offset, value=nc_n)
        ws_res.cell(row=6, column=col_offset + 1, value=round(nc_n / app * 100, 1) if app > 0 else 0)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=audit_ars1000_{session_id[:8]}.xlsx"},
    )
