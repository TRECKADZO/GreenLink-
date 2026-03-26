"""
Cooperative Members Management Routes
GreenLink Agritech - Côte d'Ivoire
"""

from fastapi import APIRouter, HTTPException, Depends, Query, Request
from fastapi.responses import StreamingResponse
from typing import List, Optional
from datetime import datetime, timezone
from bson import ObjectId
import hashlib
import logging
import io

from database import db
from routes.auth import get_current_user
from routes.cooperative import verify_cooperative, coop_id_query, CoopMemberCreate
from routes.ussd import generate_farmer_code
from slowapi import Limiter
from slowapi.util import get_remote_address

limiter = Limiter(key_func=get_remote_address)

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/cooperative", tags=["Cooperative Members"])


def hash_pin(pin: str) -> str:
    return hashlib.sha256(pin.encode()).hexdigest()


@router.get("/members")
async def get_coop_members(
    current_user: dict = Depends(get_current_user),
    status_filter: Optional[str] = Query(None, alias="status"),
    village: Optional[str] = None,
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 50
):
    """Liste des membres de la coopérative"""
    user_type = current_user.get("user_type")
    if user_type in ("field_agent", "agent_terrain"):
        coop_id = current_user.get("cooperative_id", "")
    elif user_type in ("cooperative",):
        coop_id = str(current_user["_id"])
    elif user_type in ("admin", "super_admin"):
        coop_id = str(current_user["_id"])
    else:
        raise HTTPException(status_code=403, detail="Accès réservé aux coopératives ou agents terrain")
    
    # Support both field names and types for backward compatibility
    query = coop_id_query(coop_id)
    if status_filter:
        query["status"] = status_filter
    if village:
        query["village"] = village
    if search:
        query["$and"] = [
            {"$or": [{"coop_id": coop_id}, {"cooperative_id": coop_id}]},
            {"$or": [
                {"full_name": {"$regex": search, "$options": "i"}},
                {"phone_number": {"$regex": search, "$options": "i"}}
            ]}
        ]
        del query["$or"]
    
    total = await db.coop_members.count_documents(query)
    members = await db.coop_members.find(query).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # Enrich with parcel data
    result = []
    for m in members:
        member_data = {
            "id": str(m["_id"]),
            "full_name": m.get("full_name", ""),
            "phone_number": m.get("phone_number", ""),
            "village": m.get("village", ""),
            "cni_number": m.get("cni_number", ""),
            "status": m.get("status", "active"),
            "is_active": m.get("is_active", True),
            "code_planteur": m.get("code_planteur", ""),
            "pin_configured": bool(m.get("pin_hash")),
            "created_at": m.get("created_at", datetime.now(timezone.utc)).isoformat() if isinstance(m.get("created_at"), datetime) else str(m.get("created_at", ""))
        }
        
        # Get member's parcels by member_id AND user_id
        member_id_str = str(m["_id"])
        parcel_or = [{"member_id": member_id_str}, {"farmer_id": member_id_str}]
        if m.get("user_id"):
            parcel_or.append({"farmer_id": m["user_id"]})
        parcels = await db.parcels.find({"$or": parcel_or}).to_list(100)
        member_data["nombre_parcelles"] = len(parcels)
        member_data["superficie_totale"] = round(sum([p.get("area_hectares", 0) or 0 for p in parcels]), 2)
        member_data["score_carbone_moyen"] = round(
            sum([p.get("carbon_score", 0) or 0 for p in parcels]) / len(parcels), 1
        ) if parcels else 0
        
        result.append(member_data)
    
    return {
        "total": total,
        "members": result
    }

@router.post("/members")
@limiter.limit("20/minute")
async def create_coop_member(
    request: Request,
    member: CoopMemberCreate,
    current_user: dict = Depends(get_current_user)
):
    """Ajouter un nouveau membre à la coopérative avec code planteur et PIN USSD"""
    verify_cooperative(current_user)
    coop_id = str(current_user["_id"])
    
    # Check if member already exists (support both string and ObjectId)
    existing = await db.coop_members.find_one({
        **coop_id_query(coop_id),
        "phone_number": member.phone_number
    })
    if existing:
        raise HTTPException(status_code=400, detail="Ce membre existe déjà dans la coopérative")
    
    # Validate PIN code - OBLIGATOIRE
    if not member.pin_code or len(member.pin_code) != 4 or not member.pin_code.isdigit():
        raise HTTPException(status_code=400, detail="Le code PIN à 4 chiffres est obligatoire")
    pin_hash = hash_pin(member.pin_code)
    
    # Auto-generate Code Planteur
    coop_code = current_user.get("coop_code", "")
    farmer_code = await generate_farmer_code(coop_code, member.village)
    
    member_doc = {
        "coop_id": coop_id,
        "full_name": member.full_name,
        "phone_number": member.phone_number,
        "village": member.village,
        "department": member.department,
        "zone": member.zone,
        "cni_number": member.cni_number,
        "consent_given": member.consent_given,
        "consent_date": datetime.now(timezone.utc) if member.consent_given else None,
        "hectares_approx": member.hectares,
        "status": "pending_validation",
        "is_active": True,
        "user_id": None,
        "code_planteur": farmer_code,
        "pin_hash": pin_hash,
        "created_at": datetime.now(timezone.utc),
        "created_by": current_user["_id"]
    }
    
    result = await db.coop_members.insert_one(member_doc)
    
    # Also create a ussd_registrations entry so USSD recognizes this farmer
    ussd_reg = {
        "full_name": member.full_name,
        "nom_complet": member.full_name,
        "phone_number": member.phone_number,
        "coop_code": coop_code,
        "cooperative_code": coop_code,
        "code_planteur": farmer_code,
        "village": member.village,
        "pin_hash": pin_hash,
        "hectares_approx": member.hectares,
        "user_type": "producteur",
        "registered_via": "cooperative_dashboard",
        "status": "active",
        "coop_member_id": str(result.inserted_id),
        "created_at": datetime.now(timezone.utc)
    }
    await db.ussd_registrations.insert_one(ussd_reg)
    
    logger.info(f"New coop member created: {member.full_name} for coop {coop_id}, code: {farmer_code}")
    
    return {
        "message": "Membre ajouté avec succès",
        "member_id": str(result.inserted_id),
        "code_planteur": farmer_code,
        "pin_configured": True
    }

@router.post("/members/import-csv")
async def import_members_csv(
    members_data: List[CoopMemberCreate],
    current_user: dict = Depends(get_current_user)
):
    """Import massif de membres via CSV"""
    verify_cooperative(current_user)
    coop_id = str(current_user["_id"])
    coop_code = current_user.get("coop_code", "")
    
    imported = 0
    errors = []
    
    for idx, member in enumerate(members_data):
        try:
            existing = await db.coop_members.find_one({
                **coop_id_query(coop_id),
                "phone_number": member.phone_number
            })
            if existing:
                errors.append(f"Ligne {idx+1}: {member.phone_number} existe déjà")
                continue
            
            # Auto-generate code planteur
            farmer_code = await generate_farmer_code(coop_code, member.village)
            
            # Hash PIN if provided
            pin_hash = None
            if member.pin_code and len(member.pin_code) == 4 and member.pin_code.isdigit():
                pin_hash = hash_pin(member.pin_code)
            
            member_doc = {
                "coop_id": coop_id,
                "full_name": member.full_name,
                "phone_number": member.phone_number,
                "village": member.village,
                "cni_number": member.cni_number,
                "consent_given": member.consent_given,
                "code_planteur": farmer_code,
                "pin_hash": pin_hash,
                "status": "pending_validation",
                "is_active": True,
                "created_at": datetime.now(timezone.utc),
                "created_by": current_user["_id"],
                "import_batch": True
            }
            
            result = await db.coop_members.insert_one(member_doc)
            
            # Create USSD registration entry
            await db.ussd_registrations.insert_one({
                "full_name": member.full_name,
                "nom_complet": member.full_name,
                "phone_number": member.phone_number,
                "coop_code": coop_code,
                "code_planteur": farmer_code,
                "village": member.village,
                "pin_hash": pin_hash,
                "user_type": "producteur",
                "registered_via": "csv_import",
                "status": "active",
                "coop_member_id": str(result.inserted_id),
                "created_at": datetime.now(timezone.utc)
            })
            
            imported += 1
            
        except Exception as e:
            errors.append(f"Ligne {idx+1}: {str(e)}")
    
    return {
        "message": f"{imported} membres importés avec succès",
        "imported": imported,
        "errors": errors[:10],
        "total_errors": len(errors)
    }


@router.get("/members/activation-stats")
async def get_activation_stats(
    current_user: dict = Depends(get_current_user)
):
    """Statistiques d'activation des membres de la coopérative"""
    user_type = current_user.get("user_type")
    if user_type in ("field_agent", "agent_terrain"):
        coop_id = current_user.get("cooperative_id", "")
    elif user_type in ("cooperative",):
        coop_id = str(current_user["_id"])
    elif user_type in ("admin", "super_admin"):
        coop_id = str(current_user["_id"])
    else:
        raise HTTPException(status_code=403, detail="Accès réservé")

    query = coop_id_query(coop_id)
    all_members = await db.coop_members.find(query).to_list(500)

    total = len(all_members)
    activated = [m for m in all_members if m.get("account_activated") or m.get("user_id")]
    pending = [m for m in all_members if not m.get("account_activated") and not m.get("user_id")]
    pin_configured = [m for m in all_members if m.get("pin_hash")]
    has_code = [m for m in all_members if m.get("code_planteur")]

    activation_rate = round((len(activated) / total * 100), 1) if total > 0 else 0

    recent = sorted(activated, key=lambda m: m.get("activation_date") or m.get("created_at") or datetime.min, reverse=True)[:5]
    recent_list = [{
        "id": str(m["_id"]),
        "full_name": m.get("full_name", ""),
        "phone_number": m.get("phone_number", ""),
        "village": m.get("village", ""),
        "activation_date": (m.get("activation_date") or m.get("created_at", "")).isoformat() if isinstance(m.get("activation_date") or m.get("created_at"), datetime) else str(m.get("activation_date") or m.get("created_at", ""))
    } for m in recent]

    pending_list = [{
        "id": str(m["_id"]),
        "full_name": m.get("full_name", ""),
        "phone_number": m.get("phone_number", ""),
        "village": m.get("village", ""),
        "code_planteur": m.get("code_planteur", ""),
        "pin_configured": bool(m.get("pin_hash")),
        "created_at": m.get("created_at", "").isoformat() if isinstance(m.get("created_at"), datetime) else str(m.get("created_at", ""))
    } for m in pending]

    return {
        "total_members": total,
        "activated_count": len(activated),
        "pending_count": len(pending),
        "activation_rate": activation_rate,
        "pin_configured_count": len(pin_configured),
        "pin_missing_count": total - len(pin_configured),
        "code_planteur_count": len(has_code),
        "recent_activations": recent_list,
        "pending_activation": pending_list
    }


@router.get("/members/export")
async def export_members(
    format: str = Query("xlsx", regex="^(xlsx|pdf)$"),
    status: Optional[str] = None,
    search: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Exporter la liste des membres en Excel ou PDF"""
    verify_cooperative(current_user)
    coop_id = str(current_user["_id"])
    coop_name = current_user.get("coop_name") or current_user.get("full_name", "Coopérative")
    coop_code = current_user.get("coop_code", "")

    query = coop_id_query(coop_id)
    if status:
        query["status"] = status
    if search:
        query["$or"] = [
            {"full_name": {"$regex": search, "$options": "i"}},
            {"phone_number": {"$regex": search}},
            {"code_planteur": {"$regex": search, "$options": "i"}}
        ]

    members = await db.coop_members.find(query).sort("created_at", -1).to_list(1000)

    rows = []
    for m in members:
        activated = bool(m.get("account_activated") or m.get("user_id"))
        rows.append({
            "Nom complet": m.get("full_name", ""),
            "Téléphone": m.get("phone_number", ""),
            "Village": m.get("village", ""),
            "Département": m.get("department", ""),
            "Code Planteur": m.get("code_planteur", ""),
            "Statut": m.get("status", ""),
            "Compte activé": "Oui" if activated else "Non",
            "PIN USSD": "Oui" if m.get("pin_hash") else "Non",
            "Hectares": m.get("hectares_approx", ""),
            "CNI": m.get("cni_number", ""),
            "Date création": m.get("created_at").strftime("%d/%m/%Y") if isinstance(m.get("created_at"), datetime) else str(m.get("created_at", ""))
        })

    if format == "xlsx":
        return _export_xlsx(rows, coop_name, coop_code)
    else:
        return _export_pdf(rows, coop_name, coop_code)


def _export_xlsx(rows, coop_name, coop_code):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Membres"

    # Header row
    green_fill = PatternFill(start_color="2D5A4D", end_color="2D5A4D", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = list(rows[0].keys()) if rows else ["Nom complet", "Téléphone", "Village", "Code Planteur", "Statut", "Compte activé", "PIN USSD", "Hectares", "Date création"]

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=f"{coop_name} ({coop_code}) - Liste des Membres")
    title_cell.font = Font(bold=True, size=14, color="2D5A4D")
    title_cell.alignment = Alignment(horizontal="center")

    # Date row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    date_cell = ws.cell(row=2, column=1, value=f"Exporté le {datetime.now(timezone.utc).strftime('%d/%m/%Y à %H:%M')}")
    date_cell.font = Font(italic=True, size=10, color="666666")
    date_cell.alignment = Alignment(horizontal="center")

    # Column headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = white_font
        cell.fill = green_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data rows
    for row_idx, row_data in enumerate(rows, 5):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left")

    # Auto-width columns
    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        for row_data in rows:
            val = str(row_data.get(header, ""))
            max_len = max(max_len, len(val))
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = min(max_len + 4, 35)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"membres_{coop_code}_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def _export_pdf(rows, coop_name, coop_code):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=10*mm, rightMargin=10*mm, topMargin=15*mm, bottomMargin=15*mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title2", parent=styles["Title"], fontSize=16, textColor=colors.HexColor("#2D5A4D"))
    subtitle_style = ParagraphStyle("Sub", parent=styles["Normal"], fontSize=9, textColor=colors.grey)

    elements = []
    elements.append(Paragraph(f"{coop_name} ({coop_code})", title_style))
    elements.append(Paragraph(f"Liste des Membres - Exporté le {datetime.now(timezone.utc).strftime('%d/%m/%Y à %H:%M')}", subtitle_style))
    elements.append(Spacer(1, 8*mm))

    # Table headers (simplified for PDF)
    pdf_headers = ["Nom", "Téléphone", "Village", "Code Planteur", "Activé", "PIN", "Hectares", "Date"]
    pdf_keys = ["Nom complet", "Téléphone", "Village", "Code Planteur", "Compte activé", "PIN USSD", "Hectares", "Date création"]

    table_data = [pdf_headers]
    for row in rows:
        table_data.append([str(row.get(k, ""))[:25] for k in pdf_keys])

    if len(table_data) == 1:
        table_data.append(["Aucun membre"] + [""] * (len(pdf_headers) - 1))

    col_widths = [55*mm, 40*mm, 35*mm, 35*mm, 20*mm, 15*mm, 22*mm, 28*mm]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2D5A4D")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F7F4")]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 5*mm))
    elements.append(Paragraph(f"Total: {len(rows)} membre(s)", subtitle_style))

    doc.build(elements)
    output.seek(0)

    filename = f"membres_{coop_code}_{datetime.now(timezone.utc).strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/members/{member_id}/send-reminder")
async def send_activation_reminder(
    member_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Envoyer un rappel SMS d'activation à un membre (MOCKED)"""
    verify_cooperative(current_user)
    coop_id = str(current_user["_id"])

    member = await db.coop_members.find_one({
        "_id": ObjectId(member_id),
        **coop_id_query(coop_id)
    })

    if not member:
        raise HTTPException(status_code=404, detail="Membre non trouvé")

    if member.get("account_activated") or member.get("user_id"):
        raise HTTPException(status_code=400, detail="Ce membre a déjà activé son compte")

    coop_name = current_user.get("coop_name") or current_user.get("full_name", "Coopérative")
    member_phone = member.get("phone_number", "")
    member_name = member.get("full_name", "")

    logger.info(f"[SMS MOCK] Rappel activation envoyé à {member_name} ({member_phone}): "
                f"Bonjour {member_name}, votre coopérative {coop_name} vous invite à activer "
                f"votre compte GreenLink. Composez *144*99# ou rendez-vous sur l'application.")

    await db.coop_members.update_one(
        {"_id": ObjectId(member_id)},
        {"$set": {"last_reminder_sent": datetime.now(timezone.utc)},
         "$inc": {"reminder_count": 1}}
    )

    return {
        "message": f"Rappel SMS envoyé à {member_name} ({member_phone})",
        "sms_sent": False,
        "sms_mocked": True,
        "note": "SMS simulé - passerelle Orange non configurée"
    }


@router.put("/members/{member_id}/validate")
async def validate_member(
    member_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Valider un membre en attente"""
    verify_cooperative(current_user)
    
    coop_id = str(current_user["_id"])
    result = await db.coop_members.update_one(
        {"_id": ObjectId(member_id), **coop_id_query(coop_id)},
        {
            "$set": {
                "status": "active",
                "validated_at": datetime.utcnow(),
                "validated_by": current_user["_id"]
            }
        }
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Membre non trouvé")
    
    return {"message": "Membre validé avec succès"}

@router.get("/members/{member_id}")
async def get_member_details(
    member_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Détails d'un membre"""
    verify_cooperative(current_user)
    
    coop_id = str(current_user["_id"])
    member = await db.coop_members.find_one({
        "_id": ObjectId(member_id),
        **coop_id_query(coop_id)
    })
    
    if not member:
        raise HTTPException(status_code=404, detail="Membre non trouvé")
    
    # Get member's parcels and harvests
    parcels = []
    harvests = []
    total_premium = 0
    
    if member.get("user_id"):
        parcels = await db.parcels.find({"$or": [{"farmer_id": member["user_id"]}, {"member_id": str(member["_id"])}, {"farmer_id": str(member["_id"])}]}).to_list(100)
        harvests = await db.harvests.find({"farmer_id": member["user_id"]}).to_list(100)
        total_premium = sum([h.get("carbon_premium", 0) for h in harvests])
    else:
        mid = str(member["_id"])
        parcels = await db.parcels.find({"$or": [{"farmer_id": mid}, {"member_id": mid}]}).to_list(100)
    
    return {
        "id": str(member["_id"]),
        "full_name": member.get("full_name", ""),
        "phone_number": member.get("phone_number", ""),
        "village": member.get("village", ""),
        "cni_number": member.get("cni_number", ""),
        "status": member.get("status", ""),
        "is_active": member.get("is_active", True),
        "consent_given": member.get("consent_given", False),
        "code_planteur": member.get("code_planteur", ""),
        "pin_configured": bool(member.get("pin_hash")),
        "created_at": member.get("created_at", ""),
        "parcels": [{
            "id": str(p["_id"]),
            "location": p.get("location", ""),
            "area_hectares": p.get("area_hectares", 0),
            "carbon_score": p.get("carbon_score", 0),
            "crop_type": p.get("crop_type", "cacao")
        } for p in parcels],
        "harvests_count": len(harvests),
        "total_premium_earned": total_premium
    }
