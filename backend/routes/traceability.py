"""
Module Tracabilite ARS 1000-2 (Clauses 11-16)
GreenLink Agritech - Cote d'Ivoire

Tracabilite complete du cacao: de la parcelle a l'export.
Segregation physique, QR codes, rapports d'audit.
"""

from fastapi import APIRouter, HTTPException, Depends, Query, status
from fastapi.responses import StreamingResponse
from typing import Optional, List
from pydantic import BaseModel, Field
from datetime import datetime, timezone
from bson import ObjectId
import logging
import io
import json
import base64
import uuid

from database import db
from routes.auth import get_current_user
from routes.cooperative import verify_cooperative

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/traceability", tags=["Tracabilite ARS 1000-2"])

# ============= CONSTANTS =============

ETAPES_FLUX = [
    "recolte",
    "fermentation",
    "sechage",
    "stockage_coop",
    "conditionnement",
    "transport",
    "export",
]

ETAPE_LABELS = {
    "recolte": "Recolte",
    "fermentation": "Fermentation",
    "sechage": "Sechage",
    "stockage_coop": "Stockage Cooperative",
    "conditionnement": "Conditionnement",
    "transport": "Transport",
    "export": "Export",
}

STATUTS_ARS = ["conforme", "non_conforme", "en_attente", "en_cours"]

# ============= PYDANTIC MODELS =============

class LotSourceCreate(BaseModel):
    farmer_id: str
    farmer_name: str = ""
    parcelle_id: str = ""
    parcelle_name: str = ""
    quantite_kg: float = 0
    date_recolte: str = ""
    campagne: str = "2025-2026"
    grade_qualite: str = "en_attente"
    certifie_ars1000: bool = False
    notes: str = ""

class LotSourceUpdate(BaseModel):
    quantite_kg: Optional[float] = None
    grade_qualite: Optional[str] = None
    certifie_ars1000: Optional[bool] = None
    notes: Optional[str] = None

class EvenementCreate(BaseModel):
    etape: str
    date_evenement: str = ""
    quantite_kg: float = 0
    lieu: str = ""
    responsable: str = ""
    temperature: Optional[float] = None
    humidite: Optional[float] = None
    duree_heures: Optional[float] = None
    observations: str = ""
    controle_qualite: dict = {}
    conforme: bool = True

class SegregationCheck(BaseModel):
    lot_ids: List[str]
    action: str = "melange"

class ObjectifARS(BaseModel):
    clause: str
    titre: str
    description: str = ""
    cible: float = 100
    unite: str = "%"

# ============= HELPERS =============

def serialize_doc(doc):
    """Remove _id and convert ObjectId fields to strings"""
    if doc is None:
        return None
    result = {}
    for k, v in doc.items():
        if k == "_id":
            continue
        if isinstance(v, ObjectId):
            result[k] = str(v)
        elif isinstance(v, datetime):
            result[k] = v.isoformat()
        elif isinstance(v, list):
            result[k] = [serialize_doc(i) if isinstance(i, dict) else (str(i) if isinstance(i, ObjectId) else i) for i in v]
        elif isinstance(v, dict):
            result[k] = serialize_doc(v)
        else:
            result[k] = v
    return result

def get_coop_id(user):
    return str(user.get("coop_id") or user.get("_id") or user.get("id", ""))

# ============= LOT ENDPOINTS =============

@router.post("/lots")
async def create_lot(lot: LotSourceCreate, current_user: dict = Depends(get_current_user)):
    """Creer un lot a la source (enregistrement initial lie au producteur/parcelle)"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    lot_code = f"LOT-{uuid.uuid4().hex[:8].upper()}"

    doc = {
        "lot_code": lot_code,
        "coop_id": coop_id,
        "farmer_id": lot.farmer_id,
        "farmer_name": lot.farmer_name,
        "parcelle_id": lot.parcelle_id,
        "parcelle_name": lot.parcelle_name,
        "quantite_initiale_kg": lot.quantite_kg,
        "quantite_actuelle_kg": lot.quantite_kg,
        "date_recolte": lot.date_recolte,
        "campagne": lot.campagne,
        "grade_qualite": lot.grade_qualite,
        "certifie_ars1000": lot.certifie_ars1000,
        "notes": lot.notes,
        "etape_courante": "recolte",
        "statut_ars": "en_cours",
        "segregation": "certifie" if lot.certifie_ars1000 else "non_certifie",
        "evenements": [],
        "qr_code": "",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    result = await db.traceability_lots.insert_one(doc)
    doc.pop("_id", None)

    return {"status": "success", "lot": doc, "lot_code": lot_code}


@router.get("/lots")
async def list_lots(
    current_user: dict = Depends(get_current_user),
    campagne: Optional[str] = None,
    etape: Optional[str] = None,
    segregation: Optional[str] = None,
    certifie: Optional[str] = None,
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
):
    """Lister les lots de la cooperative avec filtres"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    query = {"coop_id": coop_id}
    if campagne:
        query["campagne"] = campagne
    if etape:
        query["etape_courante"] = etape
    if segregation:
        query["segregation"] = segregation
    if certifie == "true":
        query["certifie_ars1000"] = True
    elif certifie == "false":
        query["certifie_ars1000"] = False
    if search:
        query["$or"] = [
            {"lot_code": {"$regex": search, "$options": "i"}},
            {"farmer_name": {"$regex": search, "$options": "i"}},
            {"parcelle_name": {"$regex": search, "$options": "i"}},
        ]

    total = await db.traceability_lots.count_documents(query)
    lots = await db.traceability_lots.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)

    return {"lots": lots, "total": total, "skip": skip, "limit": limit}


@router.get("/lots/{lot_code}")
async def get_lot_detail(lot_code: str, current_user: dict = Depends(get_current_user)):
    """Detail d'un lot avec sa timeline complete"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    lot = await db.traceability_lots.find_one(
        {"lot_code": lot_code, "coop_id": coop_id}, {"_id": 0}
    )
    if not lot:
        raise HTTPException(status_code=404, detail="Lot non trouve")

    return {"lot": lot}


@router.put("/lots/{lot_code}")
async def update_lot(lot_code: str, update: LotSourceUpdate, current_user: dict = Depends(get_current_user)):
    """Mettre a jour un lot"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()

    if "certifie_ars1000" in update_data:
        update_data["segregation"] = "certifie" if update_data["certifie_ars1000"] else "non_certifie"

    result = await db.traceability_lots.update_one(
        {"lot_code": lot_code, "coop_id": coop_id},
        {"$set": update_data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Lot non trouve")

    lot = await db.traceability_lots.find_one(
        {"lot_code": lot_code, "coop_id": coop_id}, {"_id": 0}
    )
    return {"status": "success", "lot": lot}


# ============= EVENEMENTS (FLUX) =============

@router.post("/lots/{lot_code}/events")
async def add_event(lot_code: str, event: EvenementCreate, current_user: dict = Depends(get_current_user)):
    """Ajouter un evenement (etape du flux) a un lot"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    if event.etape not in ETAPES_FLUX:
        raise HTTPException(status_code=400, detail=f"Etape invalide. Valeurs: {', '.join(ETAPES_FLUX)}")

    lot = await db.traceability_lots.find_one({"lot_code": lot_code, "coop_id": coop_id})
    if not lot:
        raise HTTPException(status_code=404, detail="Lot non trouve")

    event_doc = {
        "event_id": str(uuid.uuid4()),
        "etape": event.etape,
        "etape_label": ETAPE_LABELS.get(event.etape, event.etape),
        "date_evenement": event.date_evenement or datetime.now(timezone.utc).isoformat(),
        "quantite_kg": event.quantite_kg,
        "lieu": event.lieu,
        "responsable": event.responsable,
        "temperature": event.temperature,
        "humidite": event.humidite,
        "duree_heures": event.duree_heures,
        "observations": event.observations,
        "controle_qualite": event.controle_qualite,
        "conforme": event.conforme,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }

    update_fields = {
        "etape_courante": event.etape,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    if event.quantite_kg > 0:
        update_fields["quantite_actuelle_kg"] = event.quantite_kg

    if not event.conforme:
        update_fields["statut_ars"] = "non_conforme"

    await db.traceability_lots.update_one(
        {"lot_code": lot_code, "coop_id": coop_id},
        {"$push": {"evenements": event_doc}, "$set": update_fields}
    )

    return {"status": "success", "event": event_doc}


@router.get("/lots/{lot_code}/timeline")
async def get_lot_timeline(lot_code: str, current_user: dict = Depends(get_current_user)):
    """Timeline complete du lot (tracabilite ascendante/descendante)"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    lot = await db.traceability_lots.find_one(
        {"lot_code": lot_code, "coop_id": coop_id}, {"_id": 0}
    )
    if not lot:
        raise HTTPException(status_code=404, detail="Lot non trouve")

    timeline = []
    for etape in ETAPES_FLUX:
        events_for_step = [e for e in lot.get("evenements", []) if e["etape"] == etape]
        timeline.append({
            "etape": etape,
            "label": ETAPE_LABELS.get(etape, etape),
            "completed": len(events_for_step) > 0,
            "current": lot.get("etape_courante") == etape,
            "events": events_for_step,
        })

    return {
        "lot_code": lot_code,
        "farmer_name": lot.get("farmer_name", ""),
        "parcelle_name": lot.get("parcelle_name", ""),
        "quantite_initiale_kg": lot.get("quantite_initiale_kg", 0),
        "quantite_actuelle_kg": lot.get("quantite_actuelle_kg", 0),
        "certifie_ars1000": lot.get("certifie_ars1000", False),
        "segregation": lot.get("segregation", ""),
        "timeline": timeline,
    }


# ============= SEGREGATION =============

@router.get("/segregation")
async def get_segregation_status(current_user: dict = Depends(get_current_user)):
    """Vue des magasins virtuels (certifie vs non-certifie)"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    pipeline_certifie = [
        {"$match": {"coop_id": coop_id, "segregation": "certifie"}},
        {"$group": {
            "_id": "$etape_courante",
            "count": {"$sum": 1},
            "total_kg": {"$sum": "$quantite_actuelle_kg"},
        }},
    ]

    pipeline_non_certifie = [
        {"$match": {"coop_id": coop_id, "segregation": "non_certifie"}},
        {"$group": {
            "_id": "$etape_courante",
            "count": {"$sum": 1},
            "total_kg": {"$sum": "$quantite_actuelle_kg"},
        }},
    ]

    certifie_raw = await db.traceability_lots.aggregate(pipeline_certifie).to_list(100)
    non_certifie_raw = await db.traceability_lots.aggregate(pipeline_non_certifie).to_list(100)

    def format_warehouse(raw_data, label):
        items = []
        total_kg = 0
        total_lots = 0
        for r in raw_data:
            items.append({
                "etape": r["_id"],
                "label": ETAPE_LABELS.get(r["_id"], r["_id"]),
                "lots": r["count"],
                "quantite_kg": r["total_kg"],
            })
            total_kg += r["total_kg"]
            total_lots += r["count"]
        return {
            "label": label,
            "items": items,
            "total_kg": total_kg,
            "total_lots": total_lots,
        }

    return {
        "certifie": format_warehouse(certifie_raw, "Magasin Certifie ARS 1000"),
        "non_certifie": format_warehouse(non_certifie_raw, "Magasin Non-Certifie"),
    }


@router.post("/segregation/check")
async def check_segregation(data: SegregationCheck, current_user: dict = Depends(get_current_user)):
    """Verifier si un melange de lots est autorise (segregation)"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    lots = await db.traceability_lots.find(
        {"lot_code": {"$in": data.lot_ids}, "coop_id": coop_id}, {"_id": 0}
    ).to_list(100)

    if len(lots) < 2:
        return {"allowed": True, "message": "Un seul lot, pas de risque de melange."}

    segregations = set(l.get("segregation", "") for l in lots)

    if len(segregations) > 1:
        return {
            "allowed": False,
            "alerte": "BLOCAGE SEGREGATION",
            "message": "Melange interdit: lots certifies et non-certifies detectes. La segregation ARS 1000 interdit ce melange.",
            "lots_certifies": [l["lot_code"] for l in lots if l.get("segregation") == "certifie"],
            "lots_non_certifies": [l["lot_code"] for l in lots if l.get("segregation") == "non_certifie"],
        }

    return {
        "allowed": True,
        "message": f"Melange autorise: tous les lots sont de type '{list(segregations)[0]}'.",
    }


# ============= QR CODE =============

@router.get("/lots/{lot_code}/qrcode")
async def generate_qr_code(lot_code: str, current_user: dict = Depends(get_current_user)):
    """Generer un QR code pour un lot (tracabilite sur sac)"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    lot = await db.traceability_lots.find_one(
        {"lot_code": lot_code, "coop_id": coop_id}, {"_id": 0}
    )
    if not lot:
        raise HTTPException(status_code=404, detail="Lot non trouve")

    import qrcode
    qr_data = json.dumps({
        "lot_code": lot_code,
        "coop_id": coop_id,
        "farmer": lot.get("farmer_name", ""),
        "parcelle": lot.get("parcelle_name", ""),
        "quantite_kg": lot.get("quantite_actuelle_kg", 0),
        "certifie": lot.get("certifie_ars1000", False),
        "campagne": lot.get("campagne", ""),
        "etape": lot.get("etape_courante", ""),
    }, ensure_ascii=False)

    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=10, border=4)
    qr.add_data(qr_data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#1A3622", back_color="white")

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)

    return StreamingResponse(buf, media_type="image/png", headers={
        "Content-Disposition": f"inline; filename=qr_{lot_code}.png"
    })


@router.get("/lots/{lot_code}/qrcode-base64")
async def get_qr_code_base64(lot_code: str, current_user: dict = Depends(get_current_user)):
    """QR code en base64 (pour integration frontend)"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    lot = await db.traceability_lots.find_one(
        {"lot_code": lot_code, "coop_id": coop_id}, {"_id": 0}
    )
    if not lot:
        raise HTTPException(status_code=404, detail="Lot non trouve")

    import qrcode
    qr_data = json.dumps({
        "lot": lot_code,
        "certifie": lot.get("certifie_ars1000", False),
        "kg": lot.get("quantite_actuelle_kg", 0),
        "etape": lot.get("etape_courante", ""),
    }, ensure_ascii=False)

    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=8, border=3)
    qr.add_data(qr_data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#1A3622", back_color="white")

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    b64 = base64.b64encode(buf.getvalue()).decode("utf-8")

    return {"qr_base64": f"data:image/png;base64,{b64}", "lot_code": lot_code}


# ============= DASHBOARD =============

@router.get("/dashboard")
async def get_traceability_dashboard(current_user: dict = Depends(get_current_user)):
    """Tableau de bord de la tracabilite"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    base_q = {"coop_id": coop_id}

    total_lots = await db.traceability_lots.count_documents(base_q)
    lots_certifies = await db.traceability_lots.count_documents({**base_q, "certifie_ars1000": True})
    lots_non_certifies = total_lots - lots_certifies
    lots_conformes = await db.traceability_lots.count_documents({**base_q, "statut_ars": "conforme"})
    lots_non_conformes = await db.traceability_lots.count_documents({**base_q, "statut_ars": "non_conforme"})
    lots_en_cours = await db.traceability_lots.count_documents({**base_q, "statut_ars": "en_cours"})

    # Volume total
    volume_pipeline = [
        {"$match": base_q},
        {"$group": {"_id": None, "total": {"$sum": "$quantite_actuelle_kg"}}},
    ]
    volume_result = await db.traceability_lots.aggregate(volume_pipeline).to_list(1)
    volume_total_kg = volume_result[0]["total"] if volume_result else 0

    # Repartition par etape
    etape_pipeline = [
        {"$match": base_q},
        {"$group": {
            "_id": "$etape_courante",
            "count": {"$sum": 1},
            "total_kg": {"$sum": "$quantite_actuelle_kg"},
        }},
    ]
    etape_result = await db.traceability_lots.aggregate(etape_pipeline).to_list(20)
    par_etape = []
    for e in ETAPES_FLUX:
        found = next((r for r in etape_result if r["_id"] == e), None)
        par_etape.append({
            "etape": e,
            "label": ETAPE_LABELS.get(e, e),
            "lots": found["count"] if found else 0,
            "quantite_kg": found["total_kg"] if found else 0,
        })

    # Derniers evenements
    recent_pipeline = [
        {"$match": base_q},
        {"$unwind": "$evenements"},
        {"$sort": {"evenements.created_at": -1}},
        {"$limit": 10},
        {"$project": {
            "_id": 0,
            "lot_code": 1,
            "farmer_name": 1,
            "event": "$evenements",
        }},
    ]
    recent_events = await db.traceability_lots.aggregate(recent_pipeline).to_list(10)

    # Alertes segregation
    alertes = []
    non_conforme_lots = await db.traceability_lots.find(
        {**base_q, "statut_ars": "non_conforme"}, {"_id": 0, "lot_code": 1, "farmer_name": 1, "etape_courante": 1}
    ).to_list(20)
    for nc in non_conforme_lots:
        alertes.append({
            "type": "non_conforme",
            "severity": "error",
            "message": f"Lot {nc['lot_code']} ({nc.get('farmer_name', '')}) - Non conforme a l'etape {ETAPE_LABELS.get(nc.get('etape_courante', ''), '')}",
        })

    return {
        "kpis": {
            "total_lots": total_lots,
            "lots_certifies": lots_certifies,
            "lots_non_certifies": lots_non_certifies,
            "lots_conformes": lots_conformes,
            "lots_non_conformes": lots_non_conformes,
            "lots_en_cours": lots_en_cours,
            "volume_total_kg": volume_total_kg,
            "taux_conformite": round((lots_conformes / total_lots * 100) if total_lots > 0 else 0, 1),
            "taux_certification": round((lots_certifies / total_lots * 100) if total_lots > 0 else 0, 1),
        },
        "par_etape": par_etape,
        "recent_events": recent_events,
        "alertes": alertes,
    }


# ============= OBJECTIFS ARS 1000 =============

@router.get("/objectives")
async def get_objectives(current_user: dict = Depends(get_current_user)):
    """Objectifs de tracabilite ARS 1000-2 (Clauses 11-16)"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    base_q = {"coop_id": coop_id}
    total = await db.traceability_lots.count_documents(base_q)
    certifies = await db.traceability_lots.count_documents({**base_q, "certifie_ars1000": True})
    conformes = await db.traceability_lots.count_documents({**base_q, "statut_ars": "conforme"})

    # Count lots with at least 3 events (good traceability depth)
    deep_trace = await db.traceability_lots.count_documents({
        **base_q, "evenements.2": {"$exists": True}
    })

    # Count lots that reached export
    exported = await db.traceability_lots.count_documents({
        **base_q, "etape_courante": "export"
    })

    objectives = [
        {
            "clause": "11",
            "titre": "Enregistrement a la source",
            "description": "Chaque lot doit etre enregistre avec producteur, parcelle, quantite et date.",
            "valeur": total,
            "cible": max(total, 10),
            "unite": "lots",
            "progression": min(100, round(total / max(total, 10) * 100)) if total > 0 else 0,
            "statut": "conforme" if total > 0 else "en_attente",
        },
        {
            "clause": "12",
            "titre": "Segregation physique",
            "description": "Separation stricte entre lots certifies et non-certifies. Aucun melange autorise.",
            "valeur": certifies,
            "cible": total,
            "unite": "lots certifies",
            "progression": round(certifies / total * 100) if total > 0 else 0,
            "statut": "conforme" if certifies > 0 else "en_attente",
        },
        {
            "clause": "13",
            "titre": "Tracabilite etape par etape",
            "description": "Suivi du lot de la recolte a l'export avec dates, quantites et controles.",
            "valeur": deep_trace,
            "cible": total,
            "unite": "lots traces",
            "progression": round(deep_trace / total * 100) if total > 0 else 0,
            "statut": "conforme" if deep_trace >= total * 0.8 else "en_cours" if deep_trace > 0 else "en_attente",
        },
        {
            "clause": "14",
            "titre": "Controles qualite",
            "description": "Chaque etape doit inclure un controle qualite conforme ARS 1000.",
            "valeur": conformes,
            "cible": total,
            "unite": "lots conformes",
            "progression": round(conformes / total * 100) if total > 0 else 0,
            "statut": "conforme" if conformes >= total * 0.9 else "en_cours" if conformes > 0 else "en_attente",
        },
        {
            "clause": "15",
            "titre": "Tracabilite ascendante/descendante",
            "description": "QR codes sur sacs pour remonter ou descendre la chaine.",
            "valeur": total,
            "cible": total,
            "unite": "lots avec QR",
            "progression": 100 if total > 0 else 0,
            "statut": "conforme" if total > 0 else "en_attente",
        },
        {
            "clause": "16",
            "titre": "Rapport d'audit",
            "description": "Rapports PDF/Excel generables pour les auditeurs ARS 1000.",
            "valeur": exported,
            "cible": total,
            "unite": "lots exportes",
            "progression": round(exported / total * 100) if total > 0 else 0,
            "statut": "conforme" if exported > 0 else "en_attente",
        },
    ]

    score_global = round(sum(o["progression"] for o in objectives) / len(objectives)) if objectives else 0

    return {
        "objectives": objectives,
        "score_global": score_global,
        "total_lots": total,
    }


# ============= RAPPORTS =============

@router.get("/reports/audit-data")
async def get_audit_report_data(
    current_user: dict = Depends(get_current_user),
    campagne: Optional[str] = None,
):
    """Donnees pour le rapport d'audit ARS 1000"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    query = {"coop_id": coop_id}
    if campagne:
        query["campagne"] = campagne

    lots = await db.traceability_lots.find(query, {"_id": 0}).to_list(5000)

    total = len(lots)
    certifies = sum(1 for l in lots if l.get("certifie_ars1000"))
    conformes = sum(1 for l in lots if l.get("statut_ars") == "conforme")
    non_conformes = sum(1 for l in lots if l.get("statut_ars") == "non_conforme")
    volume = sum(l.get("quantite_actuelle_kg", 0) for l in lots)

    by_farmer = {}
    for l in lots:
        fn = l.get("farmer_name", "Inconnu")
        if fn not in by_farmer:
            by_farmer[fn] = {"lots": 0, "volume_kg": 0, "certifie": 0}
        by_farmer[fn]["lots"] += 1
        by_farmer[fn]["volume_kg"] += l.get("quantite_actuelle_kg", 0)
        if l.get("certifie_ars1000"):
            by_farmer[fn]["certifie"] += 1

    farmers_summary = [{"farmer": k, **v} for k, v in by_farmer.items()]

    by_etape = {}
    for l in lots:
        et = l.get("etape_courante", "recolte")
        if et not in by_etape:
            by_etape[et] = {"lots": 0, "volume_kg": 0}
        by_etape[et]["lots"] += 1
        by_etape[et]["volume_kg"] += l.get("quantite_actuelle_kg", 0)

    etapes_summary = [{"etape": k, "label": ETAPE_LABELS.get(k, k), **v} for k, v in by_etape.items()]

    return {
        "resume": {
            "total_lots": total,
            "lots_certifies": certifies,
            "lots_conformes": conformes,
            "lots_non_conformes": non_conformes,
            "volume_total_kg": volume,
            "taux_conformite": round(conformes / total * 100, 1) if total else 0,
        },
        "par_producteur": farmers_summary,
        "par_etape": etapes_summary,
        "lots": lots,
    }


@router.get("/reports/export/excel")
async def export_excel_report(
    current_user: dict = Depends(get_current_user),
    campagne: Optional[str] = None,
):
    """Exporter le rapport de tracabilite en Excel"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    query = {"coop_id": coop_id}
    if campagne:
        query["campagne"] = campagne

    lots = await db.traceability_lots.find(query, {"_id": 0}).to_list(5000)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # Sheet 1: Resume
    ws1 = wb.active
    ws1.title = "Resume Tracabilite"

    header_fill = PatternFill(start_color="1A3622", end_color="1A3622", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws1.append(["RAPPORT DE TRACABILITE ARS 1000-2"])
    ws1.merge_cells("A1:H1")
    ws1["A1"].font = Font(bold=True, size=16, color="1A3622")

    ws1.append([f"Cooperative: {coop_id}", "", "", f"Date: {datetime.now(timezone.utc).strftime('%d/%m/%Y')}"])
    ws1.append([])

    headers = ["Code Lot", "Producteur", "Parcelle", "Quantite (kg)", "Campagne", "Certifie ARS", "Etape", "Statut"]
    ws1.append(headers)
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for lot in lots:
        ws1.append([
            lot.get("lot_code", ""),
            lot.get("farmer_name", ""),
            lot.get("parcelle_name", ""),
            lot.get("quantite_actuelle_kg", 0),
            lot.get("campagne", ""),
            "Oui" if lot.get("certifie_ars1000") else "Non",
            ETAPE_LABELS.get(lot.get("etape_courante", ""), ""),
            lot.get("statut_ars", ""),
        ])

    for col in range(1, 9):
        ws1.column_dimensions[chr(64 + col)].width = 18

    # Sheet 2: Evenements
    ws2 = wb.create_sheet("Evenements")
    evt_headers = ["Code Lot", "Etape", "Date", "Quantite (kg)", "Lieu", "Responsable", "Conforme", "Observations"]
    ws2.append(evt_headers)
    for col, h in enumerate(evt_headers, 1):
        cell = ws2.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for lot in lots:
        for evt in lot.get("evenements", []):
            ws2.append([
                lot.get("lot_code", ""),
                evt.get("etape_label", ""),
                evt.get("date_evenement", ""),
                evt.get("quantite_kg", 0),
                evt.get("lieu", ""),
                evt.get("responsable", ""),
                "Oui" if evt.get("conforme") else "Non",
                evt.get("observations", ""),
            ])

    for col in range(1, 9):
        ws2.column_dimensions[chr(64 + col)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=tracabilite_ars1000_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"},
    )


@router.get("/reports/export/pdf")
async def export_pdf_report(
    current_user: dict = Depends(get_current_user),
    campagne: Optional[str] = None,
):
    """Exporter le rapport de tracabilite en PDF"""
    verify_cooperative(current_user)
    coop_id = get_coop_id(current_user)

    query = {"coop_id": coop_id}
    if campagne:
        query["campagne"] = campagne

    lots = await db.traceability_lots.find(query, {"_id": 0}).to_list(5000)
    total = len(lots)
    certifies = sum(1 for l in lots if l.get("certifie_ars1000"))
    conformes = sum(1 for l in lots if l.get("statut_ars") == "conforme")
    volume = sum(l.get("quantite_actuelle_kg", 0) for l in lots)

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle("Title", parent=styles["Title"], fontSize=18, textColor=colors.HexColor("#1A3622"))
    elements.append(Paragraph("RAPPORT DE TRACABILITE ARS 1000-2", title_style))
    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph(f"Date: {datetime.now(timezone.utc).strftime('%d/%m/%Y')}", styles["Normal"]))
    elements.append(Spacer(1, 0.8*cm))

    # KPIs
    kpi_data = [
        ["Total Lots", str(total)],
        ["Lots Certifies ARS 1000", str(certifies)],
        ["Lots Conformes", str(conformes)],
        ["Volume Total (kg)", f"{volume:,.0f}"],
        ["Taux Conformite", f"{round(conformes/total*100,1) if total else 0}%"],
    ]
    kpi_table = Table(kpi_data, colWidths=[8*cm, 6*cm])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E8F0EA")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#1A3622")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E5E0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 1*cm))

    # Lots Table
    elements.append(Paragraph("Detail des Lots", ParagraphStyle("H2", parent=styles["Heading2"], textColor=colors.HexColor("#1A3622"))))
    elements.append(Spacer(1, 0.3*cm))

    lot_headers = ["Code", "Producteur", "Qte (kg)", "Certifie", "Etape", "Statut"]
    lot_rows = [lot_headers]
    for l in lots[:100]:
        lot_rows.append([
            l.get("lot_code", "")[:12],
            l.get("farmer_name", "")[:20],
            str(l.get("quantite_actuelle_kg", 0)),
            "Oui" if l.get("certifie_ars1000") else "Non",
            ETAPE_LABELS.get(l.get("etape_courante", ""), "")[:15],
            l.get("statut_ars", ""),
        ])

    lot_table = Table(lot_rows, colWidths=[2.5*cm, 4*cm, 2*cm, 2*cm, 3.5*cm, 2.5*cm])
    lot_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A3622")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E5E0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
    ]))
    elements.append(lot_table)

    doc.build(elements)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=tracabilite_ars1000_{datetime.now(timezone.utc).strftime('%Y%m%d')}.pdf"},
    )


# ============= USSD TRACABILITE =============

@router.post("/ussd/trace")
async def ussd_trace_lot(data: dict):
    """Endpoint USSD simule pour tracer un lot par code (ex: *144*99#)"""
    lot_code = data.get("lot_code", "").strip()
    if not lot_code:
        return {"message": "Veuillez entrer un code lot. Ex: LOT-A1B2C3D4"}

    lot = await db.traceability_lots.find_one({"lot_code": lot_code}, {"_id": 0})
    if not lot:
        return {"message": f"Lot {lot_code} introuvable."}

    etape = ETAPE_LABELS.get(lot.get("etape_courante", ""), "Inconnue")
    certif = "Certifie ARS 1000" if lot.get("certifie_ars1000") else "Non certifie"
    qty = lot.get("quantite_actuelle_kg", 0)
    farmer = lot.get("farmer_name", "")

    msg = (
        f"=== TRACABILITE LOT ===\n"
        f"Code: {lot_code}\n"
        f"Producteur: {farmer}\n"
        f"Quantite: {qty} kg\n"
        f"Etape: {etape}\n"
        f"Statut: {certif}\n"
        f"========================"
    )

    return {"message": msg, "lot_code": lot_code}
